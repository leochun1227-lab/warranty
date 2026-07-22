from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import shutil
import subprocess
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DEFAULT_ANALYSIS_TICKET_JS = ROOT / "outputs" / "analysis_ticket_csv.js"
DEFAULT_FAILURE_TIMING_CSV = ROOT / "outputs" / "analysis_ticket_failure_timing.csv"
DEFAULT_PARTS_CLASSIFIED_CSV = ROOT / "outputs" / "parts_classified.csv"
DEFAULT_REPAIRERS_JSON = ROOT / "outputs" / "repairers_2026" / "repairers_2026_data.json"
DEFAULT_OUTPUT = ROOT / "generated_exports" / "ticket_timeline_segments_2025_2026.xlsx"
DEFAULT_SUMMARY_JSON = ROOT / "generated_exports" / "ticket_timeline_summary.json"
DEFAULT_COMPLETION_ANALYTICS_JSON = ROOT / "generated_exports" / "ticket_timeline_completion_analytics_2026.json"
DEFAULT_PRICE_MIX_PPT_SCRIPT = ROOT / "generate_ticket_timeline_price_mix_ppt.mjs"
DEFAULT_START_DATE = date(date.today().year, 1, 1)

APPROVED_STATUS_TEXTS = {
    "sales order approved",
    "dispatch parts",
    "partially picked",
    "repair in progress",
    "repairer invoiced received",
    "repairer invoiced processed",
    "approved claims closed",
    "approved claims closed (closed)",
}
UNAPPROVED_STATUS_TEXTS = {
    "unapproved claims closed",
    "unapproved claims closed (closed)",
}


@dataclass(frozen=True)
class Paths:
    analysis_ticket_js: Path
    failure_timing_csv: Path
    parts_classified_csv: Path
    repairers_json: Path
    output: Path
    summary_json: Path
    completion_analytics_json: Path


def clean(value: Any) -> str:
    return str(value or "").strip()


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip()


def normalize_po(value: Any) -> str:
    text = clean(value)
    if not text or text == "#":
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return text.split(".", 1)[0]
    return text


def normalize_vehicle_key(value: Any) -> str:
    text = clean(value)
    if not text or text == "#":
        return ""
    return text.upper()


def parse_amount(value: Any) -> float | None:
    text = clean(value)
    if not text or text == "#":
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    text = clean(value)
    if not text or text == "#":
        return None
    text = text.replace(" AUSACT", "")
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y%m%d",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_date(value: Any) -> date | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(value)


def iso_or_blank(value: Any) -> str:
    normalized = to_date(value)
    if normalized is not None:
        return normalized.isoformat()
    return ""


def day_diff(later: Any, earlier: Any) -> int | None:
    later_date = to_date(later)
    earlier_date = to_date(earlier)
    if later_date is None or earlier_date is None:
        return None
    return (later_date - earlier_date).days


def mean_or_blank(series: Iterable[float | int | None]) -> float | None:
    values = [float(value) for value in series if value is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def median_or_blank(series: Iterable[float | int | None]) -> float | None:
    values = sorted(float(value) for value in series if value is not None)
    if not values:
        return None
    mid = len(values) // 2
    if len(values) % 2:
        return round(values[mid], 2)
    return round((values[mid - 1] + values[mid]) / 2, 2)


def read_embedded_csv_text(js_path: Path) -> str:
    text = js_path.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'globalThis\.ANALYSIS_TICKET_CSV_TEXT\s*=\s*"(.*)";\s*$', text, re.S)
    if not match:
        raise ValueError(f"Could not find embedded CSV text in {js_path}")
    return match.group(1).encode("utf-8").decode("unicode_escape")


def ticket_claim_scope(ticket_type: str) -> str:
    text = normalize_spaces(ticket_type).lower()
    if "pre delivery" in text:
        return "Pre Delivery"
    if "in field" in text:
        return "In Field"
    return "Other"


def is_unapproved_status(status_text: str) -> bool:
    text = normalize_spaces(status_text).lower()
    return text in UNAPPROVED_STATUS_TEXTS or "unapproved" in text


def is_approved_like_status(status_text: str) -> bool:
    text = normalize_spaces(status_text).lower()
    return text in APPROVED_STATUS_TEXTS or "approved" in text or "repairer invoiced" in text or "repair in progress" in text


def is_van_recalled_status(status_text: str) -> bool:
    return normalize_spaces(status_text).lower() == "van recalled"


def is_approved_closed_status(status_text: str) -> bool:
    return "approved claims closed" in normalize_spaces(status_text).lower()


def should_include_approval_evidence_universe(row: pd.Series) -> bool:
    if is_unapproved_status(clean(row.get("status_text"))):
        return False
    if is_van_recalled_status(clean(row.get("status_text"))):
        return False
    if is_approved_closed_status(clean(row.get("status_text"))) and to_date(row.get("approved_date")) is None:
        return False
    return True


def load_analysis_ticket_base(js_path: Path) -> pd.DataFrame:
    raw_csv = read_embedded_csv_text(js_path)
    reader = csv.reader(io.StringIO(raw_csv))
    next(reader, None)  # header row is position-based and intentionally skipped

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row = list(raw_row) + [""] * max(0, 34 - len(raw_row))
        created_date = parse_date(row[14])
        approved_date = parse_date(row[19])
        purchase_date = parse_date(row[18])
        changed_date = parse_date(row[32])
        posting_date = parse_date(row[33])
        status_text = clean(row[25])
        ticket_number = clean(row[1] or row[3])
        ticket_key = clean(row[0] or row[2])
        po = normalize_po(row[12])

        rows.append(
            {
                "ticket_number": ticket_number,
                "ticket_key": ticket_key,
                "ticket_id_text": clean(row[2]),
                "ticket_type": clean(row[4]),
                "claim_scope": ticket_claim_scope(row[4]),
                "agent_name": clean(row[5]),
                "agent_id": clean(row[6]),
                "serial_id": clean(row[7]),
                "chassis_number": clean(row[8]),
                "account_name": clean(row[9]),
                "account_id": clean(row[10]),
                "erp_free_order_id": clean(row[11]),
                "po": po,
                "erp_service_order_id": clean(row[13]),
                "created_on": clean(row[14]),
                "created_date": created_date,
                "dealer_code": clean(row[15]),
                "dealer_name": clean(row[16]),
                "country_region": clean(row[17]),
                "date_of_purchase": clean(row[18]),
                "purchase_date": purchase_date,
                "claim_approved_on": clean(row[19]),
                "approved_date": approved_date,
                "postal_code": clean(row[20]),
                "registered_product": clean(row[21]),
                "registered_product_code": clean(row[22]),
                "product": clean(row[23]),
                "product_code": clean(row[24]),
                "status_text": status_text,
                "service_technician": clean(row[26]),
                "service_technician_id": clean(row[27]),
                "claim_total_amount": parse_amount(row[28]),
                "factory_parts_claim_total_amount": parse_amount(row[29]),
                "labour_hours_total_amount": parse_amount(row[30]),
                "repairer_parts_claim_total_amount": parse_amount(row[31]),
                "changed_on": clean(row[32]),
                "changed_date": changed_date,
                "posting_date": clean(row[33]),
                "posting_date_value": posting_date,
                "is_unapproved": is_unapproved_status(status_text),
                "is_approved_like_status": is_approved_like_status(status_text),
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["created_year"] = out["created_date"].map(lambda value: value.year if value else None)
    return out


def load_failure_timing_rows(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    if df.empty:
        return pd.DataFrame(
            columns=[
                "ticket_number",
                "failure_ticket_created_on",
                "failure_ticket_created_date",
                "failure_pgi_date",
                "failure_vehicle_delivery_date",
                "failure_delivery_date",
                "failure_timing_source",
                "failure_days",
                "failure_matched_vehicle_key",
                "failure_matched_serial",
                "failure_matched_chassis",
                "failure_matched_sales_order",
                "failure_ticket_sales_order",
                "failure_ticket_serial_id",
                "failure_ticket_chassis_number",
            ]
        )

    work = df.copy()
    work["ticket_number"] = work["Ticket ID"].map(clean)
    work["failure_ticket_created_on"] = work["Created On"].map(clean)
    work["failure_ticket_created_date"] = work.apply(
        lambda row: parse_date(row.get("Created On ISO")) or parse_date(row.get("Created On")),
        axis=1,
    )
    work["failure_pgi_date"] = work["PGI Date"].map(clean)
    work["failure_vehicle_delivery_date"] = work["Vehicle Delivery Date"].map(clean)
    work["failure_delivery_date"] = work["failure_vehicle_delivery_date"].map(parse_date)
    work["failure_timing_source"] = work["Timing Date Source"].map(clean)
    work["failure_days"] = pd.to_numeric(work["Failure Days"].map(clean), errors="coerce")
    work["failure_days"] = work.apply(
        lambda row: (
            row["failure_days"]
            if pd.notna(row["failure_days"])
            else day_diff(row["failure_ticket_created_date"], row["failure_delivery_date"])
        ),
        axis=1,
    )
    work["failure_matched_vehicle_key"] = work["Matched Vehicle Key"].map(clean)
    work["failure_matched_serial"] = work["Matched Serial"].map(clean)
    work["failure_matched_chassis"] = work["Matched Chassis"].map(clean)
    work["failure_matched_sales_order"] = work["Matched Sales Order"].map(clean)
    work["failure_ticket_sales_order"] = work["Ticket Sales Order"].map(clean)
    work["failure_ticket_serial_id"] = work["Ticket Serial ID"].map(clean)
    work["failure_ticket_chassis_number"] = work["Ticket Chassis Number"].map(clean)
    work = work.sort_values(["ticket_number", "failure_ticket_created_on"], na_position="last")
    return work.drop_duplicates(subset=["ticket_number"], keep="first").reset_index(drop=True)


def load_parts_ticket_summary(csv_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    if df.empty:
        empty = pd.DataFrame(
            columns=[
                "ticket_number",
                "parts_po",
                "parts_sales_order",
                "so_created_date",
                "complete_issue_date",
                "parts_item_rows",
                "issued_item_rows",
            ]
        )
        return empty, empty.rename(columns={"ticket_number": "po"})

    work = df.copy()
    work["ticket_number"] = work["Ticket ID"].map(clean)
    work["parts_po"] = work["ERP Purchase Order"].map(normalize_po)
    work["parts_sales_order"] = work["Sales Order"].map(clean)
    work["so_created_date"] = work["SO Created Date"].map(parse_date)
    work["item_issue_date"] = work["First Issue Date"].map(parse_date)
    work["is_issued_item"] = work["item_issue_date"].notna()

    by_ticket = (
        work.groupby("ticket_number", dropna=False)
        .agg(
            parts_po=("parts_po", lambda values: next((value for value in values if clean(value)), "")),
            parts_sales_order=("parts_sales_order", lambda values: next((value for value in values if clean(value)), "")),
            so_created_date=("so_created_date", "min"),
            complete_issue_date=("item_issue_date", "max"),
            parts_item_rows=("ticket_number", "size"),
            issued_item_rows=("is_issued_item", "sum"),
        )
        .reset_index()
    )

    by_po = (
        work[work["parts_po"] != ""]
        .groupby("parts_po", dropna=False)
        .agg(
            parts_sales_order=("parts_sales_order", lambda values: next((value for value in values if clean(value)), "")),
            so_created_date=("so_created_date", "min"),
            complete_issue_date=("item_issue_date", "max"),
            parts_item_rows=("parts_po", "size"),
            issued_item_rows=("is_issued_item", "sum"),
        )
        .reset_index()
        .rename(columns={"parts_po": "po"})
    )

    return by_ticket, by_po


def load_repairer_invoice_rows(json_path: Path) -> pd.DataFrame:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    rows = payload.get("details", []) if isinstance(payload, dict) else []

    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        po = normalize_po(row.get("ERP Purchase Order ID"))
        if not po:
            continue
        invoice_date = parse_date(row.get("invoice_date"))
        normalized_rows.append(
            {
                "po": po,
                "invoice_status": clean(row.get("invoice_status")),
                "invoice_number": clean(row.get("invoice_number")),
                "invoice_date": invoice_date,
                "invoice_date_text": clean(row.get("invoice_date")),
                "repairer_status_text": clean(row.get("Status")),
                "repairer_ticket_key": clean(row.get("Ticket ID") or row.get("TicketID")),
                "repairer_name_from_repairs": clean(row.get("repairer_name") or row.get("Service Technician")),
            }
        )

    if not normalized_rows:
        return pd.DataFrame(
            columns=[
                "po",
                "invoice_status",
                "invoice_number",
                "invoice_date",
                "invoice_date_text",
                "repairer_status_text",
                "repairer_ticket_key",
                "repairer_name_from_repairs",
            ]
        )

    df = pd.DataFrame(normalized_rows)
    df = df.sort_values(["po", "invoice_date"], na_position="last")
    return df.drop_duplicates(subset=["po"], keep="last").reset_index(drop=True)


def build_qualifying_universe(base_df: pd.DataFrame, start_date: date, end_date: date | None) -> pd.DataFrame:
    work = base_df.copy()
    date_mask = work["created_date"].map(lambda value: value is not None and value >= start_date and (end_date is None or value <= end_date))
    qualify_mask = (
        date_mask
        & work["po"].ne("")
        & work["approved_date"].notna()
        & ~work["is_unapproved"]
        & ~work["status_text"].map(is_van_recalled_status)
    )
    out = work[qualify_mask].copy()
    out["approval_days"] = out.apply(lambda row: day_diff(row["approved_date"], row["created_date"]), axis=1)
    out["approval_days_valid"] = out["approval_days"].map(lambda value: value is not None and value >= 0)
    return out.sort_values(["created_date", "ticket_number", "ticket_key"], na_position="last").reset_index(drop=True)


def build_failure_segment(
    qualifying_df: pd.DataFrame,
    failure_df: pd.DataFrame,
    anomalies: list[dict[str, Any]],
    start_year: int,
    end_year: int,
) -> pd.DataFrame:
    merged = qualifying_df.merge(failure_df, on="ticket_number", how="left")
    merged["failure_vehicle_key"] = merged.apply(
        lambda row: next(
            (
                candidate
                for candidate in [
                    clean(row.get("failure_matched_vehicle_key")),
                    clean(row.get("failure_matched_chassis")),
                    clean(row.get("failure_matched_serial")),
                    clean(row.get("failure_matched_sales_order")),
                    clean(row.get("chassis_number")),
                    clean(row.get("serial_id")),
                    clean(row.get("failure_ticket_chassis_number")),
                    clean(row.get("failure_ticket_serial_id")),
                    clean(row.get("failure_ticket_sales_order")),
                ]
                if candidate
            ),
            "",
        ),
        axis=1,
    )
    merged["failure_days_valid"] = merged["failure_days"].notna() & merged["failure_days"].map(lambda value: value >= 0)
    merged["has_failure_timing_source"] = merged["failure_timing_source"].map(clean).ne("")
    merged["has_failure_delivery_basis"] = merged["has_failure_timing_source"] & merged["failure_timing_source"].ne("missing")
    merged["failure_delivery_year"] = merged["failure_delivery_date"].map(lambda value: value.year if value else None)
    merged["failure_delivery_year_in_scope"] = merged["failure_delivery_year"].between(start_year, end_year, inclusive="both")

    for _, row in merged[merged["failure_timing_source"].eq("")].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="missing failure timing row",
                row=row,
            )
        )

    for _, row in merged[merged["failure_timing_source"].eq("missing")].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="failure timing source is missing",
                row=row,
            )
        )

    for _, row in merged[merged["has_failure_delivery_basis"] & merged["failure_vehicle_key"].map(clean).eq("")].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="missing matched vehicle key",
                row=row,
            )
        )

    for _, row in merged[merged["has_failure_delivery_basis"] & ~merged["failure_days_valid"]].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="missing or negative failure days",
                row=row,
            )
        )

    for _, row in merged[merged["has_failure_delivery_basis"] & ~merged["failure_delivery_year_in_scope"]].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="vehicle delivery year out of scope for failure timing",
                row=row,
                detail=f"delivery_year={clean(row.get('failure_delivery_year'))}",
            )
        )

    valid = merged[
        merged["has_failure_delivery_basis"]
        & merged["failure_days_valid"]
        & merged["failure_vehicle_key"].map(clean).ne("")
        & merged["failure_delivery_year_in_scope"]
    ].copy()
    if valid.empty:
        return valid

    valid = valid.sort_values(["failure_vehicle_key", "created_date", "ticket_number", "ticket_key"], na_position="last")
    valid["failure_rank_in_vehicle"] = valid.groupby("failure_vehicle_key").cumcount() + 1

    for _, row in valid[valid["failure_rank_in_vehicle"] > 1].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Failure Timing",
                reason="later qualifying ticket on the same vehicle",
                row=row,
            )
        )

    first_only = valid[valid["failure_rank_in_vehicle"] == 1].copy()
    first_only["failure_days"] = first_only["failure_days"].round(2)
    return first_only


def merge_parts_data(
    qualifying_df: pd.DataFrame,
    parts_by_ticket_df: pd.DataFrame,
    parts_by_po_df: pd.DataFrame,
    anomalies: list[dict[str, Any]],
) -> pd.DataFrame:
    merged = qualifying_df.merge(parts_by_ticket_df, on="ticket_number", how="left", suffixes=("", "_by_ticket"))
    merged = merged.merge(parts_by_po_df, on="po", how="left", suffixes=("", "_by_po"))

    def choose_parts_value(row: pd.Series, ticket_col: str, po_col: str) -> Any:
        ticket_value = row.get(ticket_col)
        if pd.notna(ticket_value):
            return ticket_value
        return row.get(po_col)

    result = merged.copy()
    result["parts_join_source"] = ""
    result["parts_po_joined"] = ""
    result["parts_sales_order_joined"] = ""
    result["so_created_date_joined"] = pd.NaT
    result["complete_issue_date_joined"] = pd.NaT
    result["parts_item_rows_joined"] = pd.NA
    result["issued_item_rows_joined"] = pd.NA

    ticket_has_match = result["parts_po"].fillna("").astype(str).str.strip().ne("")
    po_has_match = result["parts_sales_order_by_po"].fillna("").astype(str).str.strip().ne("") | result["so_created_date_by_po"].notna() | result["complete_issue_date_by_po"].notna()

    result.loc[ticket_has_match, "parts_join_source"] = "ticket_number"
    result.loc[~ticket_has_match & po_has_match, "parts_join_source"] = "po_fallback"

    result["parts_po_joined"] = result.apply(
        lambda row: clean(row["parts_po"]) if clean(row["parts_join_source"]) == "ticket_number" else clean(row.get("po")),
        axis=1,
    )
    result["parts_sales_order_joined"] = result.apply(
        lambda row: clean(row["parts_sales_order"]) if clean(row["parts_join_source"]) == "ticket_number" else clean(row.get("parts_sales_order_by_po")),
        axis=1,
    )
    result["so_created_date_joined"] = result.apply(
        lambda row: row["so_created_date"] if clean(row["parts_join_source"]) == "ticket_number" else row.get("so_created_date_by_po"),
        axis=1,
    )
    result["complete_issue_date_joined"] = result.apply(
        lambda row: row["complete_issue_date"] if clean(row["parts_join_source"]) == "ticket_number" else row.get("complete_issue_date_by_po"),
        axis=1,
    )
    result["parts_item_rows_joined"] = result.apply(
        lambda row: row["parts_item_rows"] if clean(row["parts_join_source"]) == "ticket_number" else row.get("parts_item_rows_by_po"),
        axis=1,
    )
    result["issued_item_rows_joined"] = result.apply(
        lambda row: row["issued_item_rows"] if clean(row["parts_join_source"]) == "ticket_number" else row.get("issued_item_rows_by_po"),
        axis=1,
    )

    both_match_mask = ticket_has_match & po_has_match
    mismatch_mask = both_match_mask & (
        (result["parts_po"].fillna("").astype(str).str.strip() != result["po"].fillna("").astype(str).str.strip())
        | (result["so_created_date"].fillna(pd.Timestamp.min) != result["so_created_date_by_po"].fillna(pd.Timestamp.min))
        | (result["complete_issue_date"].fillna(pd.Timestamp.min) != result["complete_issue_date_by_po"].fillna(pd.Timestamp.min))
    )
    for _, row in result[mismatch_mask].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Parts Issuing",
                reason="ticket-number join and PO join disagree",
                row=row,
                detail=f"ticket_parts_po={clean(row.get('parts_po'))}, po_join={clean(row.get('po'))}",
            )
        )

    for _, row in result[result["parts_join_source"].eq("")].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Parts Issuing",
                reason="missing parts summary match",
                row=row,
            )
        )

    return result


def build_parts_segment(parts_joined_df: pd.DataFrame, anomalies: list[dict[str, Any]]) -> pd.DataFrame:
    work = parts_joined_df.copy()
    work["parts_issuing_days"] = work.apply(
        lambda row: day_diff(row["complete_issue_date_joined"], row["so_created_date_joined"]),
        axis=1,
    )
    work["parts_issuing_days_valid"] = work["parts_issuing_days"].map(lambda value: value is not None and value >= 0)

    for _, row in work[work["parts_join_source"].ne("") & work["so_created_date_joined"].isna()].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Parts Issuing",
                reason="missing SO created date",
                row=row,
            )
        )

    for _, row in work[work["parts_join_source"].ne("") & work["complete_issue_date_joined"].isna()].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Parts Issuing",
                reason="missing complete issue date",
                row=row,
            )
        )

    for _, row in work[work["parts_issuing_days"].map(lambda value: value is not None and value < 0)].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Parts Issuing",
                reason="negative parts issuing days",
                row=row,
            )
        )

    return work[work["parts_issuing_days_valid"]].copy()


def build_repairer_segment(
    parts_joined_df: pd.DataFrame,
    invoice_df: pd.DataFrame,
    anomalies: list[dict[str, Any]],
) -> pd.DataFrame:
    merged = parts_joined_df.merge(invoice_df, on="po", how="left")
    merged["po_last_invoice_date"] = merged["invoice_date"]
    merged["chassis_invoice_key"] = merged["chassis_number"].map(normalize_vehicle_key)

    chassis_invoice_lookup = (
        merged.loc[merged["chassis_invoice_key"].ne("") & merged["po_last_invoice_date"].notna(), ["chassis_invoice_key", "po_last_invoice_date"]]
        .sort_values(["chassis_invoice_key", "po_last_invoice_date"], na_position="last")
        .drop_duplicates(subset=["chassis_invoice_key"], keep="last")
        .set_index("chassis_invoice_key")["po_last_invoice_date"]
        .to_dict()
    )
    merged["chassis_last_invoice_date"] = merged["chassis_invoice_key"].map(chassis_invoice_lookup)

    merged["created_to_invoice_days"] = merged.apply(
        lambda row: day_diff(row["invoice_date"], row["created_date"]),
        axis=1,
    )
    merged["created_to_chassis_invoice_days"] = merged.apply(
        lambda row: day_diff(row["chassis_last_invoice_date"], row["created_date"]),
        axis=1,
    )
    merged["invoice_minus_complete_issue_days"] = merged.apply(
        lambda row: day_diff(row["invoice_date"], row["complete_issue_date_joined"]),
        axis=1,
    )
    merged["chassis_invoice_minus_complete_issue_days"] = merged.apply(
        lambda row: day_diff(row["chassis_last_invoice_date"], row["complete_issue_date_joined"]),
        axis=1,
    )
    merged["repairer_repair_days"] = merged.apply(
        lambda row: (
            None
            if row["created_to_invoice_days"] is None
            or row["approval_days"] is None
            or row["approval_days"] < 0
            or row["parts_issuing_days"] is None
            else row["created_to_invoice_days"] - row["approval_days"] - row["parts_issuing_days"]
        ),
        axis=1,
    )
    merged["repairer_repair_days_by_chassis_invoice"] = merged.apply(
        lambda row: (
            None
            if row["created_to_chassis_invoice_days"] is None
            or row["approval_days"] is None
            or row["approval_days"] < 0
            or row["parts_issuing_days"] is None
            else row["created_to_chassis_invoice_days"] - row["approval_days"] - row["parts_issuing_days"]
        ),
        axis=1,
    )
    merged["po_vs_chassis_invoice_gap_days"] = merged.apply(
        lambda row: day_diff(row["chassis_last_invoice_date"], row["po_last_invoice_date"]),
        axis=1,
    )
    merged["repairer_repair_days_valid"] = merged["repairer_repair_days"].map(lambda value: value is not None and value >= 0)

    for _, row in merged[merged["invoice_date"].isna()].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Repairer Repair Time",
                reason="missing invoice date",
                row=row,
            )
        )

    for _, row in merged[merged["created_to_invoice_days"].map(lambda value: value is not None and value < 0)].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Repairer Repair Time",
                reason="invoice date is earlier than created date",
                row=row,
            )
        )

    for _, row in merged[merged["repairer_repair_days"].map(lambda value: value is not None and value < 0)].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Repairer Repair Time",
                reason="negative repairer repair days after subtracting approval and parts time",
                row=row,
            )
        )

    return merged[merged["repairer_repair_days_valid"]].copy()


def anomaly_row(stage: str, reason: str, row: pd.Series, detail: str = "") -> dict[str, Any]:
    return {
        "stage": stage,
        "reason": reason,
        "ticket_number": clean(row.get("ticket_number")),
        "ticket_key": clean(row.get("ticket_key")),
        "claim_scope": clean(row.get("claim_scope")),
        "ticket_type": clean(row.get("ticket_type")),
        "status_text": clean(row.get("status_text")),
        "dealer_name": clean(row.get("dealer_name")),
        "repair_shop": clean(row.get("service_technician")),
        "po": clean(row.get("po")),
        "created_on": clean(row.get("created_on")),
        "claim_approved_on": clean(row.get("claim_approved_on")),
        "failure_pgi_date": clean(row.get("failure_pgi_date")),
        "so_created_date": iso_or_blank(row.get("so_created_date_joined")),
        "complete_issue_date": iso_or_blank(row.get("complete_issue_date_joined")),
        "invoice_date": iso_or_blank(row.get("invoice_date")),
        "approval_days": row.get("approval_days"),
        "parts_issuing_days": row.get("parts_issuing_days"),
        "failure_days": row.get("failure_days"),
        "created_to_invoice_days": row.get("created_to_invoice_days"),
        "repairer_repair_days": row.get("repairer_repair_days"),
        "detail": detail,
    }


def build_summary_sheet(
    qualifying_df: pd.DataFrame,
    approval_segment_df: pd.DataFrame,
    parts_segment_df: pd.DataFrame,
    repairer_segment_df: pd.DataFrame,
    anomalies_df: pd.DataFrame,
    start_date: date,
    end_date: date | None,
) -> pd.DataFrame:
    created_filter = f">= {start_date.isoformat()}" if end_date is None else f"{start_date.isoformat()} to {end_date.isoformat()}"
    rows = [
        {"Metric": "Run At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "Created On Filter", "Value": created_filter},
        {"Metric": "Qualification Filter", "Value": "Created On in scope + matched PO + Claim Approved On present + not unapproved"},
        {"Metric": "Warranty Approval Basis", "Value": "Claim Approved On - Created On"},
        {"Metric": "Parts Issuing Basis", "Value": "Complete Issue Date - SO Created Date"},
        {"Metric": "Repairer Basis", "Value": "Invoice Date - Created On - Approval Days - Parts Issuing Days"},
        {"Metric": "Denominator Rule", "Value": "Each segment uses its own valid rows only; missing dates for a segment are excluded from that segment denominator."},
        {"Metric": "Qualifying Tickets", "Value": int(len(qualifying_df))},
        {"Metric": "Warranty Approval Rows", "Value": int(len(approval_segment_df))},
        {"Metric": "Warranty Approval Avg Days", "Value": mean_or_blank(approval_segment_df.get("approval_days", []))},
        {"Metric": "Warranty Approval Median Days", "Value": median_or_blank(approval_segment_df.get("approval_days", []))},
        {"Metric": "Parts Issuing Rows", "Value": int(len(parts_segment_df))},
        {"Metric": "Parts Issuing Avg Days", "Value": mean_or_blank(parts_segment_df.get("parts_issuing_days", []))},
        {"Metric": "Parts Issuing Median Days", "Value": median_or_blank(parts_segment_df.get("parts_issuing_days", []))},
        {"Metric": "Repairer Rows", "Value": int(len(repairer_segment_df))},
        {"Metric": "Repairer Avg Days", "Value": mean_or_blank(repairer_segment_df.get("repairer_repair_days", []))},
        {"Metric": "Repairer Median Days", "Value": median_or_blank(repairer_segment_df.get("repairer_repair_days", []))},
        {"Metric": "Anomaly Rows", "Value": int(len(anomalies_df))},
    ]
    return pd.DataFrame(rows)


def build_qualifying_sheet(
    qualifying_df: pd.DataFrame,
    approval_segment_df: pd.DataFrame,
    parts_segment_df: pd.DataFrame,
    repairer_segment_df: pd.DataFrame,
) -> pd.DataFrame:
    out = qualifying_df.copy()
    out["used_in_warranty_approval"] = out["ticket_number"].isin(set(approval_segment_df["ticket_number"]))
    out["used_in_parts_issuing"] = out["ticket_number"].isin(set(parts_segment_df["ticket_number"]))
    out["used_in_repairer_time"] = out["ticket_number"].isin(set(repairer_segment_df["ticket_number"]))

    parts_days_map = parts_segment_df.set_index("ticket_number")["parts_issuing_days"].to_dict() if not parts_segment_df.empty else {}
    repairer_days_map = repairer_segment_df.set_index("ticket_number")["repairer_repair_days"].to_dict() if not repairer_segment_df.empty else {}

    out["parts_issuing_days"] = out["ticket_number"].map(parts_days_map)
    out["repairer_repair_days"] = out["ticket_number"].map(repairer_days_map)

    return out[
        [
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "agent_name",
            "agent_id",
            "dealer_code",
            "dealer_name",
            "account_name",
            "account_id",
            "country_region",
            "service_technician",
            "service_technician_id",
            "po",
            "erp_service_order_id",
            "erp_free_order_id",
            "created_on",
            "claim_approved_on",
            "changed_on",
            "posting_date",
            "date_of_purchase",
            "serial_id",
            "chassis_number",
            "registered_product",
            "registered_product_code",
            "product",
            "product_code",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
            "approval_days",
            "used_in_warranty_approval",
            "used_in_parts_issuing",
            "parts_issuing_days",
            "used_in_repairer_time",
            "repairer_repair_days",
        ]
    ].copy()


COMPLETION_BUCKETS = [
    ("0-7", "#0f8f8c"),
    ("8-14", "#4f8f42"),
    ("15-21", "#2474a6"),
    ("22-30", "#d88a18"),
    ("31-60", "#ef9f42"),
    ("60+", "#c9513f"),
]
AMOUNT_BUCKETS = [
    ("$0-500", "#5bb8b0", 0, 500),
    ("$500-2k", "#4f87c7", 500, 2000),
    ("$2k-5k", "#f0a33a", 2000, 5000),
    ("$5k+", "#c9513f", 5000, None),
]
TOTAL_COMPLETION_BUCKETS = [
    ("0-30", "#0f8f8c"),
    ("31-60", "#4f8f42"),
    ("61-90", "#2474a6"),
    ("91-120", "#d88a18"),
    ("121-180", "#ef9f42"),
    ("180+", "#c9513f"),
]
MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def completed_in_year(value: Any, year: int, as_of: date) -> bool:
    completed_date = to_date(value)
    return completed_date is not None and completed_date.year == year and completed_date <= as_of


def duration_bucket_name(value: float | int, buckets: list[tuple[str, str]] | None = None) -> str:
    labels = [label for label, _ in (buckets or COMPLETION_BUCKETS)]
    if labels and labels[-1] == "180+":
        if value <= 30:
            return "0-30"
        if value <= 60:
            return "31-60"
        if value <= 90:
            return "61-90"
        if value <= 120:
            return "91-120"
        if value <= 180:
            return "121-180"
        return "180+"
    if value <= 7:
        return "0-7"
    if value <= 14:
        return "8-14"
    if value <= 21:
        return "15-21"
    if value <= 30:
        return "22-30"
    if value <= 60:
        return "31-60"
    return "60+"


def is_valid_duration_value(value: Any) -> bool:
    try:
        return bool(pd.notna(value) and float(value) >= 0)
    except (TypeError, ValueError):
        return False


def valid_duration_values(values: Iterable[Any]) -> list[float]:
    return [float(value) for value in values if is_valid_duration_value(value)]


def valid_amount_value(value: Any) -> float | None:
    amount = parse_amount(value)
    if amount is None:
        return None
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(amount) or amount <= 0:
        return None
    return amount


def amount_bucket_name(value: float) -> str:
    for label, _color, low, high in AMOUNT_BUCKETS:
        if value >= low and (high is None or value < high):
            return label
    return AMOUNT_BUCKETS[-1][0]


def build_amount_distribution(df: pd.DataFrame, amount_col: str = "claim_total_amount") -> dict[str, Any]:
    if df is None or df.empty or amount_col not in df.columns:
        return {"totalAmount": 0, "pricedTickets": 0, "missingTickets": int(len(df)) if df is not None else 0, "buckets": []}
    amounts: list[float] = []
    missing = 0
    for value in df[amount_col].tolist():
        amount = valid_amount_value(value)
        if amount is None:
            missing += 1
        else:
            amounts.append(amount)
    total_amount = sum(amounts)
    rows: list[dict[str, Any]] = []
    for label, color, _low, _high in AMOUNT_BUCKETS:
        bucket_values = [amount for amount in amounts if amount_bucket_name(amount) == label]
        bucket_total = sum(bucket_values)
        rows.append(
            {
                "label": label,
                "color": color,
                "amount": round(bucket_total, 2),
                "pct": round(bucket_total / total_amount * 100, 1) if total_amount else 0,
                "count": len(bucket_values),
            }
        )
    return {
        "totalAmount": round(total_amount, 2),
        "pricedTickets": len(amounts),
        "missingTickets": missing,
        "buckets": rows,
    }


def build_completion_buckets(values: Iterable[Any], buckets: list[tuple[str, str]] | None = None) -> list[list[Any]]:
    bucket_defs = buckets or COMPLETION_BUCKETS
    valid_values = valid_duration_values(values)
    total = len(valid_values)
    rows: list[list[Any]] = []
    for label, color in bucket_defs:
        count = sum(1 for value in valid_values if duration_bucket_name(value, bucket_defs) == label)
        pct = round(count / total * 100, 1) if total else 0
        rows.append([label, pct, color, count])
    return rows


def build_duration_amount_distributions(
    df: pd.DataFrame,
    duration_col: str,
    bucket_defs: list[tuple[str, str]] | None = None,
) -> dict[str, Any]:
    defs = bucket_defs or COMPLETION_BUCKETS
    out: dict[str, Any] = {}
    if df is None or df.empty or duration_col not in df.columns:
        return {label: build_amount_distribution(pd.DataFrame()) for label, _color in defs}
    valid_df = df[df[duration_col].map(is_valid_duration_value)].copy()
    for label, _color in defs:
        bucket_df = valid_df[valid_df[duration_col].map(lambda value, label=label: duration_bucket_name(float(value), defs) == label)]
        out[label] = build_amount_distribution(bucket_df)
    return out


def build_price_duration_mix(
    df: pd.DataFrame,
    duration_col: str,
    amount_col: str = "claim_total_amount",
    bucket_defs: list[tuple[str, str]] | None = None,
) -> dict[str, Any]:
    duration_buckets = bucket_defs or COMPLETION_BUCKETS
    empty_price_buckets = [
        {
            "label": label,
            "color": color,
            "count": 0,
            "average": None,
            "distribution": [[duration_label, 0, duration_color, 0] for duration_label, duration_color in duration_buckets],
        }
        for label, color, _low, _high in AMOUNT_BUCKETS
    ]
    if df is None or df.empty or duration_col not in df.columns or amount_col not in df.columns:
        return {"pricedTickets": 0, "missingAmount": int(len(df)) if df is not None else 0, "buckets": empty_price_buckets}

    priced_rows: list[dict[str, Any]] = []
    missing_amount = 0
    for _, row in df.iterrows():
        duration_value = row.get(duration_col)
        if not is_valid_duration_value(duration_value):
            continue
        amount = valid_amount_value(row.get(amount_col))
        if amount is None:
            missing_amount += 1
            continue
        priced_rows.append(
            {
                "amount": amount,
                "duration": float(duration_value),
                "amountBucket": amount_bucket_name(amount),
            }
        )

    buckets: list[dict[str, Any]] = []
    for label, color, _low, _high in AMOUNT_BUCKETS:
        bucket_rows = [row for row in priced_rows if row["amountBucket"] == label]
        duration_values = [row["duration"] for row in bucket_rows]
        total = len(duration_values)
        distribution: list[list[Any]] = []
        for duration_label, duration_color in duration_buckets:
            count = sum(1 for value in duration_values if duration_bucket_name(value, duration_buckets) == duration_label)
            pct = round(count / total * 100, 1) if total else 0
            distribution.append([duration_label, pct, duration_color, count])
        buckets.append(
            {
                "label": label,
                "color": color,
                "count": total,
                "average": round(sum(duration_values) / total, 2) if total else None,
                "distribution": distribution,
            }
        )

    return {
        "pricedTickets": len(priced_rows),
        "missingAmount": missing_amount,
        "buckets": buckets,
    }


def build_completion_month_trend(
    df: pd.DataFrame,
    completed_col: str,
    duration_col: str,
    threshold_days: int,
    year: int,
    as_of: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    monthly_avgs: list[float | None] = []
    for month_index, month_label in enumerate(MONTH_LABELS, start=1):
        if date(year, month_index, 1) > as_of:
            break
        month_df = df[
            df[completed_col].map(
                lambda value, month_index=month_index: (
                    (completed_date := to_date(value)) is not None
                    and completed_date.year == year
                    and completed_date.month == month_index
                    and completed_date <= as_of
                )
            )
        ].copy()
        values = valid_duration_values(month_df.get(duration_col, []))
        avg = round(sum(values) / len(values), 2) if values else None
        monthly_avgs.append(avg)
        rolling_values = [value for value in monthly_avgs[-3:] if value is not None]
        over_count = sum(1 for value in values if value > threshold_days)
        prior_created = 0
        if not month_df.empty and "created_date" in month_df.columns:
            prior_created = sum(
                1
                for value in month_df["created_date"].tolist()
                if (created_date := to_date(value)) is not None and created_date.year < year
            )
        rows.append(
            {
                "label": month_label,
                "avg": avg,
                "completed": len(values),
                "over": over_count,
                "overPct": round(over_count / len(values) * 100, 1) if values else 0,
                "priorCreated": prior_created,
                "roll3": round(sum(rolling_values) / len(rolling_values), 2) if rolling_values else None,
            }
        )
    return rows


def build_completion_month_buckets(
    df: pd.DataFrame,
    completed_col: str,
    duration_col: str,
    year: int,
    as_of: date,
    bucket_defs: list[tuple[str, str]] | None = None,
) -> dict[str, list[list[Any]]]:
    month_buckets: dict[str, list[list[Any]]] = {}
    for month_index, month_label in enumerate(MONTH_LABELS, start=1):
        if date(year, month_index, 1) > as_of:
            break
        month_df = df[
            df[completed_col].map(
                lambda value, month_index=month_index: (
                    (completed_date := to_date(value)) is not None
                    and completed_date.year == year
                    and completed_date.month == month_index
                    and completed_date <= as_of
                )
            )
        ]
        month_buckets[month_label] = build_completion_buckets(month_df.get(duration_col, []), bucket_defs)
    return month_buckets


def latest_overtime_ticket_rows(
    df: pd.DataFrame,
    completed_col: str,
    duration_col: str,
    threshold_days: int,
    owner_col: str,
) -> list[list[str]]:
    if df.empty:
        return []
    work = df[df[duration_col].map(lambda value: pd.notna(value) and float(value) > threshold_days)].copy()
    if work.empty:
        return []
    work["_completed_sort_date"] = work[completed_col].map(to_date)
    work = work.sort_values(["_completed_sort_date", "ticket_number"], ascending=[False, False]).head(3)
    rows: list[list[str]] = []
    for _, row in work.iterrows():
        rows.append(
            [
                f"#{clean(row.get('ticket_number'))}",
                clean(row.get(owner_col)),
                f"{int(round(float(row.get(duration_col))))} days",
            ]
        )
    return rows


def completion_stage_payload(
    df: pd.DataFrame,
    completed_col: str,
    duration_col: str,
    threshold_days: int,
    year: int,
    as_of: date,
    responsibility: str,
    export_label: str,
    owner_col: str,
    bucket_defs: list[tuple[str, str]] | None = None,
    title: str = "Handling Duration Distribution",
) -> dict[str, Any]:
    values = valid_duration_values(df.get(duration_col, []))
    over_count = sum(1 for value in values if value > threshold_days)
    return {
        "responsibility": responsibility,
        "exportLabel": export_label,
        "title": title,
        "overLabel": f"{round(over_count / len(values) * 100, 1) if values else 0}% over",
        "buckets": build_completion_buckets(values, bucket_defs),
        "monthBuckets": build_completion_month_buckets(df, completed_col, duration_col, year, as_of, bucket_defs),
        "trend": build_completion_month_trend(df, completed_col, duration_col, threshold_days, year, as_of),
        "tickets": latest_overtime_ticket_rows(df, completed_col, duration_col, threshold_days, owner_col),
    }


def build_total_handling_df(parts_joined_df: pd.DataFrame, invoice_df: pd.DataFrame) -> pd.DataFrame:
    total = parts_joined_df.merge(invoice_df, on="po", how="left")
    total["created_to_invoice_days"] = total.apply(
        lambda row: day_diff(row.get("invoice_date"), row.get("created_date")),
        axis=1,
    )
    total["created_to_invoice_days_completed"] = total["created_to_invoice_days"]
    return total


def build_created_month_trend(
    completed_df: pd.DataFrame,
    universe_df: pd.DataFrame,
    duration_col: str,
    threshold_days: int,
    year: int,
    as_of: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    monthly_avgs: list[float | None] = []
    for month_index, month_label in enumerate(MONTH_LABELS, start=1):
        if date(year, month_index, 1) > as_of:
            break
        universe_month_df = universe_df[
            universe_df["created_date"].map(
                lambda value, month_index=month_index: (
                    (created_date := to_date(value)) is not None
                    and created_date.year == year
                    and created_date.month == month_index
                )
            )
        ]
        month_df = completed_df[
            completed_df["created_date"].map(
                lambda value, month_index=month_index: (
                    (created_date := to_date(value)) is not None
                    and created_date.year == year
                    and created_date.month == month_index
                )
            )
        ]
        completed_month_df = month_df[month_df[duration_col].map(is_valid_duration_value)].copy() if duration_col in month_df.columns else month_df.iloc[0:0].copy()
        values = valid_duration_values(completed_month_df.get(duration_col, []))
        avg = round(sum(values) / len(values), 2) if values else None
        monthly_avgs.append(avg)
        rolling_values = [value for value in monthly_avgs[-3:] if value is not None]
        over_count = sum(1 for value in values if value > threshold_days)
        created_count = int(len(universe_month_df))
        completed_count = int(len(values))
        rows.append(
            {
                "label": month_label,
                "avg": avg,
                "count": completed_count,
                "created": created_count,
                "completed": completed_count,
                "unfinished": max(created_count - completed_count, 0),
                "over": over_count,
                "roll3": round(sum(rolling_values) / len(rolling_values), 2) if rolling_values else None,
                "amountDistribution": build_amount_distribution(completed_month_df),
            }
        )
    return rows


def build_timeline_evidence(
    completed_df: pd.DataFrame,
    universe_df: pd.DataFrame,
    duration_col: str,
    threshold_days: int,
    year: int,
    as_of: date,
) -> dict[str, Any]:
    values = valid_duration_values(completed_df.get(duration_col, []))
    return {
        "standard": threshold_days,
        "count": len(values),
        "createdCount": int(len(universe_df)),
        "unfinishedCount": max(int(len(universe_df)) - len(values), 0),
        "average": round(sum(values) / len(values), 2) if values else None,
        "trend": build_created_month_trend(completed_df, universe_df, duration_col, threshold_days, year, as_of),
        "distribution": build_completion_buckets(values),
        "amountDistribution": build_amount_distribution(completed_df),
        "distributionAmount": build_duration_amount_distributions(completed_df, duration_col),
        "priceDurationMix": build_price_duration_mix(completed_df, duration_col),
    }


def completion_detail_records(
    df: pd.DataFrame,
    stage: str,
    completed_col: str,
    duration_col: str,
    threshold_days: int,
    year: int,
    as_of: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        completed_date = to_date(row.get(completed_col))
        if completed_date is None or completed_date.year != year or completed_date > as_of:
            continue
        duration = row.get(duration_col)
        try:
            if pd.isna(duration) or float(duration) < 0:
                continue
        except (TypeError, ValueError):
            continue
        created_date = to_date(row.get("created_date"))
        rows.append(
            {
                "stage": stage,
                "completed_month": completed_date.strftime("%Y-%m"),
                "completed_date": completed_date.isoformat(),
                "duration_days": round(float(duration), 2),
                "standard_days": threshold_days,
                "over_standard": float(duration) > threshold_days,
                "created_year": created_date.year if created_date else "",
                "ticket_number": clean(row.get("ticket_number")),
                "ticket_key": clean(row.get("ticket_key")),
                "ticket_type": clean(row.get("ticket_type")),
                "status_text": clean(row.get("status_text")),
                "agent_name": clean(row.get("agent_name")),
                "dealer_name": clean(row.get("dealer_name")),
                "service_technician": clean(row.get("service_technician")),
                "po": clean(row.get("po")),
                "erp_service_order_id": clean(row.get("erp_service_order_id")),
                "created_on": clean(row.get("created_on")),
                "claim_approved_on": clean(row.get("claim_approved_on")),
                "parts_so_created_date": iso_or_blank(row.get("so_created_date_joined")),
                "parts_complete_issue_date": iso_or_blank(row.get("complete_issue_date_joined")),
                "invoice_date": iso_or_blank(row.get("invoice_date")),
                "serial_id": clean(row.get("serial_id")),
                "chassis_number": clean(row.get("chassis_number")),
                "registered_product": clean(row.get("registered_product")),
                "product": clean(row.get("product")),
                "claim_total_amount": row.get("claim_total_amount"),
                "factory_parts_claim_total_amount": row.get("factory_parts_claim_total_amount"),
                "labour_hours_total_amount": row.get("labour_hours_total_amount"),
                "repairer_parts_claim_total_amount": row.get("repairer_parts_claim_total_amount"),
            }
        )
    return rows


def build_completion_analytics_payload(
    base_df: pd.DataFrame,
    parts_by_ticket_df: pd.DataFrame,
    parts_by_po_df: pd.DataFrame,
    invoice_df: pd.DataFrame,
    year: int,
    as_of: date,
) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame]:
    base_all_df = base_df[base_df["po"].ne("") & ~base_df["is_unapproved"]].copy()
    base_all_df["approval_days"] = base_all_df.apply(lambda row: day_diff(row["approved_date"], row["created_date"]), axis=1)
    base_all_df["approval_days_completed"] = base_all_df["approval_days"]

    approval_completed_df = base_all_df[
        base_all_df["approved_date"].map(lambda value: completed_in_year(value, year, as_of))
        & base_all_df["approval_days_completed"].map(lambda value: value is not None and value >= 0)
    ].copy()

    scratch_anomalies: list[dict[str, Any]] = []
    all_parts_joined_df = merge_parts_data(base_all_df, parts_by_ticket_df, parts_by_po_df, scratch_anomalies)
    all_parts_joined_df["parts_issuing_days"] = all_parts_joined_df.apply(
        lambda row: day_diff(row["complete_issue_date_joined"], row["so_created_date_joined"]),
        axis=1,
    )
    all_parts_joined_df["parts_issuing_days_completed"] = all_parts_joined_df["parts_issuing_days"]
    parts_completed_df = all_parts_joined_df[
        all_parts_joined_df["complete_issue_date_joined"].map(lambda value: completed_in_year(value, year, as_of))
        & all_parts_joined_df["parts_issuing_days_completed"].map(lambda value: value is not None and value >= 0)
    ].copy()

    valid_parts_for_repair_df = all_parts_joined_df[
        all_parts_joined_df["parts_issuing_days"].map(lambda value: value is not None and value >= 0)
    ].copy()
    repairer_all_df = build_repairer_segment(valid_parts_for_repair_df, invoice_df, scratch_anomalies)
    repairer_completed_df = repairer_all_df[
        repairer_all_df["invoice_date"].map(lambda value: completed_in_year(value, year, as_of))
    ].copy()
    total_handling_all_df = build_total_handling_df(all_parts_joined_df, invoice_df)
    total_handling_completed_df = total_handling_all_df[
        total_handling_all_df["invoice_date"].map(lambda value: completed_in_year(value, year, as_of))
        & total_handling_all_df["created_to_invoice_days_completed"].map(lambda value: value is not None and value >= 0)
    ].copy()

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "year": year,
        "basis": "completed month in selected year; ticket Created On can be earlier than the selected year",
        "stages": {
            "approval": completion_stage_payload(
                approval_completed_df,
                completed_col="approved_date",
                duration_col="approval_days_completed",
                threshold_days=27,
                year=year,
                as_of=as_of,
                responsibility="Warranty employee",
                export_label="Warranty Approval",
                owner_col="agent_name",
            ),
            "parts": completion_stage_payload(
                parts_completed_df,
                completed_col="complete_issue_date_joined",
                duration_col="parts_issuing_days_completed",
                threshold_days=21,
                year=year,
                as_of=as_of,
                responsibility="Parts tickets + days",
                export_label="Parts Preparation",
                owner_col="",
            ),
            "repair": completion_stage_payload(
                repairer_completed_df,
                completed_col="invoice_date",
                duration_col="repairer_repair_days",
                threshold_days=33,
                year=year,
                as_of=as_of,
                responsibility="Repairer information",
                export_label="Repairer Time",
                owner_col="service_technician",
            ),
            "closed": completion_stage_payload(
                total_handling_completed_df,
                completed_col="invoice_date",
                duration_col="created_to_invoice_days_completed",
                threshold_days=81,
                year=year,
                as_of=as_of,
                responsibility="Total handling",
                export_label="Total Handling",
                owner_col="service_technician",
                bucket_defs=TOTAL_COMPLETION_BUCKETS,
                title="Total Handling Duration Distribution",
            ),
        },
    }

    detail_rows = (
        completion_detail_records(approval_completed_df, "Warranty Approval", "approved_date", "approval_days_completed", 27, year, as_of)
        + completion_detail_records(parts_completed_df, "Parts Preparation", "complete_issue_date_joined", "parts_issuing_days_completed", 21, year, as_of)
        + completion_detail_records(repairer_completed_df, "Repairer Time", "invoice_date", "repairer_repair_days", 33, year, as_of)
        + completion_detail_records(total_handling_completed_df, "Total Handling", "invoice_date", "created_to_invoice_days_completed", 81, year, as_of)
    )
    detail_df = pd.DataFrame(detail_rows)
    if not detail_df.empty:
        detail_df = detail_df.sort_values(["stage", "completed_date", "ticket_number"], na_position="last").reset_index(drop=True)
    total_detail_df = pd.DataFrame(
        completion_detail_records(total_handling_completed_df, "Total Handling", "invoice_date", "created_to_invoice_days_completed", 81, year, as_of)
    )
    if not total_detail_df.empty:
        total_detail_df = total_detail_df.sort_values(["completed_date", "ticket_number"], na_position="last").reset_index(drop=True)
    return payload, detail_df, total_detail_df


def prepare_failure_sheet(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        [
            "failure_vehicle_key",
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "agent_name",
            "dealer_code",
            "dealer_name",
            "account_name",
            "country_region",
            "service_technician",
            "service_technician_id",
            "po",
            "erp_service_order_id",
            "created_on",
            "claim_approved_on",
            "changed_on",
            "posting_date",
            "date_of_purchase",
            "failure_pgi_date",
            "failure_vehicle_delivery_date",
            "failure_days",
            "failure_timing_source",
            "failure_matched_vehicle_key",
            "failure_matched_serial",
            "failure_matched_chassis",
            "failure_matched_sales_order",
            "failure_ticket_sales_order",
            "failure_ticket_serial_id",
            "failure_ticket_chassis_number",
            "serial_id",
            "chassis_number",
            "registered_product",
            "registered_product_code",
            "product",
            "product_code",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
        ]
    ].copy()


def prepare_approval_sheet(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        [
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "agent_name",
            "dealer_code",
            "dealer_name",
            "account_name",
            "country_region",
            "service_technician",
            "service_technician_id",
            "po",
            "erp_service_order_id",
            "created_on",
            "claim_approved_on",
            "approval_days",
            "changed_on",
            "posting_date",
            "date_of_purchase",
            "serial_id",
            "chassis_number",
            "registered_product",
            "registered_product_code",
            "product",
            "product_code",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
        ]
    ].copy()


def prepare_parts_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["so_created_date"] = out["so_created_date_joined"].map(iso_or_blank)
    out["complete_issue_date"] = out["complete_issue_date_joined"].map(iso_or_blank)
    return out[
        [
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "agent_name",
            "dealer_code",
            "dealer_name",
            "account_name",
            "country_region",
            "service_technician",
            "service_technician_id",
            "po",
            "erp_service_order_id",
            "parts_join_source",
            "parts_sales_order_joined",
            "so_created_date",
            "complete_issue_date",
            "parts_issuing_days",
            "parts_item_rows_joined",
            "issued_item_rows_joined",
            "created_on",
            "claim_approved_on",
            "changed_on",
            "posting_date",
            "date_of_purchase",
            "serial_id",
            "chassis_number",
            "registered_product",
            "registered_product_code",
            "product",
            "product_code",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
        ]
    ].copy()


def prepare_repairer_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["so_created_date"] = out["so_created_date_joined"].map(iso_or_blank)
    out["complete_issue_date"] = out["complete_issue_date_joined"].map(iso_or_blank)
    out["po_last_invoice_date"] = out["po_last_invoice_date"].map(iso_or_blank)
    out["chassis_last_invoice_date"] = out["chassis_last_invoice_date"].map(iso_or_blank)
    out["repairer_formula_text_po_invoice"] = out.apply(
        lambda row: (
            ""
            if pd.isna(row.get("created_to_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days"))
            else (
                f"{int(row['created_to_invoice_days'])}"
                f" - {int(row['approval_days'])}"
                f" - {int(row['parts_issuing_days'])}"
                f" = {int(row['repairer_repair_days'])}"
            )
        ),
        axis=1,
    )
    out["repairer_formula_text_chassis_invoice"] = out.apply(
        lambda row: (
            ""
            if pd.isna(row.get("created_to_chassis_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days_by_chassis_invoice"))
            else (
                f"{int(row['created_to_chassis_invoice_days'])}"
                f" - {int(row['approval_days'])}"
                f" - {int(row['parts_issuing_days'])}"
                f" = {int(row['repairer_repair_days_by_chassis_invoice'])}"
            )
        ),
        axis=1,
    )
    out = out[
        [
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "agent_name",
            "dealer_code",
            "dealer_name",
            "account_name",
            "country_region",
            "service_technician",
            "service_technician_id",
            "po",
            "erp_service_order_id",
            "created_on",
            "claim_approved_on",
            "changed_on",
            "posting_date",
            "date_of_purchase",
            "so_created_date",
            "complete_issue_date",
            "po_last_invoice_date",
            "chassis_last_invoice_date",
            "approval_days",
            "parts_issuing_days",
            "created_to_invoice_days",
            "created_to_chassis_invoice_days",
            "repairer_repair_days",
            "repairer_repair_days_by_chassis_invoice",
            "invoice_minus_complete_issue_days",
            "chassis_invoice_minus_complete_issue_days",
            "po_vs_chassis_invoice_gap_days",
            "repairer_formula_text_po_invoice",
            "repairer_formula_text_chassis_invoice",
            "invoice_status",
            "invoice_number",
            "serial_id",
            "chassis_number",
            "registered_product",
            "registered_product_code",
            "product",
            "product_code",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
        ]
    ].copy()
    return out.rename(
        columns={
            "created_on": "ticket_created_on",
            "claim_approved_on": "claim_approved_on_ticket",
            "changed_on": "ticket_changed_on",
            "posting_date": "ticket_posting_date",
            "so_created_date": "parts_so_created_date",
            "complete_issue_date": "parts_complete_issue_date",
            "approval_days": "warranty_approval_days",
            "parts_issuing_days": "parts_preparation_days",
            "created_to_invoice_days": "total_cycle_days_created_to_po_invoice",
            "created_to_chassis_invoice_days": "total_cycle_days_created_to_chassis_invoice",
            "repairer_repair_days": "repairer_net_days_po_invoice",
            "repairer_repair_days_by_chassis_invoice": "repairer_net_days_chassis_invoice",
            "invoice_minus_complete_issue_days": "po_invoice_minus_complete_issue_days_audit",
            "chassis_invoice_minus_complete_issue_days": "chassis_invoice_minus_complete_issue_days_audit",
        }
    )


def prepare_formula_breakdown_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["ticket_created_on"] = out["created_on"]
    out["claim_approved_on_ticket"] = out["claim_approved_on"]
    out["parts_so_created_date"] = out["so_created_date_joined"].map(iso_or_blank)
    out["parts_complete_issue_date"] = out["complete_issue_date_joined"].map(iso_or_blank)
    out["po_last_invoice_date"] = out["po_last_invoice_date"].map(iso_or_blank)
    out["chassis_last_invoice_date"] = out["chassis_last_invoice_date"].map(iso_or_blank)
    out["warranty_approval_days"] = out["approval_days"]
    out["parts_preparation_days"] = out["parts_issuing_days"]
    out["total_cycle_days_created_to_po_invoice"] = out["created_to_invoice_days"]
    out["total_cycle_days_created_to_chassis_invoice"] = out["created_to_chassis_invoice_days"]
    out["repairer_net_days_po_invoice"] = out["repairer_repair_days"]
    out["repairer_net_days_chassis_invoice"] = out["repairer_repair_days_by_chassis_invoice"]
    out["po_invoice_minus_complete_issue_days_audit"] = out["invoice_minus_complete_issue_days"]
    out["chassis_invoice_minus_complete_issue_days_audit"] = out["chassis_invoice_minus_complete_issue_days"]
    out["approved_to_so_created_days_audit"] = out.apply(
        lambda row: day_diff(row.get("so_created_date_joined"), row.get("approved_date")),
        axis=1,
    )
    out["repairer_formula_text_po_invoice"] = out.apply(
        lambda row: (
            ""
            if pd.isna(row.get("created_to_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days"))
            else (
                f"({int(row['created_to_invoice_days'])}) total cycle"
                f" - ({int(row['approval_days'])}) warranty approval"
                f" - ({int(row['parts_issuing_days'])}) parts prep"
                f" = ({int(row['repairer_repair_days'])}) repairer"
            )
        ),
        axis=1,
    )
    out["repairer_formula_text_chassis_invoice"] = out.apply(
        lambda row: (
            ""
            if pd.isna(row.get("created_to_chassis_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days_by_chassis_invoice"))
            else (
                f"({int(row['created_to_chassis_invoice_days'])}) total cycle"
                f" - ({int(row['approval_days'])}) warranty approval"
                f" - ({int(row['parts_issuing_days'])}) parts prep"
                f" = ({int(row['repairer_repair_days_by_chassis_invoice'])}) repairer"
            )
        ),
        axis=1,
    )
    out["formula_balance_check_po_invoice"] = out.apply(
        lambda row: (
            None
            if pd.isna(row.get("created_to_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days"))
            else int(row["created_to_invoice_days"]) - int(row["approval_days"]) - int(row["parts_issuing_days"]) - int(row["repairer_repair_days"])
        ),
        axis=1,
    )
    out["formula_balance_check_chassis_invoice"] = out.apply(
        lambda row: (
            None
            if pd.isna(row.get("created_to_chassis_invoice_days"))
            or pd.isna(row.get("approval_days"))
            or pd.isna(row.get("parts_issuing_days"))
            or pd.isna(row.get("repairer_repair_days_by_chassis_invoice"))
            else int(row["created_to_chassis_invoice_days"]) - int(row["approval_days"]) - int(row["parts_issuing_days"]) - int(row["repairer_repair_days_by_chassis_invoice"])
        ),
        axis=1,
    )
    return out[
        [
            "ticket_number",
            "ticket_key",
            "ticket_id_text",
            "ticket_type",
            "claim_scope",
            "status_text",
            "dealer_name",
            "service_technician",
            "po",
            "erp_service_order_id",
            "ticket_created_on",
            "claim_approved_on_ticket",
            "parts_so_created_date",
            "parts_complete_issue_date",
            "po_last_invoice_date",
            "chassis_last_invoice_date",
            "total_cycle_days_created_to_po_invoice",
            "total_cycle_days_created_to_chassis_invoice",
            "warranty_approval_days",
            "parts_preparation_days",
            "repairer_net_days_po_invoice",
            "repairer_net_days_chassis_invoice",
            "po_invoice_minus_complete_issue_days_audit",
            "chassis_invoice_minus_complete_issue_days_audit",
            "approved_to_so_created_days_audit",
            "po_vs_chassis_invoice_gap_days",
            "formula_balance_check_po_invoice",
            "formula_balance_check_chassis_invoice",
            "repairer_formula_text_po_invoice",
            "repairer_formula_text_chassis_invoice",
            "serial_id",
            "chassis_number",
            "registered_product",
            "product",
            "claim_total_amount",
            "factory_parts_claim_total_amount",
            "labour_hours_total_amount",
            "repairer_parts_claim_total_amount",
        ]
    ].copy()


def autosize_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    for sheet in workbook.worksheets:
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            sheet.column_dimensions[letter].width = min(max(12, max_length + 2), 48)
    workbook.save(output_path)


def write_workbook_file(output_path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            export_df = df.copy()
            for column in export_df.columns:
                if pd.api.types.is_datetime64_any_dtype(export_df[column]):
                    export_df[column] = export_df[column].dt.strftime("%Y-%m-%d")
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
    autosize_workbook(output_path)


def fallback_output_path(output_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
    index = 1
    while candidate.exists():
        index += 1
        candidate = output_path.with_name(f"{output_path.stem}_{timestamp}_{index}{output_path.suffix}")
    return candidate


def export_workbook(output_path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        write_workbook_file(output_path, sheets)
        return output_path
    except PermissionError:
        generated_exports_dir = ROOT / "generated_exports"
        generated_exports_dir.mkdir(parents=True, exist_ok=True)
        fallback_seed = generated_exports_dir / output_path.name
        fallback_path = fallback_output_path(fallback_seed)
        write_workbook_file(fallback_path, sheets)
        return fallback_path


def relative_path_text(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def build_timeline_summary_payload(
    paths: Paths,
    qualifying_df: pd.DataFrame,
    approval_evidence_universe_df: pd.DataFrame,
    approval_evidence_completed_df: pd.DataFrame,
    parts_evidence_universe_df: pd.DataFrame,
    repair_evidence_universe_df: pd.DataFrame,
    approval_segment_df: pd.DataFrame,
    parts_segment_df: pd.DataFrame,
    repairer_segment_df: pd.DataFrame,
    anomalies_df: pd.DataFrame,
    start_date: date,
    end_date: date | None,
    workbook_path: Path,
) -> dict[str, Any]:
    created_filter = f">= {start_date.isoformat()}" if end_date is None else f"{start_date.isoformat()} to {end_date.isoformat()}"
    approval_values = valid_duration_values(approval_evidence_completed_df.get("approval_days", []))
    approval_avg = round(sum(approval_values) / len(approval_values), 2) if approval_values else None
    approval_threshold = int(round(approval_avg)) if approval_avg is not None else 27
    parts_avg = mean_or_blank(parts_segment_df.get("parts_issuing_days", []))
    repair_avg = mean_or_blank(repairer_segment_df.get("repairer_repair_days", []))
    process_values = [value for value in [approval_avg, parts_avg, repair_avg] if value is not None]
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "year": start_date.year,
        "scope": {
            "createdOnFilter": created_filter,
            "qualificationFilter": "Created On in scope + matched PO + Claim Approved On present + not unapproved",
            "denominatorRule": "Each segment uses its own valid rows only; missing dates for a segment are excluded from that segment denominator.",
        },
        "source": {
            "workbookPath": relative_path_text(workbook_path),
            "analysisTicketJs": relative_path_text(paths.analysis_ticket_js),
            "partsClassifiedCsv": relative_path_text(paths.parts_classified_csv),
            "repairersJson": relative_path_text(paths.repairers_json),
        },
        "totals": {
            "qualifyingTickets": int(len(qualifying_df)),
            "averageProcessDays": round(sum(process_values), 2) if process_values else None,
            "anomalyRows": int(len(anomalies_df)),
        },
        "stages": [
            {
                "key": "approval",
                "avg": approval_avg,
                "median": median_or_blank(approval_values),
                "count": int(len(approval_values)),
                "evidence": build_timeline_evidence(approval_evidence_completed_df, approval_evidence_universe_df, "approval_days", approval_threshold, start_date.year, date.today()),
            },
            {
                "key": "parts",
                "avg": parts_avg,
                "median": median_or_blank(parts_segment_df.get("parts_issuing_days", [])),
                "count": int(len(parts_segment_df)),
                "evidence": build_timeline_evidence(parts_segment_df, parts_evidence_universe_df, "parts_issuing_days", 21, start_date.year, date.today()),
            },
            {
                "key": "repair",
                "avg": repair_avg,
                "median": median_or_blank(repairer_segment_df.get("repairer_repair_days", [])),
                "count": int(len(repairer_segment_df)),
                "evidence": build_timeline_evidence(repairer_segment_df, repair_evidence_universe_df, "repairer_repair_days", 33, start_date.year, date.today()),
            },
            {
                "key": "closed",
                "avg": None,
                "median": None,
                "count": int(len(repairer_segment_df)),
            },
        ],
    }


def write_timeline_summary_json(path: Path, payload: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def yearly_summary_json_path(path: Path, year: int) -> Path:
    return path.with_name(f"{path.stem}_{year}{path.suffix}")


def try_generate_price_mix_ppt(script_path: Path = DEFAULT_PRICE_MIX_PPT_SCRIPT) -> Path | None:
    if not script_path.exists():
        print(f"Price mix PPT skipped: generator not found at {script_path}")
        return None
    node_path = shutil.which("node")
    if not node_path:
        print("Price mix PPT skipped: node executable was not found.")
        return None
    try:
        result = subprocess.run(
            [node_path, str(script_path)],
            cwd=ROOT,
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except Exception as exc:
        print(f"Price mix PPT skipped: {exc}")
        return None
    if result.stdout:
        print(result.stdout.strip())
    if result.returncode != 0:
        print(f"Price mix PPT skipped: generator exited with {result.returncode}")
        if result.stderr:
            print(result.stderr.strip())
        return None
    return ROOT / "generated_exports" / "ticket_timeline_price_mix_2026.pptx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export Ticket Timeline segment ticket sheets for approved PO-backed tickets created in the selected year."
    )
    parser.add_argument("--analysis-ticket-js", default=str(DEFAULT_ANALYSIS_TICKET_JS))
    parser.add_argument("--failure-timing-csv", default=str(DEFAULT_FAILURE_TIMING_CSV))
    parser.add_argument("--parts-classified-csv", default=str(DEFAULT_PARTS_CLASSIFIED_CSV))
    parser.add_argument("--repairers-json", default=str(DEFAULT_REPAIRERS_JSON))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--summary-json", default=str(DEFAULT_SUMMARY_JSON))
    parser.add_argument("--completion-analytics-json", default=str(DEFAULT_COMPLETION_ANALYTICS_JSON))
    parser.add_argument("--start-date", default=DEFAULT_START_DATE.isoformat())
    parser.add_argument("--end-date", default="")
    return parser.parse_args()


def validate_paths(paths: Paths) -> None:
    missing = [path for path in [paths.analysis_ticket_js, paths.parts_classified_csv, paths.repairers_json] if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required source files:\n" + "\n".join(str(path) for path in missing))


def main() -> int:
    args = parse_args()
    paths = Paths(
        analysis_ticket_js=Path(args.analysis_ticket_js),
        failure_timing_csv=Path(args.failure_timing_csv),
        parts_classified_csv=Path(args.parts_classified_csv),
        repairers_json=Path(args.repairers_json),
        output=Path(args.output),
        summary_json=Path(args.summary_json),
        completion_analytics_json=Path(args.completion_analytics_json),
    )
    validate_paths(paths)
    start_date = parse_date(args.start_date)
    end_date = parse_date(args.end_date) if clean(args.end_date) else None
    if start_date is None:
        raise ValueError(f"Invalid --start-date: {args.start_date}")

    base_df = load_analysis_ticket_base(paths.analysis_ticket_js)
    qualifying_df = build_qualifying_universe(base_df, start_date=start_date, end_date=end_date)
    evidence_date_mask = base_df["created_date"].map(lambda value: value is not None and value >= start_date and (end_date is None or value <= end_date))
    approval_evidence_universe_df = base_df[
        evidence_date_mask
        & base_df.apply(should_include_approval_evidence_universe, axis=1)
    ].copy()
    approval_evidence_universe_df["approval_days"] = approval_evidence_universe_df.apply(lambda row: day_diff(row["approved_date"], row["created_date"]), axis=1)
    approval_evidence_completed_df = approval_evidence_universe_df[
        approval_evidence_universe_df["approval_days"].map(is_valid_duration_value)
    ].copy()

    parts_by_ticket_df, parts_by_po_df = load_parts_ticket_summary(paths.parts_classified_csv)
    invoice_df = load_repairer_invoice_rows(paths.repairers_json)

    anomalies: list[dict[str, Any]] = []

    approval_segment_df = qualifying_df[qualifying_df["approval_days_valid"]].copy()

    for _, row in qualifying_df[~qualifying_df["approval_days_valid"]].iterrows():
        anomalies.append(
            anomaly_row(
                stage="Warranty Approval",
                reason="missing or negative approval days",
                row=row,
            )
        )

    parts_joined_df = merge_parts_data(qualifying_df, parts_by_ticket_df, parts_by_po_df, anomalies)
    parts_evidence_universe_df = parts_joined_df.copy()
    parts_segment_df = build_parts_segment(parts_joined_df, anomalies)
    repair_evidence_universe_df = parts_segment_df.copy()
    repairer_segment_df = build_repairer_segment(parts_segment_df, invoice_df, anomalies)
    anomalies_df = pd.DataFrame(anomalies).sort_values(["stage", "reason", "ticket_number", "ticket_key"], na_position="last").reset_index(drop=True)

    summary_df = build_summary_sheet(
        qualifying_df=qualifying_df,
        approval_segment_df=approval_segment_df,
        parts_segment_df=parts_segment_df,
        repairer_segment_df=repairer_segment_df,
        anomalies_df=anomalies_df,
        start_date=start_date,
        end_date=end_date,
    )
    qualifying_sheet_df = build_qualifying_sheet(
        qualifying_df=qualifying_df,
        approval_segment_df=approval_segment_df,
        parts_segment_df=parts_segment_df,
        repairer_segment_df=repairer_segment_df,
    )
    completion_analytics_payload, completion_detail_df, total_handling_detail_df = build_completion_analytics_payload(
        base_df=base_df,
        parts_by_ticket_df=parts_by_ticket_df,
        parts_by_po_df=parts_by_po_df,
        invoice_df=invoice_df,
        year=start_date.year,
        as_of=date.today(),
    )

    sheets = [
        ("summary", summary_df),
        ("Timeline Tickets Detail", qualifying_sheet_df),
        ("All Page Tickets Detail", completion_detail_df),
        ("total_handling", total_handling_detail_df),
        ("qualifying_tickets", qualifying_sheet_df),
        ("warranty_approval", prepare_approval_sheet(approval_segment_df)),
        ("parts_issuing", prepare_parts_sheet(parts_segment_df)),
        ("repairer_time", prepare_repairer_sheet(repairer_segment_df)),
        ("formula_breakdown", prepare_formula_breakdown_sheet(repairer_segment_df)),
        ("anomalies", anomalies_df),
    ]
    written_path = export_workbook(paths.output, sheets)
    summary_payload = build_timeline_summary_payload(
        paths=paths,
        qualifying_df=qualifying_df,
        approval_evidence_universe_df=approval_evidence_universe_df,
        approval_evidence_completed_df=approval_evidence_completed_df,
        parts_evidence_universe_df=parts_evidence_universe_df,
        repair_evidence_universe_df=repair_evidence_universe_df,
        approval_segment_df=approval_segment_df,
        parts_segment_df=parts_segment_df,
        repairer_segment_df=repairer_segment_df,
        anomalies_df=anomalies_df,
        start_date=start_date,
        end_date=end_date,
        workbook_path=written_path,
    )
    written_summary_path = write_timeline_summary_json(paths.summary_json, summary_payload)
    written_yearly_summary_path = write_timeline_summary_json(
        yearly_summary_json_path(paths.summary_json, start_date.year),
        summary_payload,
    )
    written_completion_analytics_path = write_timeline_summary_json(paths.completion_analytics_json, completion_analytics_payload)
    written_price_mix_ppt_path = try_generate_price_mix_ppt()

    print(f"Workbook written: {written_path}")
    print(f"Summary JSON written: {written_summary_path}")
    print(f"Year summary JSON written: {written_yearly_summary_path}")
    print(f"Completion analytics JSON written: {written_completion_analytics_path}")
    if written_price_mix_ppt_path:
        print(f"Price mix PPT written: {written_price_mix_ppt_path}")
    print(f"Qualifying tickets: {len(qualifying_df)}")
    print(f"Warranty approval rows: {len(approval_segment_df)}")
    print(f"Parts issuing rows: {len(parts_segment_df)}")
    print(f"Repairer rows: {len(repairer_segment_df)}")
    print(f"Anomaly rows: {len(anomalies_df)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
