import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT_JSON = ROOT / "outputs" / "repairers_2026" / "sap_authoritative_repair_payments.json"
OUT_XLSX = ROOT / "outputs" / "repairers_2026" / "sap_authoritative_repair_payments.xlsx"
FALLBACK_OUT_XLSX = (
    ROOT / "outputs" / "repairers_2026" / "sap_authoritative_repair_payments_with_cancelled.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="14213D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
META_FILL = PatternFill("solid", fgColor="EAF3FF")
META_FONT = Font(color="14213D")
TITLE_FONT = Font(bold=True, size=16, color="14213D")
THIN_HEADER = Side(style="thin", color="C8D3E6")
THIN_BODY = Side(style="thin", color="E4EAF4")
HEADER_BORDER = Border(left=THIN_HEADER, right=THIN_HEADER, top=THIN_HEADER, bottom=THIN_HEADER)
BODY_BORDER = Border(left=THIN_BODY, right=THIN_BODY, top=THIN_BODY, bottom=THIN_BODY)


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def matrix_headers(rows: list[dict[str, Any]], preferred_headers: list[str]) -> list[str]:
    seen = set()
    headers = []
    for header in preferred_headers:
        if header not in seen:
            headers.append(header)
            seen.add(header)
    for row in rows:
        for key in (row or {}).keys():
            if key not in seen:
                headers.append(key)
                seen.add(key)
    return headers or ["No Data"]


def width_for_header(header: str) -> int:
    lowered = header.lower()
    if any(token in lowered for token in ("short text", "rule", "repairer", "invoice docs", "pos", "compare", "match")):
        return 32
    if any(token in lowered for token in ("ticket id", "vendor", "date", "currency", "po", "po item", "invoice doc")):
        return 18
    if any(token in lowered for token in ("count", "qty", "line")):
        return 14
    return 20


def is_money_header(header: str) -> bool:
    lowered = header.lower()
    return any(token in lowered for token in ("amount", "cost", "value", "price"))


def style_sheet(ws, headers: list[str], row_count: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_count, 1)}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = HEADER_BORDER
    for row in ws.iter_rows(min_row=2, max_row=row_count, max_col=len(headers)):
        for cell in row:
            cell.border = BODY_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = width_for_header(header)
        if is_money_header(header):
            for cell in ws[letter][1:]:
                cell.number_format = '$#,##0.00'


def write_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]], preferred_headers: list[str]) -> None:
    ws = wb.create_sheet(name)
    safe_rows = rows if rows else [{}]
    headers = matrix_headers(safe_rows, preferred_headers)
    ws.append(headers)
    for row in safe_rows:
        ws.append([normalize_cell((row or {}).get(header)) for header in headers])
    style_sheet(ws, headers, ws.max_row)


def write_meta(wb: Workbook, meta_rows: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws["A1"] = "SAP Authoritative Repair PO Costs"
    ws["A2"] = (
        "C4C is used only to decide repairer-analysis eligibility and remove customer/self repairs. "
        "SAP is authoritative for PO date, ticket ID parsed from PO short text, PO net value, "
        "cancellation status, and repairer/vendor. Invoice fields are status/completion info only."
    )
    for row in ws["A1:D2"]:
        for cell in row:
            cell.fill = META_FILL
            cell.font = META_FONT
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = TITLE_FONT
    ws.append([])
    ws.append(["Metric", "Value"])
    for row in meta_rows:
        ws.append([normalize_cell(row.get("Metric")), normalize_cell(row.get("Value"))])
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = BODY_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A5"


def save_workbook(wb: Workbook) -> Path:
    try:
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except PermissionError:
        wb.save(FALLBACK_OUT_XLSX)
        return FALLBACK_OUT_XLSX


def main() -> int:
    payload = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    wb = Workbook()
    write_meta(wb, payload.get("meta") or [])
    write_sheet(wb, "SAP Ticket Summary", payload.get("summary") or [], [
        "SAP Ticket ID",
        "SAP First PO Date",
        "SAP Last PO Date",
        "SAP Last Invoice Date",
        "SAP Repairer Name",
        "SAP Repairer Vendor ID",
        "SAP PO Net Value",
        "SAP Cancelled PO Net Value",
        "SAP PO Cancelled",
        "SAP Invoice Status",
        "SAP Signed Invoice Amount",
        "SAP Currency",
        "SAP Line Count",
        "SAP Invoice Row Count",
        "SAP Invoice Docs",
        "SAP POs",
        "C4C Eligible For Repairer Analysis",
        "C4C Eligibility Match Source",
        "C4C Eligibility Reason",
        "C4C Service Technician",
        "C4C Status",
        "C4C PO",
        "C4C Chassis Number",
        "C4C Repair Detail Exists",
        "C4C Repair Detail Repairer",
        "C4C Repair Detail PO",
        "C4C Repair Detail Approved Cost",
    ])
    po_lines = payload.get("po_lines") or payload.get("invoice_lines") or []
    write_sheet(wb, "SAP PO Lines", po_lines, [
        "SAP Ticket ID",
        "SAP PO Date",
        "SAP Last Invoice Date",
        "SAP Last Invoice Doc",
        "SAP PO",
        "SAP PO Item",
        "SAP Repairer Name",
        "SAP Repairer Vendor ID",
        "SAP PO Net Value",
        "SAP Active PO Net Value",
        "SAP Cancelled PO Net Value",
        "SAP PO Cancelled",
        "SAP PO Header Deletion Indicator",
        "SAP PO Item Deletion Indicator",
        "SAP PO Currency",
        "SAP Invoice Status",
        "SAP Signed Invoice Amount Doc",
        "SAP Short Text",
        "C4C Eligible For Repairer Analysis",
        "C4C Eligibility Match Source",
        "C4C Eligibility Reason",
        "C4C Service Technician",
        "C4C Status",
        "C4C PO",
        "C4C Chassis Number",
        "C4C Repair Detail Exists",
        "C4C Repair Detail Repairer",
        "C4C Repair Detail PO",
        "C4C Repair Detail Approved Cost",
    ])
    write_sheet(wb, "C4C Ineligible Excluded", payload.get("c4c_ineligible") or [], [
        "SAP Ticket ID",
        "SAP PO Date",
        "SAP PO",
        "SAP PO Item",
        "SAP Repairer Name",
        "SAP Repairer Vendor ID",
        "SAP PO Net Value",
        "SAP PO Cancelled",
        "SAP Short Text",
        "C4C Eligibility Match Source",
        "C4C Eligibility Reason",
        "C4C Service Technician",
        "C4C Status",
        "C4C PO",
        "C4C Chassis Number",
    ])
    write_sheet(wb, "Short Text Unreadable", payload.get("short_text_unreadable") or [], [
        "SAP PO",
        "SAP PO Item",
        "SAP PO Date",
        "SAP Last Invoice Doc",
        "SAP Last Invoice Date",
        "SAP Repairer Name",
        "SAP PO Net Value",
        "SAP PO Cancelled",
        "SAP PO Currency",
        "SAP Short Text",
        "Short Text Parse Note",
    ])
    write_sheet(wb, "C4C Compare", payload.get("c4c_compare") or [], [
        "SAP Ticket ID",
        "SAP Last PO Date",
        "SAP Last Invoice Date",
        "SAP Repairer Name",
        "C4C Repair Detail Repairer",
        "Repairer Match",
        "SAP POs",
        "C4C Repair Detail PO",
        "Any SAP PO Equals C4C PO",
        "SAP PO Net Value",
        "C4C Repair Detail Approved Cost",
        "SAP Minus C4C Approved Cost",
        "C4C Repair Detail Exists",
    ])

    saved_path = save_workbook(wb)
    print(json.dumps({
        "output": str(saved_path),
        "summary": len(payload.get("summary") or []),
        "po_lines": len(po_lines),
        "c4c_ineligible": len(payload.get("c4c_ineligible") or []),
        "unreadable": len(payload.get("short_text_unreadable") or []),
        "compare": len(payload.get("c4c_compare") or []),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
