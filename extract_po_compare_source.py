import json
from datetime import date, datetime, time

import openpyxl


INPUT = r"C:\Users\Leo.Li\Downloads\claim_ytd_comparison_tickets_detail_20260709_135559.xlsx"
OUTPUT = "po_compare_source.json"
APPROVED_COST_CACHE = "outputs/analysis_approved_cost_by_ticket.json"


def clean(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_ticket_number(value):
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits


def load_approved_cost_map():
    try:
        with open(APPROVED_COST_CACHE, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        return {}
    by_ticket = payload.get("byTicket") if isinstance(payload, dict) else {}
    return by_ticket if isinstance(by_ticket, dict) else {}


def approved_cost_for_ticket(by_ticket, ticket_id):
    ticket = normalize_ticket_number(ticket_id)
    if not ticket:
        return None
    rec = by_ticket.get(f"ticket_{ticket}") or by_ticket.get(ticket) or {}
    if not isinstance(rec, dict):
        return None
    cost = number(rec.get("amount") if rec.get("amount") not in (None, "") else rec.get("netOrderValue"))
    return cost if cost and cost > 0 else None


wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
tickets = wb["Tickets"]
sheet1 = wb["Sheet1"]
approved_cost_by_ticket = load_approved_cost_map()

ticket_map = {}
for row in tickets.iter_rows(min_row=5, values_only=True):
    ticket_id = row[0]
    if ticket_id in (None, ""):
        continue
    ticket_map[ticket_id] = {
        "claim_scope": clean(row[1]),
        "status_group": clean(row[2]),
        "status_text": clean(row[3]),
        "claim_approved_on": clean(row[6]),
        "dealer_name": clean(row[9]),
        "ticket_type": clean(row[10]),
        "ticket_type_text": clean(row[11]),
        "created_date": clean(row[12]),
        "approved_date": clean(row[13]),
        "resolved_date": clean(row[14]),
        "amount_including_tax": number(row[8]),
        "amount_value": number(row[15]),
    }

rows = []
for row in sheet1.iter_rows(min_row=2, values_only=True):
    ticket_id = row[0]
    approved_cost = approved_cost_for_ticket(approved_cost_by_ticket, ticket_id)
    po_price = approved_cost if approved_cost is not None else number(row[4])
    if ticket_id in (None, "") or po_price is None:
        continue
    ticket = ticket_map.get(ticket_id, {})
    rows.append(
        {
            "ticket": ticket_id,
            "po_price": po_price,
            "po_price_source": "analysis_approved_cost_by_ticket" if approved_cost is not None else "workbook_sheet1",
            "amount_including_tax": ticket.get("amount_including_tax"),
            "status_group": ticket.get("status_group"),
            "status_text": ticket.get("status_text"),
            "claim_scope": ticket.get("claim_scope"),
            "claim_approved_on": ticket.get("claim_approved_on"),
            "dealer_name": ticket.get("dealer_name"),
            "ticket_type": ticket.get("ticket_type"),
            "ticket_type_text": ticket.get("ticket_type_text"),
            "created_date": ticket.get("created_date"),
            "approved_date": ticket.get("approved_date"),
            "resolved_date": ticket.get("resolved_date"),
            "po_date": clean(row[3]),
            "amount_value": ticket.get("amount_value"),
        }
    )

with open(OUTPUT, "w", encoding="utf-8") as handle:
    json.dump(rows, handle, ensure_ascii=False, indent=2)

print(json.dumps({"rows_with_po": len(rows)}, ensure_ascii=False))
