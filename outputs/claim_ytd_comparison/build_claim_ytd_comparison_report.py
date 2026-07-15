from __future__ import annotations

import argparse
import csv
import math
import os
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import firebase_admin
    from firebase_admin import credentials, db
except ImportError:  # pragma: no cover
    firebase_admin = None
    credentials = None
    db = None

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
OUTPUT_DIR = SCRIPT_DIR

DEFAULT_FIREBASE_DB_URL = os.getenv(
    "FIREBASE_DB_URL",
    "https://snowy-hr-report-default-rtdb.asia-southeast1.firebasedatabase.app",
)
DEFAULT_FIREBASE_SA_PATH = os.getenv(
    "FIREBASE_SA_PATH",
    str(REPO_ROOT / "firebase-service-account.json"),
)
DEFAULT_FIREBASE_ROOT = os.getenv("FIREBASE_ROOT", "c4cTickets_test")
DEFAULT_LOCAL_CSV = os.getenv(
    "CLAIM_COMPARISON_CSV",
    str(REPO_ROOT / "SAPAnalyticsReport_ZF8C06456D7698BCB54F44D_.csv"),
)
DEFAULT_DETAILS_PREFIX = "claim_ytd_comparison_tickets_detail"

MONTH_LABELS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]

DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%dT%H:%M:%S.%fZ",
    "%d.%m.%Y",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M:%S AUSACT",
)


@dataclass(frozen=True)
class MetricSpec:
    key: str
    title: str
    period_kind: str
    value_kind: str
    description: str
    accent: str


METRICS = (
    MetricSpec(
        key="claims_received",
        title="Claims received",
        period_kind="created",
        value_kind="count",
        description="Use CreatedOn as the ticket intake date.",
        accent="2F80ED",
    ),
    MetricSpec(
        key="claims_closed",
        title="Claims closed",
        period_kind="resolved",
        value_kind="count",
        description="Count approved closed + unapproved closed tickets by ResolvedOnDateTime.",
        accent="27AE60",
    ),
    MetricSpec(
        key="amount_approved",
        title="Amount approved",
        period_kind="approved_amount",
        value_kind="amount",
        description="Sum tickets with ClaimApprovedOnDateTime whose current status is in Z9/Y0/Y1/Y2/Y4/YB, using approvedAmount from fetch.",
        accent="F2994A",
    ),
)

APPROVED_STATUS_CODES = {"Z9", "Y0", "Y1", "Y2", "Y4", "YB"}
APPROVED_STATUS_TEXTS = {
    "sales order approved",
    "partially picked",
    "dispatch parts",
    "repair in progress",
    "repairer invoiced received",
    "repairer invoiced processed",
}
APPROVED_CLOSED_STATUS_CODES = {"Y7"}
APPROVED_CLOSED_STATUS_TEXTS = {"approved claims closed (closed)"}
UNAPPROVED_CLOSED_STATUS_CODES = {"Y8"}
UNAPPROVED_CLOSED_STATUS_TEXTS = {"unapproved claims closed (closed)"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def sql_quote(value: Any) -> str:
    return str(value).replace("'", "''")


def first_non_blank(*values: Any) -> str:
    for value in values:
        text = clean(value)
        if text:
            return text
    return ""


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip().lower()


def classify_claim_scope(ticket_type: Any, ticket_type_text: Any = None) -> str:
    text = normalize_text(first_non_blank(ticket_type_text, ticket_type))
    if not text:
        return ""
    if any(token in text for token in ("pre delivery", "pre-delivery", "predelivery", "pdi")):
        return "Pre Delivery"
    if any(token in text for token in ("in field", "field warranty")):
        return "In Field"
    code = normalize_text(ticket_type)
    if code == "z005":
        return "Pre Delivery"
    if code == "z006":
        return "In Field"
    return ""


def parse_any_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean(value)
    if not text or text == "00000000":
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    try:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
        if not pd.isna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None


def parse_amount(value: Any) -> float:
    text = clean(value)
    if not text:
        return 0.0
    text = text.replace(",", "")
    text = text.replace("$", "")
    text = text.replace("AUD", "").strip()
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if match:
            return float(match.group(0))
    return 0.0


def iso_date(value: Any) -> str:
    parsed = parse_any_date(value)
    return parsed.isoformat() if parsed else ""


def firebase_init(db_url: str, sa_path: str) -> None:
    if firebase_admin is None or credentials is None or db is None:
        raise RuntimeError("firebase_admin is not installed in this environment.")
    if getattr(firebase_admin, "_apps", None):
        return
    if not Path(sa_path).exists():
        raise FileNotFoundError(f"Firebase service account file not found: {sa_path}")
    cred = credentials.Certificate(sa_path)
    firebase_admin.initialize_app(cred, {"databaseURL": db_url})


def firebase_node_to_dict(node: Any) -> Dict[str, Any]:
    if isinstance(node, dict):
        return node
    if isinstance(node, list):
        return {str(i): value for i, value in enumerate(node) if value is not None}
    return {}


def classify_status_group(ticket: Dict[str, Any]) -> str:
    status_text = normalize_text(
        first_non_blank(
            ticket.get("TicketStatusText"),
            ticket.get("Status"),
            ticket.get("TicketStatus"),
        )
    )
    status_code = normalize_text(
        first_non_blank(
            ticket.get("TicketStatus"),
            ticket.get("TicketStatusCode"),
            ticket.get("StatusCode"),
        )
    )

    combined = f"{status_code} {status_text}".strip()
    if status_code in UNAPPROVED_CLOSED_STATUS_CODES or status_text in UNAPPROVED_CLOSED_STATUS_TEXTS or "unapproved claims closed" in combined:
        return "unapproved_closed"
    if status_code in APPROVED_CLOSED_STATUS_CODES or status_text in APPROVED_CLOSED_STATUS_TEXTS or "approved claims closed" in combined:
        return "approved_closed"
    if status_code in APPROVED_STATUS_CODES or status_text in APPROVED_STATUS_TEXTS:
        return "approved"
    return ""


def load_tickets_from_firebase(firebase_root: str, firebase_db_url: str, firebase_sa_path: str) -> pd.DataFrame:
    firebase_init(firebase_db_url, firebase_sa_path)
    raw = db.reference(f"{firebase_root}/tickets").get()
    tickets = firebase_node_to_dict(raw)
    rows: List[Dict[str, Any]] = []

    for ticket_id, node in tickets.items():
        ticket = (node or {}).get("ticket", {}) if isinstance(node, dict) else {}
        if not isinstance(ticket, dict):
            continue

        status_group = classify_status_group(ticket)
        rows.append(
            {
                "TicketID": first_non_blank(ticket_id, ticket.get("TicketID")),
                "StatusGroup": status_group,
                "StatusText": first_non_blank(ticket.get("TicketStatusText"), ticket.get("Status"), ticket.get("TicketStatus")),
                "StatusCode": first_non_blank(ticket.get("TicketStatus"), ticket.get("TicketStatusCode"), ticket.get("StatusCode")),
                "CreatedOn": first_non_blank(ticket.get("CreatedOn"), ticket.get("Created On"), ticket.get("createdOn"), ticket.get("CreatedAt")),
                "ClaimApprovedOnDateTime": first_non_blank(
                    ticket.get("ClaimApprovedOnDateTime"),
                    ticket.get("Claim Approved On"),
                    ticket.get("ClaimApprovedOnDate"),
                    ticket.get("ClaimApprovedOn"),
                ),
                "ClaimApprovedOn": first_non_blank(
                    ticket.get("Claim Approved On"),
                    ticket.get("ClaimApprovedOnDateTime"),
                    ticket.get("ClaimApprovedOnDate"),
                    ticket.get("ClaimApprovedOn"),
                    ticket.get("ApprovalDate"),
                ),
                "ResolvedOn": first_non_blank(
                    ticket.get("ResolvedOnDateTime"),
                    ticket.get("ResolvedOnDate"),
                    ticket.get("ResolvedOn"),
                    ticket.get("Resolved On"),
                ),
                "ERPPurchaseOrder": first_non_blank(
                    ticket.get("ERPPurchaseOrder"),
                    ticket.get("ERP Purchase Order ID"),
                    ticket.get("Purchasing Document"),
                ),
                "NetValue": first_non_blank(
                    ticket.get("Net Value"),
                    ticket.get("NetValue"),
                ),
                "AmountIncludingTax": first_non_blank(
                    ticket.get("AmountIncludingTax"),
                    ticket.get("Amount Including Tax"),
                    ticket.get("ClaimTotalAmount"),
                ),
                "ApprovedAmount": first_non_blank(
                    ticket.get("approvedAmount"),
                    ticket.get("ApprovedAmount"),
                ),
                "ApprovedAmountSource": first_non_blank(
                    ticket.get("approvedAmountSource"),
                    ticket.get("ApprovedAmountSource"),
                ),
                "DealerName": first_non_blank(ticket.get("DealerName"), ticket.get("Dealer Name")),
                "TicketType": first_non_blank(ticket.get("TicketType"), ticket.get("Ticket Type")),
                "TicketTypeText": first_non_blank(ticket.get("TicketTypeText"), ticket.get("Ticket Type Text")),
                "ClaimScope": classify_claim_scope(
                    first_non_blank(ticket.get("TicketType"), ticket.get("Ticket Type")),
                    first_non_blank(ticket.get("TicketTypeText"), ticket.get("Ticket Type Text")),
                ),
            }
        )

    return pd.DataFrame(rows)


def load_tickets_from_csv(csv_path: str) -> pd.DataFrame:
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"Local CSV not found: {csv_path}")

    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    cols = {re.sub(r"\s+", " ", str(col)).strip().lower(): col for col in df.columns}

    def pick(row: pd.Series, *names: str) -> str:
        for name in names:
            col = cols.get(re.sub(r"\s+", " ", name).strip().lower())
            if col:
                value = clean(row.get(col))
                if value:
                    return value
        return ""

    rows: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        ticket = {
            "TicketStatusText": pick(row, "Status", "TicketStatusText", "Ticket Status"),
            "TicketStatus": pick(row, "TicketStatus", "Ticket Status"),
        }
        rows.append(
            {
                "TicketID": pick(row, "Ticket ID", "TicketID"),
                "StatusGroup": classify_status_group(ticket),
                "StatusText": pick(row, "Status", "TicketStatusText", "Ticket Status"),
                "StatusCode": pick(row, "TicketStatus", "Ticket Status"),
                "CreatedOn": pick(row, "Created On", "CreatedOn"),
                "ClaimApprovedOnDateTime": pick(row, "ClaimApprovedOnDateTime", "Claim Approved On", "ClaimApprovedOnDate", "ClaimApprovedOn"),
                "ClaimApprovedOn": pick(row, "Claim Approved On", "ClaimApprovedOnDateTime", "ClaimApprovedOn"),
                "ResolvedOn": pick(row, "ResolvedOnDateTime", "Resolved On", "ResolvedOn"),
                "ERPPurchaseOrder": pick(row, "ERPPurchaseOrder", "ERP Purchase Order ID", "Purchasing Document"),
                "NetValue": pick(row, "Net Value", "NetValue"),
                "AmountIncludingTax": pick(row, "Amount Including Tax", "ClaimTotalAmount"),
                "ApprovedAmount": pick(row, "approvedAmount", "ApprovedAmount", "Approved Amount"),
                "ApprovedAmountSource": pick(row, "approvedAmountSource", "ApprovedAmountSource"),
                "DealerName": pick(row, "Dealer Name", "Dealer"),
                "TicketType": pick(row, "Ticket Type", "TicketType"),
                "TicketTypeText": pick(row, "Ticket Type Text", "TicketTypeText"),
                "ClaimScope": classify_claim_scope(
                    pick(row, "Ticket Type", "TicketType"),
                    pick(row, "Ticket Type Text", "TicketTypeText"),
                ),
            }
        )
    return pd.DataFrame(rows)


def normalize_dataset(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "TicketID",
                "StatusGroup",
                "StatusText",
                "StatusCode",
                "CreatedOn",
                "ClaimApprovedOnDateTime",
                "ClaimApprovedOn",
                "ERPPurchaseOrder",
                "NetValue",
                "ResolvedOn",
                "AmountIncludingTax",
                "ApprovedAmount",
                "ApprovedAmountSource",
                "DealerName",
                "TicketType",
                "TicketTypeText",
                "ClaimScope",
                "CreatedDate",
                "ApprovedDate",
                "ResolvedDate",
                "ApprovedAmountValue",
                "NetValueAmount",
                "ApprovedAmountEligible",
            ]
        )

    out = df.copy()
    for col in ["TicketID", "StatusGroup", "StatusText", "StatusCode", "CreatedOn", "ClaimApprovedOnDateTime", "ClaimApprovedOn", "ERPPurchaseOrder", "NetValue", "ResolvedOn", "AmountIncludingTax", "ApprovedAmount", "ApprovedAmountSource", "DealerName", "TicketType", "TicketTypeText", "ClaimScope"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    out["CreatedDate"] = out["CreatedOn"].map(parse_any_date)
    out["ApprovedDate"] = out["ClaimApprovedOnDateTime"].map(parse_any_date)
    out["ApprovedDate"] = out["ApprovedDate"].where(out["ApprovedDate"].notna(), out["ClaimApprovedOn"].map(parse_any_date))
    out["ResolvedDate"] = out["ResolvedOn"].map(parse_any_date)
    out["NetValueAmount"] = out["NetValue"].map(parse_amount)
    out["ApprovedAmountValue"] = out["ApprovedAmount"].map(parse_amount)
    out["ApprovedAmountEligible"] = out["ApprovedDate"].notna() & out["StatusGroup"].eq("approved")
    out["Received2025"] = out["CreatedDate"].map(lambda d: d is not None and d.year == 2025)
    out["Received2026YTD"] = out["CreatedDate"].map(lambda d: d is not None and d.year == 2026)
    out["Closed2025"] = out["ResolvedDate"].map(lambda d: d is not None and d.year == 2025)
    out["Closed2026YTD"] = out["ResolvedDate"].map(lambda d: d is not None and d.year == 2026)
    out["ApprovedAmount2025"] = out["ApprovedAmountEligible"] & out["ApprovedDate"].map(lambda d: d is not None and d.year == 2025)
    out["ApprovedAmount2026YTD"] = out["ApprovedAmountEligible"] & out["ApprovedDate"].map(lambda d: d is not None and d.year == 2026)
    return out


def date_window(year: int, today: date) -> Tuple[date, date]:
    start = date(year, 1, 1)
    end = date(year, 12, 31) if year < today.year else today
    return start, end


def count_in_window(df: pd.DataFrame, date_col: str, start: date, end: date, mask: Optional[pd.Series] = None) -> int:
    if df.empty:
        return 0
    work = df
    if mask is not None:
        work = work[mask]
    if work.empty:
        return 0
    dates = pd.to_datetime(work[date_col], errors="coerce")
    window_mask = dates.notna() & (dates.dt.date >= start) & (dates.dt.date <= end)
    return int(window_mask.sum())


def metric_total(df: pd.DataFrame, metric: MetricSpec, start: date, end: date) -> float:
    if df.empty:
        return 0.0

    if metric.key == "claims_received":
        return float(count_in_window(df, "CreatedDate", start, end))

    if metric.key == "claims_closed":
        mask = df["StatusGroup"].isin(["approved_closed", "unapproved_closed"])
        return float(count_in_window(df, "ResolvedDate", start, end, mask=mask))

    if metric.key == "amount_approved":
        approved_mask = df["ApprovedAmountEligible"].fillna(False)
        if not approved_mask.any():
            return 0.0
        work = df.loc[approved_mask].copy()
        dates = pd.to_datetime(work["ApprovedDate"], errors="coerce")
        valid = dates.notna() & (dates.dt.date >= start) & (dates.dt.date <= end)
        if not valid.any():
            return 0.0
        return float(pd.to_numeric(work.loc[valid, "ApprovedAmountValue"], errors="coerce").fillna(0).sum())

    return 0.0


def monthly_series(df: pd.DataFrame, metric: MetricSpec, year: int, today: date) -> Dict[int, float]:
    start, end = date_window(year, today)
    months = {m: 0.0 for m in range(1, 13)}
    if df.empty:
        return months

    if metric.key == "claims_received":
        dates = pd.to_datetime(df["CreatedDate"], errors="coerce")
        valid = dates.notna() & (dates.dt.date >= start) & (dates.dt.date <= end)
        for month, cnt in dates[valid].dt.month.value_counts().sort_index().items():
            months[int(month)] = float(cnt)
        return months

    if metric.key == "claims_closed":
        mask = df["StatusGroup"].isin(["approved_closed", "unapproved_closed"])
        dates = pd.to_datetime(df.loc[mask, "ResolvedDate"], errors="coerce")
        valid = dates.notna() & (dates.dt.date >= start) & (dates.dt.date <= end)
        for month, cnt in dates[valid].dt.month.value_counts().sort_index().items():
            months[int(month)] = float(cnt)
        return months

    if metric.key == "amount_approved":
        approved_mask = df["ApprovedAmountEligible"].fillna(False)
        approved_subset = df.loc[approved_mask].copy()
        approved_dates = pd.to_datetime(approved_subset["ApprovedDate"], errors="coerce")
        approved_valid = approved_dates.notna() & (approved_dates.dt.date >= start) & (approved_dates.dt.date <= end)
        approved_months = pd.DataFrame(
            {
                "month": approved_dates[approved_valid].dt.month,
                "amount": pd.to_numeric(approved_subset.loc[approved_valid, "ApprovedAmountValue"], errors="coerce").fillna(0),
            }
        )
        for month, value in approved_months.groupby("month")["amount"].sum().items():
            months[int(month)] += float(value)

        return months

    return months


def fmt_number(value: float, value_kind: str) -> str:
    if value_kind == "amount":
        return f"{value:,.2f}"
    return f"{int(round(value)):,.0f}"


def pct_change(new_value: float, old_value: float) -> Optional[float]:
    if old_value == 0:
        return None
    return (new_value - old_value) / old_value


def header_fill(accent: str) -> PatternFill:
    return PatternFill("solid", fgColor=accent)


def pastel_fill(accent: str) -> PatternFill:
    return PatternFill("solid", fgColor=f"F2F7FF" if accent == "2F80ED" else ("F1FBF4" if accent == "27AE60" else "FFF7ED"))


def apply_border(cell) -> None:
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_range(ws, cell_range: str, fill: Optional[PatternFill] = None, font: Optional[Font] = None, alignment: Optional[Alignment] = None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            apply_border(cell)


def auto_size(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        column_letter = get_column_letter(col_cells[0].column)
        widths = [len(clean(cell.value)) for cell in col_cells if cell.value is not None]
        width = min(max_width, max(min_width, max(widths, default=min_width) + 2))
        ws.column_dimensions[column_letter].width = width


def format_money_cell(cell) -> None:
    cell.number_format = '#,##0.00;[Red]-#,##0.00'


def format_count_cell(cell) -> None:
    cell.number_format = '#,##0'


def write_summary_sheet(wb: Workbook, summary_rows: List[Dict[str, Any]], today: date) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Claims YTD Comparison"
    ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"2025 full year vs 2026 YTD through {today.isoformat()}."
    ws["A2"].font = Font(size=10, color="475467")
    ws["A2"].alignment = Alignment(horizontal="left")

    headers = ["Metric", "2025 Total", "2026 YTD", "Change", "Change %", "Rule"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell)

    for idx, row in enumerate(summary_rows, start=5):
        accent = row["accent"]
        values = [
            row["metric"],
            row["value_2025"],
            row["value_2026"],
            row["delta"],
            row["pct_change"],
            row["description"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            apply_border(cell)
            if col == 1:
                cell.font = Font(bold=True, color="0F172A")
                cell.fill = pastel_fill(accent)
            elif col in (2, 3, 4):
                cell.number_format = '#,##0.00' if row["value_kind"] == "amount" else '#,##0'
            elif col == 5:
                cell.number_format = "0.0%"
            elif col == 6:
                cell.font = Font(color="475467")

            if col == 4:
                if isinstance(value, (int, float)) and value < 0:
                    cell.font = Font(color="C62828", bold=True)
                else:
                    cell.font = Font(color="1B5E20", bold=True)
            if col == 5:
                if value is None:
                    cell.value = "n/a"
                    cell.number_format = "General"
                elif value < 0:
                    cell.font = Font(color="C62828", bold=True)
                else:
                    cell.font = Font(color="1B5E20", bold=True)

    ws["A9"] = "Notes"
    ws["A9"].font = Font(bold=True, color="0F172A")
    ws["A10"] = "• Claims received uses CreatedOn."
    ws["A11"] = "• Claims closed uses ResolvedOnDateTime for approved closed and unapproved closed tickets."
    ws["A12"] = "• Amount approved uses ClaimApprovedOnDateTime and only current approved statuses Z9/Y0/Y1/Y2/Y4/YB."
    ws["A13"] = "• Amount approved is sourced from fetch approvedAmount (3111 PO)."
    for cell_ref in ("A10", "A11", "A12", "A13"):
        ws[cell_ref].font = Font(color="475467")

    auto_size(ws, min_width=14, max_width=52)
    ws.column_dimensions["F"].width = 54


def build_scope_summary_rows(df: pd.DataFrame, scope_name: str, today: date) -> List[Dict[str, Any]]:
    scope_df = df[df["ClaimScope"].eq(scope_name)].copy()
    return build_summary_rows(scope_df, today)


def write_scope_summary_sheet(wb: Workbook, df: pd.DataFrame, scope_name: str, today: date) -> None:
    ws = wb.create_sheet(scope_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"{scope_name} Comparison"
    ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"2025 full year vs 2026 YTD through {today.isoformat()} for {scope_name} tickets."
    ws["A2"].font = Font(size=10, color="475467")
    ws["A2"].alignment = Alignment(horizontal="left")

    summary_rows = build_scope_summary_rows(df, scope_name, today)
    headers = ["Metric", "2025 Total", "2026 YTD", "Change", "Change %", "Rule"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell)

    for idx, row in enumerate(summary_rows, start=5):
        accent = row["accent"]
        values = [
            row["metric"],
            row["value_2025"],
            row["value_2026"],
            row["delta"],
            row["pct_change"],
            row["description"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            apply_border(cell)
            if col == 1:
                cell.font = Font(bold=True, color="0F172A")
                cell.fill = pastel_fill(accent)
            elif col in (2, 3, 4):
                cell.number_format = '#,##0.00' if row["value_kind"] == "amount" else '#,##0'
            elif col == 5:
                cell.number_format = "0.0%"
            elif col == 6:
                cell.font = Font(color="475467")

            if col == 4:
                if isinstance(value, (int, float)) and value < 0:
                    cell.font = Font(color="C62828", bold=True)
                else:
                    cell.font = Font(color="1B5E20", bold=True)
            if col == 5:
                if value is None:
                    cell.value = "n/a"
                    cell.number_format = "General"
                elif value < 0:
                    cell.font = Font(color="C62828", bold=True)
                else:
                    cell.font = Font(color="1B5E20", bold=True)

    auto_size(ws, min_width=14, max_width=52)
    ws.column_dimensions["F"].width = 54


def build_monthly_table(metric: MetricSpec, df: pd.DataFrame, today: date) -> List[Dict[str, Any]]:
    series_2025 = monthly_series(df, metric, 2025, today)
    series_2026 = monthly_series(df, metric, 2026, today)
    rows: List[Dict[str, Any]] = []
    for month_idx, month_name in enumerate(MONTH_LABELS, start=1):
        val_2026 = series_2026[month_idx] if month_idx <= today.month else None
        rows.append(
            {
                "Month": month_name,
                "2025": series_2025[month_idx],
                "2026 YTD": val_2026,
            }
        )
    return rows


def render_monthly_table_section(
    ws,
    metric: MetricSpec,
    frame: pd.DataFrame,
    start_row: int,
    section_title: str,
    today: date,
) -> int:
    end_col = 3
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=1, value=section_title)
    title_cell.font = Font(size=12, bold=True, color="0F172A")
    title_cell.fill = pastel_fill(metric.accent)
    title_cell.alignment = Alignment(horizontal="left")
    apply_border(title_cell)

    monthly_rows = build_monthly_table(metric, frame, today)
    table_start = start_row + 1
    headers = ["Month", "2025", "2026 YTD"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        apply_border(cell)

    for row_idx, row in enumerate(monthly_rows, start=table_start + 1):
        ws.cell(row=row_idx, column=1, value=row["Month"])
        ws.cell(row=row_idx, column=2, value=row["2025"])
        ws.cell(row=row_idx, column=3, value=row["2026 YTD"])
        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            apply_border(cell)
            cell.alignment = Alignment(horizontal="center")
            if col == 1:
                cell.font = Font(bold=True, color="0F172A")
            elif metric.value_kind == "amount":
                format_money_cell(cell)
            else:
                format_count_cell(cell)

    return table_start + len(monthly_rows) + 2


def write_metric_sheet(wb: Workbook, metric: MetricSpec, df: pd.DataFrame, today: date) -> None:
    ws = wb.create_sheet(metric.title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    ws.merge_cells("A1:H1")
    ws["A1"] = metric.title
    ws["A1"].font = Font(size=17, bold=True, color="0F172A")

    ws.merge_cells("A2:H2")
    ws["A2"] = metric.description
    ws["A2"].font = Font(size=10, color="475467")

    ws["A4"] = "2025 Full Year"
    ws["B4"] = metric_total(df, metric, date(2025, 1, 1), date(2025, 12, 31))
    ws["C4"] = "2026 YTD"
    ws["D4"] = metric_total(df, metric, date(2026, 1, 1), today)
    ws["E4"] = "Change"
    ws["F4"] = ws["D4"].value - ws["B4"].value
    ws["G4"] = "Change %"
    change_pct = pct_change(ws["D4"].value, ws["B4"].value)
    ws["H4"] = change_pct if change_pct is not None else "n/a"

    for cell_ref in ("A4", "C4", "E4", "G4"):
        ws[cell_ref].font = Font(bold=True, color="FFFFFF")
        ws[cell_ref].fill = PatternFill("solid", fgColor=metric.accent)
        ws[cell_ref].alignment = Alignment(horizontal="center")
        apply_border(ws[cell_ref])

    for cell_ref in ("B4", "D4", "F4", "H4"):
        apply_border(ws[cell_ref])
        ws[cell_ref].alignment = Alignment(horizontal="center")
        if metric.value_kind == "amount" and cell_ref != "H4":
            format_money_cell(ws[cell_ref])
        elif cell_ref != "H4":
            format_count_cell(ws[cell_ref])

    ws["H4"].number_format = "0.0%" if isinstance(change_pct, float) else "General"
    if isinstance(change_pct, float) and change_pct < 0:
        ws["H4"].font = Font(color="C62828", bold=True)
    elif isinstance(change_pct, float):
        ws["H4"].font = Font(color="1B5E20", bold=True)

    for ref in ("B4", "D4", "F4"):
        if isinstance(ws[ref].value, (int, float)):
            if ws[ref].value < 0:
                ws[ref].font = Font(color="C62828", bold=True)
            else:
                ws[ref].font = Font(color="1B5E20", bold=True)

    next_row = 7
    next_row = render_monthly_table_section(ws, metric, df, next_row, "All Fields", today)
    next_row = render_monthly_table_section(
        ws,
        metric,
        df[df["ClaimScope"].eq("In Field")].copy(),
        next_row,
        "In Field",
        today,
    )
    next_row = render_monthly_table_section(
        ws,
        metric,
        df[df["ClaimScope"].eq("Pre Delivery")].copy(),
        next_row,
        "Pre Delivery",
        today,
    )

    ws[f"E{next_row}"] = "2026 YTD stops at the current system date."
    ws[f"E{next_row}"].font = Font(size=9, italic=True, color="6B7280")

    auto_size(ws, min_width=11, max_width=28)
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


def write_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data")
    ws.sheet_state = "hidden"
    if df.empty:
        return

    export_df = df.copy()
    export_df["CreatedDate"] = export_df["CreatedDate"].map(iso_date)
    export_df["ApprovedDate"] = export_df["ApprovedDate"].map(iso_date)
    export_df["ResolvedDate"] = export_df["ResolvedDate"].map(iso_date)
    for row in dataframe_to_rows(export_df, index=False, header=True):
        ws.append(row)


def comparison_detail_frames(df: pd.DataFrame, today: date) -> List[Tuple[MetricSpec, pd.DataFrame]]:
    frames: List[Tuple[MetricSpec, pd.DataFrame]] = []

    for metric in METRICS:
        if metric.key == "claims_received":
            work = df[df["CreatedDate"].notna()].copy()
            work = work[work["CreatedDate"].map(lambda d: d.year in {2025, 2026} and d <= today)].copy()
            work["Counted In"] = work["CreatedDate"].map(lambda d: "2025" if d.year == 2025 else "2026 YTD")
            work["Count Rule"] = "CreatedOn"
            work["Counted Date"] = work["CreatedDate"]
            work["Counted Value"] = 1
            ordered_cols = [
                "TicketID",
                "Counted In",
                "Count Rule",
                "Counted Date",
                "CreatedOn",
                "ClaimScope",
                "StatusGroup",
                "StatusText",
                "StatusCode",
                "DealerName",
                "TicketType",
                "TicketTypeText",
            ]
            work = work[ordered_cols].copy()
            work["Counted Date"] = work["Counted Date"].map(iso_date)
            frames.append((metric, work.sort_values(["Counted In", "Counted Date", "TicketID"], ascending=[True, False, False]).reset_index(drop=True)))
            continue

        if metric.key == "claims_closed":
            work = df[df["StatusGroup"].isin(["approved_closed", "unapproved_closed"])].copy()
            work = work[work["ResolvedDate"].notna()].copy()
            work = work[work["ResolvedDate"].map(lambda d: d.year in {2025, 2026} and d <= today)].copy()
            work["Counted In"] = work["ResolvedDate"].map(lambda d: "2025" if d.year == 2025 else "2026 YTD")
            work["Count Rule"] = "ResolvedOnDateTime"
            work["Counted Date"] = work["ResolvedDate"]
            work["Counted Value"] = 1
            ordered_cols = [
                "TicketID",
                "Counted In",
                "Count Rule",
                "Counted Date",
                "ResolvedOn",
                "ClaimScope",
                "StatusGroup",
                "StatusText",
                "StatusCode",
                "DealerName",
                "TicketType",
                "TicketTypeText",
            ]
            work = work[ordered_cols].copy()
            work["Counted Date"] = work["Counted Date"].map(iso_date)
            frames.append((metric, work.sort_values(["Counted In", "Counted Date", "TicketID"], ascending=[True, False, False]).reset_index(drop=True)))
            continue

        if metric.key == "amount_approved":
            approved_work = df[df["ApprovedAmountEligible"].fillna(False)].copy()
            approved_work = approved_work[approved_work["ApprovedDate"].notna()].copy()
            approved_work = approved_work[approved_work["ApprovedDate"].map(lambda d: d.year in {2025, 2026} and d <= today)].copy()
            approved_work["Counted In"] = approved_work["ApprovedDate"].map(lambda d: "2025" if d.year == 2025 else "2026 YTD")
            approved_work["Count Rule"] = "ClaimApprovedOnDateTime"
            approved_work["Approved Bucket"] = approved_work["StatusGroup"]
            approved_work["Counted Date"] = approved_work["ApprovedDate"]
            approved_work["Counted Value"] = pd.to_numeric(approved_work["ApprovedAmountValue"], errors="coerce").fillna(0)

            work = approved_work.copy()
            ordered_cols = [
                "TicketID",
                "Counted In",
                "Approved Bucket",
                "Count Rule",
                "Counted Date",
                "Counted Value",
                "ClaimScope",
                "StatusGroup",
                "StatusText",
                "StatusCode",
                "CreatedOn",
                "ClaimApprovedOnDateTime",
                "ClaimApprovedOn",
                "ERPPurchaseOrder",
                "NetValue",
                "ApprovedAmount",
                "ApprovedAmountSource",
                "ApprovedAmountValue",
                "ResolvedOn",
                "DealerName",
                "TicketType",
                "TicketTypeText",
            ]
            work = work[ordered_cols].copy()
            work["Counted Date"] = work["Counted Date"].map(iso_date)
            work["Counted Value"] = work["Counted Value"].map(lambda v: round(float(v), 2))
            frames.append((metric, work.sort_values(["Counted In", "Counted Date", "TicketID"], ascending=[True, False, False]).reset_index(drop=True)))
            continue

    return frames


def write_metric_detail_sheet(wb: Workbook, metric: MetricSpec, detail_df: pd.DataFrame, today: date) -> None:
    ws = wb.create_sheet(metric.title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{metric.title} detail"
    ws["A1"].font = Font(size=17, bold=True, color="0F172A")

    ws.merge_cells("A2:H2")
    if metric.key == "claims_received":
        desc = "Each row is a ticket counted by CreatedOn. Counted In shows whether it lands in 2025 or 2026 YTD."
    elif metric.key == "claims_closed":
        desc = "Each row is a ticket counted by ResolvedOnDateTime from approved closed or unapproved closed."
    else:
        desc = "Each row is a ticket counted by ClaimApprovedOnDateTime and current approved statuses Z9/Y0/Y1/Y2/Y4/YB, with amount sourced from fetch approvedAmount."
    ws["A2"] = desc
    ws["A2"].font = Font(size=10, color="475467")

    if detail_df.empty:
        ws["A4"] = "No rows found for this comparison."
        ws["A4"].font = Font(italic=True, color="6B7280")
        return

    headers = list(detail_df.columns)
    total_2025 = detail_df[detail_df["Counted In"].eq("2025")]
    total_2026 = detail_df[detail_df["Counted In"].eq("2026 YTD")]

    def overall_total_value(frame: pd.DataFrame) -> float:
        if metric.value_kind == "amount" and "Counted Value" in frame.columns:
            return float(pd.to_numeric(frame["Counted Value"], errors="coerce").fillna(0).sum())
        return float(len(frame))

    ws["A4"] = "2025"
    ws["B4"] = overall_total_value(total_2025)
    ws["C4"] = "2026 YTD"
    ws["D4"] = overall_total_value(total_2026)
    ws["E4"] = "Total"
    ws["F4"] = overall_total_value(detail_df)
    for ref in ("A4", "C4", "E4"):
        ws[ref].font = Font(bold=True, color="FFFFFF")
        ws[ref].fill = PatternFill("solid", fgColor=metric.accent)
        ws[ref].alignment = Alignment(horizontal="center")
        apply_border(ws[ref])
    for ref in ("B4", "D4", "F4"):
        ws[ref].font = Font(bold=True, color="0F172A")
        ws[ref].alignment = Alignment(horizontal="center")
        apply_border(ws[ref])
        if metric.value_kind == "amount":
            format_money_cell(ws[ref])
        else:
            format_count_cell(ws[ref])

    def section_summary(frame: pd.DataFrame) -> Tuple[float, float, float]:
        frame_2025 = frame[frame["Counted In"].eq("2025")]
        frame_2026 = frame[frame["Counted In"].eq("2026 YTD")]
        return (
            overall_total_value(frame_2025),
            overall_total_value(frame_2026),
            overall_total_value(frame),
        )

    def render_section(start_row: int, scope_name: str, scope_df: pd.DataFrame) -> int:
        end_col = max(6, len(headers))
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
        title_cell = ws.cell(row=start_row, column=1, value=scope_name)
        title_cell.font = Font(size=12, bold=True, color="0F172A")
        title_cell.fill = pastel_fill(metric.accent)
        title_cell.alignment = Alignment(horizontal="left")
        apply_border(title_cell)

        scope_2025, scope_2026, scope_total = section_summary(scope_df)
        row = start_row + 1
        summary_pairs = [("2025", scope_2025), ("2026 YTD", scope_2026), ("Total", scope_total)]
        for idx, (label, value) in enumerate(summary_pairs):
            label_col = 1 + idx * 2
            value_col = label_col + 1
            label_cell = ws.cell(row=row, column=label_col, value=label)
            label_cell.font = Font(bold=True, color="FFFFFF")
            label_cell.fill = PatternFill("solid", fgColor=metric.accent)
            label_cell.alignment = Alignment(horizontal="center")
            apply_border(label_cell)
            value_cell = ws.cell(row=row, column=value_col, value=value)
            value_cell.font = Font(bold=True, color="0F172A")
            value_cell.alignment = Alignment(horizontal="center")
            apply_border(value_cell)
            if metric.value_kind == "amount":
                format_money_cell(value_cell)
            else:
                format_count_cell(value_cell)

        row += 2
        if scope_df.empty:
            ws.cell(row=row, column=1, value="No rows found for this scope.")
            ws.cell(row=row, column=1).font = Font(italic=True, color="6B7280")
            return row + 2

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            apply_border(cell)

        for row_idx, row_values in enumerate(scope_df.itertuples(index=False), start=row + 1):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_border(cell)
                header_name = headers[col_idx - 1]
                if header_name in {"Counted In", "Count Rule", "Approved Bucket", "ClaimScope"}:
                    cell.alignment = Alignment(horizontal="center")
                elif header_name == "Counted Value":
                    if metric.value_kind == "amount":
                        format_money_cell(cell)
                    else:
                        format_count_cell(cell)

        return row + len(scope_df) + 2

    next_row = 6
    for scope_name in ("In Field", "Pre Delivery"):
        scope_df = detail_df[detail_df["ClaimScope"].eq(scope_name)].copy()
        next_row = render_section(next_row, scope_name, scope_df)

    auto_size(ws, min_width=12, max_width=34)
    for col_name in ["B", "D", "F"]:
        ws.column_dimensions[col_name].width = 16


def write_detail_workbook(df: pd.DataFrame, output_dir: Path, prefix: str, today: date) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    ws.merge_cells("A1:L1")
    ws["A1"] = "Claims YTD Comparison Detail"
    ws["A1"].font = Font(size=17, bold=True, color="0F172A")
    ws.merge_cells("A2:L2")
    ws["A2"] = "Single ticket-level table with scope, status, date, and fetch approvedAmount columns."
    ws["A2"].font = Font(size=10, color="475467")

    detail_df = df.copy()
    detail_df = detail_df[
        [
            "TicketID",
            "ClaimScope",
            "StatusGroup",
            "StatusText",
            "StatusCode",
            "CreatedOn",
            "ClaimApprovedOn",
            "ResolvedOn",
            "ERPPurchaseOrder",
            "ApprovedAmount",
            "ApprovedAmountSource",
            "ApprovedAmountValue",
            "DealerName",
            "TicketType",
            "TicketTypeText",
            "CreatedDate",
            "ApprovedDate",
            "ResolvedDate",
        ]
    ].copy()
    detail_df["CreatedDate"] = detail_df["CreatedDate"].map(iso_date)
    detail_df["ApprovedDate"] = detail_df["ApprovedDate"].map(iso_date)
    detail_df["ResolvedDate"] = detail_df["ResolvedDate"].map(iso_date)

    headers = list(detail_df.columns)
    table_start = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(cell)

    for row_idx, row in enumerate(detail_df.itertuples(index=False), start=table_start + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_border(cell)
            header_name = headers[col_idx - 1]
            if header_name in {"ClaimScope", "StatusGroup", "StatusCode"}:
                cell.alignment = Alignment(horizontal="center")
            elif header_name == "ApprovedAmountValue":
                format_money_cell(cell)

    auto_size(ws, min_width=12, max_width=34)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamped = output_dir / f"{prefix}_{timestamp}.xlsx"
    latest = output_dir / f"{prefix}_latest.xlsx"
    wb.save(timestamped)
    try:
        shutil.copyfile(timestamped, latest)
        return latest
    except PermissionError:
        return timestamped


def build_summary_rows(df: pd.DataFrame, today: date) -> List[Dict[str, Any]]:
    summary_rows: List[Dict[str, Any]] = []
    for metric in METRICS:
        total_2025 = metric_total(df, metric, date(2025, 1, 1), date(2025, 12, 31))
        total_2026 = metric_total(df, metric, date(2026, 1, 1), today)
        delta = total_2026 - total_2025
        change = pct_change(total_2026, total_2025)
        summary_rows.append(
            {
                "metric": metric.title,
                "value_2025": round(total_2025, 2) if metric.value_kind == "amount" else int(round(total_2025)),
                "value_2026": round(total_2026, 2) if metric.value_kind == "amount" else int(round(total_2026)),
                "delta": round(delta, 2) if metric.value_kind == "amount" else int(round(delta)),
                "pct_change": change,
                "description": metric.description,
                "accent": metric.accent,
                "value_kind": metric.value_kind,
            }
        )
    return summary_rows


def infer_source(args: argparse.Namespace) -> pd.DataFrame:
    if args.source == "firebase":
        return load_tickets_from_firebase(args.firebase_root, args.firebase_db_url, args.firebase_sa_path)
    if args.source == "csv":
        return load_tickets_from_csv(args.csv_path)
    raise ValueError(f"Unsupported source: {args.source}")


def build_workbook(df: pd.DataFrame, today: date) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = build_summary_rows(df, today)
    write_summary_sheet(wb, summary_rows, today)
    for metric in METRICS:
        write_metric_sheet(wb, metric, df, today)
    write_data_sheet(wb, df)

    wb.active = 0
    return wb


def save_workbook(wb: Workbook, output_dir: Path, prefix: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamped = output_dir / f"{prefix}_{timestamp}.xlsx"
    latest = output_dir / f"{prefix}_latest.xlsx"
    wb.save(timestamped)
    try:
        shutil.copyfile(timestamped, latest)
        return latest
    except PermissionError:
        return timestamped


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the 2025 vs 2026 YTD claims comparison workbook.")
    parser.add_argument("--source", choices=("firebase", "csv"), default="firebase")
    parser.add_argument("--firebase-root", default=DEFAULT_FIREBASE_ROOT)
    parser.add_argument("--firebase-db-url", default=DEFAULT_FIREBASE_DB_URL)
    parser.add_argument("--firebase-sa-path", default=DEFAULT_FIREBASE_SA_PATH)
    parser.add_argument("--csv-path", default=DEFAULT_LOCAL_CSV)
    parser.add_argument(
        "--output-prefix",
        default="claim_ytd_comparison_2025_vs_2026_ytd",
        help="Base filename without suffix.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    today = date.today()
    raw_df = infer_source(args)
    df = normalize_dataset(raw_df)
    wb = build_workbook(df, today)
    latest = save_workbook(wb, OUTPUT_DIR, args.output_prefix)
    detail_latest = write_detail_workbook(df, OUTPUT_DIR, DEFAULT_DETAILS_PREFIX, today)
    print(f"Workbook written: {latest}")
    print(f"Detail workbook written: {detail_latest}")
    print(f"Source rows: {len(df):,}")
    print(f"Cutoff date: {today.isoformat()}")


if __name__ == "__main__":
    main()
