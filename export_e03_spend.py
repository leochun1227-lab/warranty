#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export E03 (Regent Spare Parts) YTD spend to Excel.

Pulls every PO item under EKGRP='E03' with AEDAT >= 2026-01-01 (deleted
lines excluded), plus a small vendor lookup, plus AUD conversion at rate 5
for CNY lines. Produces one workbook with two sheets:
    Detail    -- one row per EKPO line, all fields for spot-checking
    Summary   -- totals rolled up by AFNAM, by doctype, and by vendor

Run:
    python export_e03_spend.py
    python export_e03_spend.py --out my-e03-spend.xlsx
    python export_e03_spend.py --year 2025    # different YTD baseline
"""
from __future__ import annotations

import argparse
import os
from datetime import date, datetime

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DSN = os.getenv(
    "SAP_HANA_DSN",
    "DRIVER={HDBODBC};SERVERNODE=10.11.2.25:30241;UID=BAOJIANFENG;PWD=Xja@2025ABC;",
)
MANDT = os.getenv("SAP_CLIENT", "800")
CNY_TO_AUD_RATE = float(os.getenv("CNY_TO_AUD_RATE", "5"))


DETAIL_COLUMNS = [
    ("PO",            "EBELN"),
    ("Item",          "EBELP"),
    ("PO Date",       "AEDAT"),
    ("Doctype",       "BSART"),
    ("Created By",    "ERNAM"),
    ("Requisitioner", "AFNAM"),
    ("Vendor Code",   "LIFNR"),
    ("Vendor Name",   "VENDOR_NAME"),
    ("Material",      "MATNR"),
    ("Description",   "TXZ01"),
    ("Qty",           "MENGE"),
    ("UoM",           "MEINS"),
    ("Unit Price",    "NETPR"),
    ("Price Unit",    "PEINH"),
    ("Net Value",     "NETWR"),
    ("Currency",      "WAERS"),
    ("Net Value AUD", "NETWR_AUD"),
]


def fmt_sap_date(s) -> str:
    s = str(s or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


def to_aud(netwr: float | None, currency: str | None) -> float:
    if netwr is None:
        return 0.0
    cur = (currency or "").strip().upper()
    if cur == "CNY":
        return round(float(netwr) / (CNY_TO_AUD_RATE or 5.0), 2)
    return round(float(netwr), 2)


def fetch_rows(conn, ytd_start: str, all_history: bool) -> list[dict]:
    # Match SAP ME2N with Purchasing Group=E03 and nothing else.
    # Any BSART is Nishi's -- do NOT exclude ZE03. Earlier attempts to exclude
    # ZE03 dropped this to 3 rows because ME2N's default selection parameter
    # happens to hide ZE03 too, but the user's actual definition of "Nishi's
    # spend" is the full EKGRP=E03 population.
    date_clause = "" if all_history else f"AND ekko.\"AEDAT\" >= '{ytd_start}'"
    sql = f"""
        SELECT
            ekpo."EBELN",  ekpo."EBELP",
            ekko."AEDAT",  ekko."BSART", ekko."ERNAM",
            ekpo."AFNAM",  ekko."LIFNR",
            ekpo."MATNR",  ekpo."TXZ01",
            ekpo."MENGE",  ekpo."MEINS",
            ekpo."NETPR",  ekpo."PEINH", ekpo."NETWR",
            ekko."WAERS"
        FROM "SAPHANADB"."EKPO" ekpo
        INNER JOIN "SAPHANADB"."EKKO" ekko
            ON ekko."MANDT"=ekpo."MANDT" AND ekko."EBELN"=ekpo."EBELN"
        WHERE ekpo."MANDT"='{MANDT}'
          AND ekko."EKGRP"='E03'
          {date_clause}
          AND (ekpo."LOEKZ" IS NULL OR ekpo."LOEKZ" = '')
          AND (ekko."LOEKZ" IS NULL OR ekko."LOEKZ" = '')
        ORDER BY ekko."AEDAT" DESC, ekpo."EBELN" DESC, ekpo."EBELP"
    """
    cur = conn.execute(sql)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def fetch_vendor_names(conn, vendor_codes: set[str]) -> dict[str, str]:
    codes = [c for c in vendor_codes if c]
    if not codes:
        return {}
    in_list = "','".join(c.replace("'", "''") for c in codes)
    sql = f"""
        SELECT "LIFNR","NAME1"
        FROM "SAPHANADB"."LFA1"
        WHERE "MANDT"='{MANDT}' AND "LIFNR" IN ('{in_list}')
    """
    try:
        cur = conn.execute(sql)
        return {r[0]: r[1] or "" for r in cur.fetchall()}
    except Exception as exc:
        print(f"  (LFA1 lookup failed, vendor names blank: {exc})")
        return {}


def group_totals(rows: list[dict], key: str) -> list[tuple[str, int, float]]:
    """Return [(value, line_count, total_aud), ...] sorted by AUD desc."""
    agg: dict[str, list[int | float]] = {}
    for r in rows:
        k = (r.get(key) or "(blank)")
        slot = agg.setdefault(k, [0, 0.0])
        slot[0] += 1
        slot[1] += r["NETWR_AUD"]
    return sorted(
        [(k, int(v[0]), round(v[1], 2)) for k, v in agg.items()],
        key=lambda t: t[2],
        reverse=True,
    )


HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
SUBHEADER_FONT = Font(bold=True)


def write_workbook(out_path: str, rows: list[dict], period_label: str) -> None:
    wb = Workbook()

    # ---- Detail sheet ----
    ws = wb.active
    ws.title = "Detail"
    headers = [label for label, _ in DETAIL_COLUMNS]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(key) for _, key in DETAIL_COLUMNS])

    # Column widths
    widths = {"PO": 14, "Item": 6, "PO Date": 12, "Doctype": 9, "Created By": 12,
              "Requisitioner": 16, "Vendor Code": 12, "Vendor Name": 32,
              "Material": 14, "Description": 40, "Qty": 8, "UoM": 6,
              "Unit Price": 12, "Price Unit": 10, "Net Value": 14,
              "Currency": 8, "Net Value AUD": 14}
    for i, (label, _) in enumerate(DETAIL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(label, 12)
    ws.freeze_panes = "A2"

    # ---- Summary sheet ----
    sm = wb.create_sheet("Summary")
    sm.append(["E03 (Regent Spare Parts) YTD Spend Summary"])
    sm["A1"].font = Font(bold=True, size=14)
    sm.append([f"Period: {period_label}", f"Detail lines: {len(rows)}"])
    total_aud = round(sum(r["NETWR_AUD"] for r in rows), 2)
    unique_pos = len({r["EBELN"] for r in rows if r.get("EBELN")})
    sm.append([f"Unique POs: {unique_pos}", f"Total (AUD, CNY converted at /{CNY_TO_AUD_RATE:g}): {total_aud:,.2f}"])
    sm.append([])

    def dump_group(title: str, key: str, name_col: str) -> None:
        sm.append([title])
        sm.cell(row=sm.max_row, column=1).font = SUBHEADER_FONT
        sm.append([name_col, "Lines", "AUD Total"])
        for c in sm[sm.max_row]:
            c.fill = SUBHEADER_FILL
            c.font = SUBHEADER_FONT
        for value, count, total in group_totals(rows, key):
            sm.append([value, count, total])
        sm.append([])

    dump_group("By Requisitioner (AFNAM)", "AFNAM", "Requisitioner")
    dump_group("By Doctype (BSART)", "BSART", "Doctype")
    dump_group("By Vendor", "VENDOR_NAME", "Vendor Name")

    for col_letter, width in [("A", 32), ("B", 10), ("C", 16)]:
        sm.column_dimensions[col_letter].width = width

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=None,
                    help="Output .xlsx path (default: e03_spend_<year>_<timestamp>.xlsx)")
    ap.add_argument("--year", type=int, default=None,
                    help="YTD baseline year (default: current year)")
    ap.add_argument("--all-history", action="store_true",
                    help="No date filter -- match ME2N's default output (all time). Off by default: YTD only.")
    args = ap.parse_args()

    ytd_year = args.year or datetime.now().year
    ytd_start = f"{ytd_year:04d}0101"
    period_label = "all history" if args.all_history else f"{ytd_start} onward"
    out_path = args.out or f"e03_spend_{'all' if args.all_history else ytd_year}_{datetime.now():%Y%m%d_%H%M}.xlsx"

    print(f"Connecting to HANA and pulling E03 spend ({period_label}, EKGRP=E03 only, no doctype filter) ...")
    with pyodbc.connect(DSN, autocommit=True) as conn:
        rows = fetch_rows(conn, ytd_start, args.all_history)
        print(f"  Got {len(rows)} EKPO lines")

        vendor_codes = {(r.get("LIFNR") or "").strip() for r in rows}
        vendor_names = fetch_vendor_names(conn, vendor_codes)
        print(f"  Resolved {len(vendor_names)} vendor names")

    for r in rows:
        r["AEDAT"] = fmt_sap_date(r.get("AEDAT"))
        r["VENDOR_NAME"] = vendor_names.get((r.get("LIFNR") or "").strip(), "")
        r["NETWR"] = round(float(r.get("NETWR") or 0), 2)
        r["NETPR"] = round(float(r.get("NETPR") or 0), 4)
        r["MENGE"] = float(r.get("MENGE") or 0)
        r["PEINH"] = int(r.get("PEINH") or 1)
        r["NETWR_AUD"] = to_aud(r["NETWR"], r.get("WAERS"))

    write_workbook(out_path, rows, period_label)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
