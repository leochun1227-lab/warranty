import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_SOURCE = Path.home() / "Desktop" / "Repairer_Business_Name_Rule_Mapping.xlsx"
DEFAULT_OUTPUT = ROOT / "assets" / "repairer_name_rule_mapping.json"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_dict(headers, values):
    return {header: clean(value) for header, value in zip(headers, values) if header}


def rows_from_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    return [row_dict(headers, row) for row in rows[1:]]


def compact_rows(rows, keys):
    output = []
    seen = set()
    for row in rows:
        item = {key: clean(row.get(key)) for key in keys}
        if not any(item.values()):
            continue
        dedupe_key = tuple(item.get(key, "") for key in keys)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        output.append(item)
    return output


def ensure_snowy_dealer_code_alias(rows, old_code, new_code, dealership, repair_shop="SNOWY RIVER RV PTY LTD Repairs"):
    old_code = clean(old_code)
    new_code = clean(new_code)
    dealership = clean(dealership)
    repair_shop = clean(repair_shop)
    for row in rows:
        if (
            clean(row.get("oldWarrantyHandlingDealerAssign")) == old_code
            and clean(row.get("newWarrantyHandlingDealerAssign")) == new_code
            and clean(row.get("dealership")).casefold() == dealership.casefold()
        ):
            return
    rows.append({
        "repairShopC4C": repair_shop,
        "oldWarrantyHandlingDealerAssign": old_code,
        "newWarrantyHandlingDealerAssign": new_code,
        "dealership": dealership,
    })


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.source, data_only=True, read_only=True)

    general_rows = rows_from_sheet(workbook, "General Mapping")
    snowy_rows = rows_from_sheet(workbook, "SNOWY RIVER RV PTY LTD (Parts)")

    general = []
    for row in general_rows:
        repair_shop = clean(row.get("Repair Shop(C4C)") or row.get("Repair Shop (C4C)") or row.get("Repair Shop"))
        mapping_name = clean(row.get("Mapping name") or row.get("Mapping Name"))
        if repair_shop and mapping_name:
            general.append({"repairShopC4C": repair_shop, "mappingName": mapping_name})

    snowy_dealer_codes = []
    for row in snowy_rows:
        repair_shop = clean(row.get("Repair Shop( C4C)") or row.get("Repair Shop(C4C)") or row.get("Repair Shop"))
        old_code = clean(row.get("Old CODE Warranty Handling Dealer(Assign)") or row.get("Old Warranty Handling Dealer(Assign)"))
        new_code = clean(row.get("New CODE Warranty Handling Dealer(Assign)") or row.get("New Warranty Handling Dealer(Assign)"))
        dealership = clean(row.get("Dealership") or row.get("Mapping name") or row.get("Mapping Name"))
        if dealership and (old_code or new_code or repair_shop):
            snowy_dealer_codes.append({
                "repairShopC4C": repair_shop,
                "oldWarrantyHandlingDealerAssign": old_code,
                "newWarrantyHandlingDealerAssign": new_code,
                "dealership": dealership,
            })

    ensure_snowy_dealer_code_alias(snowy_dealer_codes, "3151", "3151", "Regent RV - Frankston")

    payload = {
        "meta": {
            "source": args.source.name,
            "source_sheets": workbook.sheetnames,
            "mapping_rule": "For repairer SNOWY RV PTY LTD / SNOWY RIVER RV PTY LTD, map Warranty Handling Dealer(Assign) old/new code to Dealership.",
        },
        "general": compact_rows(general, ["repairShopC4C", "mappingName"]),
        "snowyRiverDealerCodes": compact_rows(
            snowy_dealer_codes,
            [
                "repairShopC4C",
                "oldWarrantyHandlingDealerAssign",
                "newWarrantyHandlingDealerAssign",
                "dealership",
            ],
        ),
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "general": len(payload["general"]),
        "snowyRiverDealerCodes": len(payload["snowyRiverDealerCodes"]),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
