# -*- coding: utf-8 -*-
"""
Read-only approved-logic validator.

Purpose:
1. Load ticket rows from Firebase, Excel, or CSV.
2. Compare multiple approved candidate rules side by side.
3. Export an Excel workbook for manual validation.

This script never writes back to Firebase.
"""

from __future__ import annotations

import argparse
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

try:
    import firebase_admin
    from firebase_admin import credentials, db
except Exception:  # pragma: no cover
    firebase_admin = None
    credentials = None
    db = None


DEFAULT_FIREBASE_DB_URL = os.getenv(
    "FIREBASE_DB_URL",
    "https://snowy-hr-report-default-rtdb.asia-southeast1.firebasedatabase.app",
)
DEFAULT_FIREBASE_ROOT = os.getenv("FIREBASE_ROOT", "c4cTickets_test")
DEFAULT_FIREBASE_SA_PATH = os.getenv(
    "FIREBASE_SA_PATH",
    str(Path.cwd() / "firebase-service-account.json"),
)
DEFAULT_INPUT_XLSX = Path("outputs") / "claim_ytd_comparison" / "claim_ytd_comparison_tickets_detail_latest.xlsx"

APPROVED_ACTIVE_STATUS_CODES = {"Z9", "Y0", "Y1", "Y2", "Y4", "YB"}
APPROVED_ACTIVE_STATUS_TEXTS = {
    "sales order approved",
    "partially picked",
    "dispatch parts",
    "repair in progress",
    "repairer invoiced received",
    "repairer invoiced processed",
}
APPROVED_CLOSED_STATUS_CODES = {"Y7"}
UNAPPROVED_CLOSED_STATUS_CODES = {"Y8"}

DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%dT%H:%M:%S.%fZ",
    "%d/%m/%Y",
    "%d/%m/%Y %H:%M:%S",
    "%d.%m.%Y",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M:%S AUSACT",
    "%Y%m%d",
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip().lower()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def first_non_blank(*values: Any) -> str:
    for value in values:
        text = clean(value)
        if text:
            return text
    return ""


def parse_any_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean(value)
    if not text or text == "00000000":
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    try:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
        if not pd.isna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None


def parse_amount(value: Any) -> float:
    text = clean(value)
    if not text:
        return 0.0
    text = text.replace(",", "").replace("$", "").replace("AUD", "").strip()
    try:
        return float(text)
    except Exception:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group(0)) if match else 0.0


def iso_date(value: Any) -> str:
    parsed = parse_any_date(value)
    return parsed.isoformat() if parsed else ""


def status_bucket(status_code: Any, status_text: Any) -> str:
    code = clean(status_code).upper()
    text = normalize_text(status_text)
    combined = f"{code.lower()} {text}".strip()

    if code in UNAPPROVED_CLOSED_STATUS_CODES or "unapproved claims closed" in combined or "unapproved closed" in combined:
        return "unapproved_closed"
    if code in APPROVED_CLOSED_STATUS_CODES or "approved claims closed" in combined or "claim approved closed" in combined or "approved closed" in combined:
        return "approved_closed"
    if code in APPROVED_ACTIVE_STATUS_CODES or text in APPROVED_ACTIVE_STATUS_TEXTS:
        return "approved_active"
    return "other"


def classify_reason(row: pd.Series) -> str:
    if not bool(row.get("HasClaimApprovedDate")):
        return "claim_approved_on_blank"
    bucket = clean(row.get("StatusBucket"))
    if bucket == "approved_active":
        return "claim_date_and_active_approved_status"
    if bucket == "approved_closed":
        return "claim_date_but_approved_closed"
    if bucket == "unapproved_closed":
        return "claim_date_but_unapproved_closed"
    return "claim_date_but_status_outside_active_whitelist"


def firebase_init(db_url: str, sa_path: str) -> None:
    if firebase_admin is None or credentials is None or db is None:
        raise RuntimeError("firebase_admin is not installed in this environment.")
    if getattr(firebase_admin, "_apps", None):
        return
    path = Path(sa_path)
    if not path.exists():
        raise FileNotFoundError(f"Firebase service account json not found: {path}")
    firebase_admin.initialize_app(credentials.Certificate(str(path)), {"databaseURL": db_url})


def firebase_node_to_dict(node: Any) -> dict[str, Any]:
    if isinstance(node, dict):
        return node
    if isinstance(node, list):
        return {str(i): value for i, value in enumerate(node) if value is not None}
    return {}


def load_from_firebase(firebase_root: str, firebase_db_url: str, firebase_sa_path: str) -> pd.DataFrame:
    firebase_init(firebase_db_url, firebase_sa_path)
    raw = db.reference(f"{firebase_root}/tickets").get()
    tickets = firebase_node_to_dict(raw)
    rows: list[dict[str, Any]] = []

    for fallback_id, node in tickets.items():
        ticket = (node or {}).get("ticket", {}) if isinstance(node, dict) else node
        if not isinstance(ticket, dict):
            continue
        rows.append(
            {
                "TicketID": first_non_blank(ticket.get("TicketID"), ticket.get("ticketID"), ticket.get("id"), fallback_id),
                "TicketName": first_non_blank(ticket.get("TicketName"), ticket.get("Name"), ticket.get("Subject")),
                "StatusCode": first_non_blank(ticket.get("TicketStatus"), ticket.get("TicketStatusCode"), ticket.get("StatusCode"), ticket.get("Status")),
                "StatusText": first_non_blank(ticket.get("TicketStatusText"), ticket.get("statusText"), ticket.get("Status")),
                "CreatedOn": first_non_blank(ticket.get("CreatedOn"), ticket.get("Created On"), ticket.get("createdOn"), ticket.get("CreatedAt")),
                "ClaimApprovedOnDateTime": first_non_blank(
                    ticket.get("ClaimApprovedOnDateTime"),
                    ticket.get("ClaimApprovedOnDate"),
                    ticket.get("ClaimApprovedOn"),
                    ticket.get("Claim Approved On"),
                    ticket.get("Claim Approved On DateTime"),
                    ticket.get("ApprovalDate"),
                ),
                "ResolvedOnDateTime": first_non_blank(
                    ticket.get("ResolvedOnDateTime"),
                    ticket.get("ResolvedOnDate"),
                    ticket.get("ResolvedOn"),
                    ticket.get("Resolved On"),
                ),
                "ChangedOnDateTime": first_non_blank(
                    ticket.get("ChangeOnDateTime"),
                    ticket.get("ChangeOnDate"),
                    ticket.get("ChangeOn"),
                    ticket.get("ChangedOn"),
                    ticket.get("Changed On"),
                    ticket.get("LastChangeDateTime"),
                    ticket.get("LastUpdatedDateTime"),
                ),
                "ERPPurchaseOrder": first_non_blank(ticket.get("ERPPurchaseOrder"), ticket.get("ERP Purchase Order ID"), ticket.get("Purchasing Document")),
                "SalesOrder": first_non_blank(ticket.get("Sales Order"), ticket.get("SalesOrder"), ticket.get("LookupSalesOrder")),
                "AmountIncludingTax": first_non_blank(ticket.get("AmountIncludingTax"), ticket.get("ClaimTotalAmount")),
                "ApprovedAmount": first_non_blank(ticket.get("approvedAmount"), ticket.get("ApprovedAmount")),
                "ApprovedAmountSource": first_non_blank(ticket.get("approvedAmountSource"), ticket.get("ApprovedAmountSource")),
                "PoNetValue": first_non_blank(ticket.get("Net Value"), ticket.get("NetValue")),
                "DealerName": first_non_blank(ticket.get("DealerName"), ticket.get("Dealer Name")),
                "TicketType": first_non_blank(ticket.get("TicketType"), ticket.get("Ticket Type")),
                "TicketTypeText": first_non_blank(ticket.get("TicketTypeText"), ticket.get("Ticket Type Text")),
                "StoredApprovalDecision": first_non_blank(ticket.get("approvalDecision"), ticket.get("decision")),
            }
        )

    return pd.DataFrame(rows)


def detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    for idx, row in enumerate(rows[:25]):
        normalized = [normalize_header(cell) for cell in row]
        if "ticketid" in normalized or "ticket" in normalized:
            return idx
    raise RuntimeError("Could not detect header row in Excel sheet.")


def load_raw_excel(path: str, sheet_name: str = "") -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = detect_header_row(rows)
    headers = [clean(v) for v in rows[header_idx]]
    data_rows = rows[header_idx + 1 :]
    data_rows = [row for row in data_rows if any(clean(cell) for cell in row)]
    return pd.DataFrame(data_rows, columns=headers)


def load_raw_csv(path: str) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def pick_from_row(row: pd.Series, cols: dict[str, str], *names: str) -> str:
    for name in names:
        col = cols.get(normalize_header(name))
        if col:
            value = clean(row.get(col))
            if value:
                return value
    return ""


def unify_columns(raw_df: pd.DataFrame) -> pd.DataFrame:
    cols = {normalize_header(col): col for col in raw_df.columns}
    rows: list[dict[str, Any]] = []

    for _, row in raw_df.iterrows():
        rows.append(
            {
                "TicketID": pick_from_row(row, cols, "TicketID", "Ticket ID", "Raw TicketID"),
                "TicketName": pick_from_row(row, cols, "TicketName", "Ticket Name", "Ticket", "Customer", "Raw TicketName"),
                "StatusCode": pick_from_row(row, cols, "StatusCode", "TicketStatus", "Ticket Status", "Raw TicketStatus"),
                "StatusText": pick_from_row(row, cols, "StatusText", "TicketStatusText", "Ticket Status Text", "Status", "Raw TicketStatusText"),
                "CreatedOn": pick_from_row(row, cols, "CreatedOn", "Created On", "CreatedDate"),
                "ClaimApprovedOnDateTime": pick_from_row(
                    row,
                    cols,
                    "ClaimApprovedOnDateTime",
                    "ClaimApprovedOn",
                    "Claim Approved On",
                    "ClaimApprovedOnDate",
                    "ApprovedDate",
                    "Decision Date",
                ),
                "ResolvedOnDateTime": pick_from_row(row, cols, "ResolvedOnDateTime", "ResolvedOn", "Resolved On", "ResolvedDate"),
                "ChangedOnDateTime": pick_from_row(row, cols, "ChangedOnDateTime", "Changed On", "ChangedOn"),
                "ERPPurchaseOrder": pick_from_row(row, cols, "ERPPurchaseOrder", "ERP Purchase Order ID", "Purchasing Document"),
                "SalesOrder": pick_from_row(row, cols, "SalesOrder", "Sales Order", "ERP Service Order ID"),
                "AmountIncludingTax": pick_from_row(row, cols, "AmountIncludingTax", "ClaimAmountValue", "ClaimTotalAmount"),
                "ApprovedAmount": pick_from_row(row, cols, "ApprovedAmount", "approvedAmount", "Approved Amount", "Amount"),
                "ApprovedAmountSource": pick_from_row(row, cols, "ApprovedAmountSource", "approvedAmountSource", "Price Source"),
                "PoNetValue": pick_from_row(row, cols, "PoNetValue", "NetValue", "Net Value", "PO Net Value"),
                "DealerName": pick_from_row(row, cols, "DealerName", "Dealer Name", "Dealer"),
                "TicketType": pick_from_row(row, cols, "TicketType", "Ticket Type"),
                "TicketTypeText": pick_from_row(row, cols, "TicketTypeText", "Ticket Type Text", "Claim Type"),
                "StoredApprovalDecision": pick_from_row(row, cols, "approvalDecision", "ApprovalDecision", "Decision"),
            }
        )

    return pd.DataFrame(rows)


def normalize_dataset(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in [
        "TicketID",
        "TicketName",
        "StatusCode",
        "StatusText",
        "CreatedOn",
        "ClaimApprovedOnDateTime",
        "ResolvedOnDateTime",
        "ChangedOnDateTime",
        "ERPPurchaseOrder",
        "SalesOrder",
        "AmountIncludingTax",
        "ApprovedAmount",
        "ApprovedAmountSource",
        "PoNetValue",
        "DealerName",
        "TicketType",
        "TicketTypeText",
        "StoredApprovalDecision",
    ]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    out["StatusBucket"] = out.apply(lambda row: status_bucket(row["StatusCode"], row["StatusText"]), axis=1)
    out["ApprovedDate"] = out["ClaimApprovedOnDateTime"].map(parse_any_date)
    out["CreatedDate"] = out["CreatedOn"].map(parse_any_date)
    out["ResolvedDate"] = out["ResolvedOnDateTime"].map(parse_any_date)
    out["ChangedDate"] = out["ChangedOnDateTime"].map(parse_any_date)
    out["ApprovedAmountValue"] = out["ApprovedAmount"].map(parse_amount)
    out["ClaimAmountValue"] = out["AmountIncludingTax"].map(parse_amount)
    out["PoNetValueAmount"] = out["PoNetValue"].map(parse_amount)

    out["HasClaimApprovedDate"] = out["ApprovedDate"].notna()
    out["IsApprovedActiveStatus"] = out["StatusBucket"].eq("approved_active")
    out["IsApprovedClosedStatus"] = out["StatusBucket"].eq("approved_closed")
    out["IsUnapprovedClosedStatus"] = out["StatusBucket"].eq("unapproved_closed")

    out["Rule_ClaimDate_NotUnapprovedClosed"] = out["HasClaimApprovedDate"] & ~out["IsUnapprovedClosedStatus"]
    out["Rule_ClaimDate_ApprovedWhitelist"] = out["HasClaimApprovedDate"] & out["IsApprovedActiveStatus"]
    out["Rule_ClaimDate_ApprovedOrClosed"] = out["HasClaimApprovedDate"] & (out["IsApprovedActiveStatus"] | out["IsApprovedClosedStatus"])
    out["Rule_StatusWhitelistOnly"] = out["IsApprovedActiveStatus"]

    out["CurrentPythonApproxApproved"] = out.apply(current_python_approx, axis=1)
    out["DiagnosticReason"] = out.apply(classify_reason, axis=1)
    out["ApprovedDateIso"] = out["ApprovedDate"].map(iso_date)
    out["CreatedDateIso"] = out["CreatedDate"].map(iso_date)
    out["ResolvedDateIso"] = out["ResolvedDate"].map(iso_date)
    out["ChangedDateIso"] = out["ChangedDate"].map(iso_date)
    return out


def current_python_approx(row: pd.Series) -> bool:
    stored = clean(row.get("StoredApprovalDecision")).lower()
    if stored == "approved":
        return True
    if stored == "unapproved":
        return False
    if clean(row.get("StatusBucket")) == "approved_closed":
        return False
    if clean(row.get("ERPPurchaseOrder")) or clean(row.get("SalesOrder")):
        return True
    if bool(row.get("IsApprovedActiveStatus")):
        return True
    if bool(row.get("IsUnapprovedClosedStatus")):
        return False
    return False


def build_summary(df: pd.DataFrame, source_label: str) -> pd.DataFrame:
    metrics = [
        ("Source", source_label),
        ("Total tickets", len(df)),
        ("ClaimApprovedOnDateTime non-empty", int(df["HasClaimApprovedDate"].sum())),
        ("Rule A: Claim date + not unapproved closed", int(df["Rule_ClaimDate_NotUnapprovedClosed"].sum())),
        ("Rule B: Claim date + approved whitelist status", int(df["Rule_ClaimDate_ApprovedWhitelist"].sum())),
        ("Rule C: Claim date + approved whitelist or approved closed", int(df["Rule_ClaimDate_ApprovedOrClosed"].sum())),
        ("Rule D: Status whitelist only", int(df["Rule_StatusWhitelistOnly"].sum())),
        ("Current Python approx", int(df["CurrentPythonApproxApproved"].sum())),
        (
            "A but not B",
            int((df["Rule_ClaimDate_NotUnapprovedClosed"] & ~df["Rule_ClaimDate_ApprovedWhitelist"]).sum()),
        ),
        (
            "B but not A",
            int((df["Rule_ClaimDate_ApprovedWhitelist"] & ~df["Rule_ClaimDate_NotUnapprovedClosed"]).sum()),
        ),
    ]
    return pd.DataFrame(metrics, columns=["Metric", "Value"])


def build_status_summary(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    grouped = (
        work.groupby(["StatusCode", "StatusText", "StatusBucket"], dropna=False)
        .agg(
            TicketCount=("TicketID", "size"),
            ClaimApprovedDateCount=("HasClaimApprovedDate", "sum"),
            RuleA=("Rule_ClaimDate_NotUnapprovedClosed", "sum"),
            RuleB=("Rule_ClaimDate_ApprovedWhitelist", "sum"),
            RuleC=("Rule_ClaimDate_ApprovedOrClosed", "sum"),
            RuleD=("Rule_StatusWhitelistOnly", "sum"),
            PythonApprox=("CurrentPythonApproxApproved", "sum"),
        )
        .reset_index()
        .sort_values(["TicketCount", "ClaimApprovedDateCount", "StatusCode"], ascending=[False, False, True])
    )
    return grouped


def build_reason_summary(df: pd.DataFrame) -> pd.DataFrame:
    work = (
        df.groupby(["DiagnosticReason", "StatusBucket", "StatusCode", "StatusText"], dropna=False)
        .size()
        .reset_index(name="TicketCount")
        .sort_values(["TicketCount", "DiagnosticReason"], ascending=[False, True])
    )
    return work


def ordered_detail_columns(df: pd.DataFrame) -> list[str]:
    cols = [
        "TicketID",
        "TicketName",
        "StatusCode",
        "StatusText",
        "StatusBucket",
        "StoredApprovalDecision",
        "ClaimApprovedOnDateTime",
        "ApprovedDateIso",
        "CreatedOn",
        "CreatedDateIso",
        "ResolvedOnDateTime",
        "ResolvedDateIso",
        "ChangedOnDateTime",
        "ChangedDateIso",
        "ERPPurchaseOrder",
        "SalesOrder",
        "DealerName",
        "TicketType",
        "TicketTypeText",
        "ClaimAmountValue",
        "ApprovedAmountValue",
        "PoNetValueAmount",
        "Rule_ClaimDate_NotUnapprovedClosed",
        "Rule_ClaimDate_ApprovedWhitelist",
        "Rule_ClaimDate_ApprovedOrClosed",
        "Rule_StatusWhitelistOnly",
        "CurrentPythonApproxApproved",
        "DiagnosticReason",
    ]
    return [col for col in cols if col in df.columns]


def auto_width_and_filter(writer: pd.ExcelWriter) -> None:
    for ws in writer.book.worksheets:
        if ws.max_row >= 2:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = 10
            col_letter = column_cells[0].column_letter
            for cell in column_cells[:250]:
                value = clean(cell.value)
                if value:
                    max_len = max(max_len, min(len(value), 60))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 64)


def write_excel(path: str, sheets: dict[str, pd.DataFrame]) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
        auto_width_and_filter(writer)


def resolve_source(args: argparse.Namespace) -> tuple[str, str]:
    source = clean(args.source).lower() or "auto"
    input_path = clean(args.input)
    if source != "auto":
        return source, input_path
    if input_path:
        ext = Path(input_path).suffix.lower()
        if ext in {".xlsx", ".xlsm", ".xls"}:
            return "xlsx", input_path
        if ext == ".csv":
            return "csv", input_path
    if DEFAULT_INPUT_XLSX.exists():
        return "xlsx", str(DEFAULT_INPUT_XLSX)
    return "firebase", input_path


def load_dataset(args: argparse.Namespace) -> tuple[pd.DataFrame, str]:
    source, input_path = resolve_source(args)
    if source == "firebase":
        df = load_from_firebase(args.firebase_root, args.firebase_db_url, args.firebase_sa_path)
        return df, f"firebase:{args.firebase_root}"
    if source == "xlsx":
        if not input_path:
            raise SystemExit("Excel source selected but no --input path was provided and no default workbook was found.")
        raw = load_raw_excel(input_path, args.sheet_name)
        return unify_columns(raw), f"xlsx:{input_path}"
    if source == "csv":
        if not input_path:
            raise SystemExit("CSV source selected but no --input path was provided.")
        raw = load_raw_csv(input_path)
        return unify_columns(raw), f"csv:{input_path}"
    raise SystemExit(f"Unsupported source: {source}")


def default_output_path() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(Path.cwd() / f"approved_logic_validation_{stamp}.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare approved candidate rules and export an Excel workbook.")
    parser.add_argument("--source", default="auto", choices=["auto", "firebase", "xlsx", "csv"])
    parser.add_argument("--input", default="")
    parser.add_argument("--sheet-name", default="")
    parser.add_argument("--output", default=default_output_path())
    parser.add_argument("--firebase-db-url", default=DEFAULT_FIREBASE_DB_URL)
    parser.add_argument("--firebase-root", default=DEFAULT_FIREBASE_ROOT)
    parser.add_argument("--firebase-sa-path", default=DEFAULT_FIREBASE_SA_PATH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    raw_df, source_label = load_dataset(args)
    df = normalize_dataset(raw_df)

    all_rows = df[ordered_detail_columns(df)].sort_values(
        ["Rule_ClaimDate_NotUnapprovedClosed", "Rule_ClaimDate_ApprovedWhitelist", "ApprovedDateIso", "TicketID"],
        ascending=[False, False, False, True],
    )
    loose_not_strict = df[
        df["Rule_ClaimDate_NotUnapprovedClosed"] & ~df["Rule_ClaimDate_ApprovedWhitelist"]
    ][ordered_detail_columns(df)].sort_values(["StatusBucket", "StatusCode", "ApprovedDateIso", "TicketID"])
    strict_rows = df[df["Rule_ClaimDate_ApprovedWhitelist"]][ordered_detail_columns(df)].sort_values(
        ["ApprovedDateIso", "TicketID"], ascending=[False, True]
    )
    approved_or_closed_rows = df[df["Rule_ClaimDate_ApprovedOrClosed"]][ordered_detail_columns(df)].sort_values(
        ["StatusBucket", "ApprovedDateIso", "TicketID"], ascending=[True, False, True]
    )
    suspicious_claim_rows = df[
        df["HasClaimApprovedDate"] & ~df["IsApprovedActiveStatus"]
    ][ordered_detail_columns(df)].sort_values(["StatusBucket", "StatusCode", "ApprovedDateIso", "TicketID"])

    sheets = {
        "Summary": build_summary(df, source_label),
        "Status_Summary": build_status_summary(df),
        "Reason_Summary": build_reason_summary(df),
        "All_Tickets": all_rows,
        "RuleB_Strict_Approved": strict_rows,
        "RuleA_Loose_Not_RuleB": loose_not_strict,
        "ClaimDate_ApprovedOrClosed": approved_or_closed_rows,
        "ClaimDate_Not_ActiveStatus": suspicious_claim_rows,
    }
    write_excel(args.output, sheets)
    print(f"Source: {source_label}")
    print(f"Rows loaded: {len(df)}")
    print(f"Excel written: {args.output}")


if __name__ == "__main__":
    main()
