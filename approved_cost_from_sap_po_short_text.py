# -*- coding: utf-8 -*-
"""
Approved cost proof-of-concept using SAP PO Short Text ticket numbers.

This script does not change the main fetch pipeline. It:
1. Reads current C4C/Firebase tickets.
2. Uses the existing approved-ticket logic from ctm_v44_history_safe_mandt800_rejection_filter.py.
3. Fetches SAP PO items from HANA by PO criteria, not by C4C PO fields.
4. Extracts ticket numbers from EKPO.TXZ01 / Short Text.
5. Calculates approved cost from SAP PO values matched by ticket number.
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

import pandas as pd
import pyodbc
from firebase_admin import db

from ctm_v44_history_safe_mandt800_rejection_filter import (
    DEFAULT_DB_URL,
    DEFAULT_SOURCE_ROOT,
    approval_decision,
    approval_decision_date,
    clean,
    init_firebase,
    normalize_row,
    normalize_ticket_id,
    parse_amount,
)

try:
    from fetch_all_tickets_fast_with_firebase_MANDT800_REJECTION_FILTER import SAP_HANA_DSN as FETCH_SAP_HANA_DSN
except Exception:
    FETCH_SAP_HANA_DSN = ""


warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable.*",
    category=UserWarning,
)


DEFAULT_OUTPUT = "approved_cost_from_sap_po_short_text.xlsx"
DEFAULT_SCHEMA = "SAPHANADB"
DEFAULT_CLIENT = "800"
DEFAULT_PURCHASING_ORG = "3111"
DEFAULT_PURCHASING_GROUP = "E06"
DEFAULT_AMOUNT_COLUMN = "Net Order Value"
APPROVED_COLUMNS = [
    "Ticket Number",
    "TicketID",
    "Approved Date",
    "Ticket Status",
    "Ticket Status Text",
    "Dealer ID",
    "Dealer Name",
    "C4C AmountIncludingTax",
    "ERP Purchase Order",
    "ERP Free Order",
    "Sales Order",
    "Created On",
    "Chassis Number",
    "Claim Approved On",
]
PO_AGG_COLUMNS = [
    "Ticket Number",
    "SAP_PO_Item_Count",
    "SAP_PO_Amount",
    "SAP_Net_Order_Value",
    "SAP_Net_Price_Sum",
    "SAP_PO_Documents",
    "SAP_Currencies",
    "SAP_Short_Text_Sample",
]

TICKET_NO_PATTERN = re.compile(
    r"\btickets?\s*no\.?\s*[:#\-]?\s*\[?\s*(\d+)\s*\]?\b",
    flags=re.IGNORECASE,
)
TICKET_BRACKET_PATTERN = re.compile(
    r"\btickets?\s*\[\s*(\d+)\s*\]",
    flags=re.IGNORECASE,
)
TICKET_OUTER_BRACKET_PATTERN = re.compile(
    r"\[\s*tickets?\s+(\d+)\s*(?:[\]\}]|$|[,，;；]|\s)?",
    flags=re.IGNORECASE,
)
logger = logging.getLogger("approved_cost_from_sap_po_short_text")


def setup_logging(log_file: str) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_file, encoding="utf-8"),
        ],
    )


def sql_quote(value: Any) -> str:
    return str(value).replace("'", "''")


def quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def full_table(schema: str, table: str) -> str:
    return f"{quote_ident(schema)}.{quote_ident(table)}"


def read_sql(conn: pyodbc.Connection, sql: str, params: Optional[Sequence[Any]] = None) -> pd.DataFrame:
    return pd.read_sql(sql, conn, params=list(params or []))


def get_table_columns(conn: pyodbc.Connection, schema: str, table: str) -> list[str]:
    sql = """
    SELECT COLUMN_NAME
    FROM SYS.TABLE_COLUMNS
    WHERE SCHEMA_NAME = ?
      AND TABLE_NAME = ?
    ORDER BY POSITION
    """
    df = read_sql(conn, sql, [schema, table])
    return df["COLUMN_NAME"].astype(str).tolist()


def real_col(existing_cols: set[str], col: str) -> Optional[str]:
    lookup = {x.upper(): x for x in existing_cols}
    return lookup.get(col.upper())


def add_select(parts: list[str], existing_cols: set[str], alias: str, col: str, output_name: str) -> None:
    actual = real_col(existing_cols, col)
    if actual:
        parts.append(f"{alias}.{quote_ident(actual)} AS {quote_ident(output_name)}")


def resolve_hana_dsn(cli_value: str) -> str:
    dsn = clean(cli_value) or clean(os.getenv("SAP_HANA_DSN")) or clean(FETCH_SAP_HANA_DSN)
    if not dsn:
        raise SystemExit("SAP HANA DSN is empty. Set SAP_HANA_DSN or pass --sap-hana-dsn.")
    return dsn


def connect_hana(dsn: str) -> pyodbc.Connection:
    logger.info("Connecting SAP HANA...")
    return pyodbc.connect(dsn, timeout=60, autocommit=True)


def fetch_sap_po_items(
    conn: pyodbc.Connection,
    schema: str,
    client: str,
    purchasing_org: str,
    purchasing_group: str,
    plant_filter: str,
    exclude_deleted: bool,
) -> pd.DataFrame:
    ekko_cols = set(get_table_columns(conn, schema, "EKKO"))
    ekpo_cols = set(get_table_columns(conn, schema, "EKPO"))

    select_parts: list[str] = []
    add_select(select_parts, ekpo_cols, "p", "EBELN", "Purchasing Document")
    add_select(select_parts, ekpo_cols, "p", "EBELP", "Item")
    add_select(select_parts, ekko_cols, "h", "EKORG", "Purchasing Organization")
    add_select(select_parts, ekko_cols, "h", "EKGRP", "Purchasing Group")
    add_select(select_parts, ekpo_cols, "p", "WERKS", "Plant")
    add_select(select_parts, ekko_cols, "h", "BSART", "Document Type")
    add_select(select_parts, ekko_cols, "h", "BEDAT", "Document Date")
    add_select(select_parts, ekko_cols, "h", "AEDAT", "Changed On")
    add_select(select_parts, ekko_cols, "h", "ERNAM", "Created By")
    add_select(select_parts, ekko_cols, "h", "LIFNR", "Supplier")
    add_select(select_parts, ekpo_cols, "p", "MATNR", "Material")
    add_select(select_parts, ekpo_cols, "p", "TXZ01", "Short Text")
    add_select(select_parts, ekpo_cols, "p", "MATKL", "Material Group")
    add_select(select_parts, ekpo_cols, "p", "KNTTP", "Account Assignment Category")
    add_select(select_parts, ekpo_cols, "p", "PSTYP", "Item Category")
    add_select(select_parts, ekpo_cols, "p", "MENGE", "Order Quantity")
    add_select(select_parts, ekpo_cols, "p", "MEINS", "Order Unit")
    add_select(select_parts, ekpo_cols, "p", "NETPR", "Net Price")
    add_select(select_parts, ekpo_cols, "p", "PEINH", "Price Unit")
    add_select(select_parts, ekpo_cols, "p", "NETWR", "Net Order Value")
    add_select(select_parts, ekpo_cols, "p", "BRTWR", "Gross Order Value")
    add_select(select_parts, ekko_cols, "h", "WAERS", "Currency")
    add_select(select_parts, ekpo_cols, "p", "MWSKZ", "Tax Code")
    add_select(select_parts, ekko_cols, "h", "LOEKZ", "Header Deletion Indicator")
    add_select(select_parts, ekpo_cols, "p", "LOEKZ", "Item Deletion Indicator")
    add_select(select_parts, ekpo_cols, "p", "ELIKZ", "Delivery Completed")
    add_select(select_parts, ekpo_cols, "p", "EREKZ", "Final Invoice")
    add_select(select_parts, ekpo_cols, "p", "LGORT", "Storage Location")

    if "Short Text" not in [part.split(" AS ")[-1].replace('"', "") for part in select_parts]:
        raise RuntimeError("EKPO.TXZ01 / Short Text is not available in the SAP query result.")

    where_parts = [
        f'h."MANDT" = \'{sql_quote(client)}\'',
        f'h."EKORG" = \'{sql_quote(purchasing_org)}\'',
        f'h."EKGRP" = \'{sql_quote(purchasing_group)}\'',
    ]
    if plant_filter:
        where_parts.append(f'p."WERKS" = \'{sql_quote(plant_filter)}\'')
    if exclude_deleted:
        if real_col(ekko_cols, "LOEKZ"):
            where_parts.append("COALESCE(h.\"LOEKZ\", '') = ''")
        if real_col(ekpo_cols, "LOEKZ"):
            where_parts.append("COALESCE(p.\"LOEKZ\", '') = ''")

    sql = f"""
    SELECT
        {", ".join(select_parts)}
    FROM {full_table(schema, "EKPO")} p
    INNER JOIN {full_table(schema, "EKKO")} h
        ON h."MANDT" = p."MANDT"
       AND h."EBELN" = p."EBELN"
    WHERE {" AND ".join(where_parts)}
    ORDER BY p."EBELN", p."EBELP"
    """

    logger.info(
        "Fetching SAP PO items: schema=%s client=%s EKORG=%s EKGRP=%s plant=%s excludeDeleted=%s",
        schema,
        client,
        purchasing_org,
        purchasing_group,
        plant_filter or "(blank)",
        exclude_deleted,
    )
    df = read_sql(conn, sql)
    logger.info("SAP PO item rows fetched: %s", len(df))
    return df


def extract_ticket_number(short_text: Any) -> tuple[str, str]:
    text = clean(short_text)
    if not text:
        return "", "Short Text is blank"

    matches: list[str] = []
    matches.extend(TICKET_NO_PATTERN.findall(text))
    matches.extend(TICKET_BRACKET_PATTERN.findall(text))
    matches.extend(TICKET_OUTER_BRACKET_PATTERN.findall(text))

    unique: list[str] = []
    seen: set[str] = set()
    for match in matches:
        ticket_id = normalize_ticket_id(match)
        if ticket_id and ticket_id not in seen:
            unique.append(ticket_id)
            seen.add(ticket_id)

    if len(unique) == 1:
        return unique[0], ""
    if len(unique) > 1:
        return "", "Multiple ticket numbers found: " + ", ".join(unique)
    if re.search(r"\btickets?\b", text, flags=re.IGNORECASE):
        return "", "Contains Ticket word but not standard Ticket No. number or Ticket [number] pattern"
    return "", "No standard Ticket No. number or Ticket [number] pattern"


def add_ticket_number_to_po(po_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if po_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    if "Short Text" not in po_df.columns:
        raise RuntimeError("SAP PO result has no Short Text column.")

    work = po_df.copy()
    extracted = work["Short Text"].apply(extract_ticket_number)
    work["Ticket Number"] = extracted.apply(lambda item: item[0])
    work["Exception Reason"] = extracted.apply(lambda item: item[1])
    work["Ticket Match Source"] = work["Ticket Number"].apply(lambda v: "short_text_ticket_number" if clean(v) else "")

    cols = list(work.columns)
    cols.remove("Ticket Number")
    cols.remove("Exception Reason")
    cols.remove("Ticket Match Source")
    short_text_idx = cols.index("Short Text")
    cols.insert(short_text_idx + 1, "Ticket Number")
    cols.insert(short_text_idx + 2, "Ticket Match Source")
    cols.append("Exception Reason")
    work = work[cols]

    regular = work[work["Ticket Number"].fillna("").astype(str).str.strip() != ""].copy()
    exceptions = work[work["Ticket Number"].fillna("").astype(str).str.strip() == ""].copy()
    regular = regular.drop(columns=["Exception Reason"], errors="ignore")

    for col in ["Net Price", "Net Order Value", "Gross Order Value", "Order Quantity"]:
        if col in regular.columns:
            regular[col] = pd.to_numeric(regular[col], errors="coerce").fillna(0.0)
        if col in exceptions.columns:
            exceptions[col] = pd.to_numeric(exceptions[col], errors="coerce").fillna(0.0)

    logger.info("SAP PO rows with Ticket Number: %s", len(regular))
    logger.info("SAP PO Short Text exception rows: %s", len(exceptions))
    return regular, exceptions


def firebase_ticket_entries(source_root: str) -> Iterable[tuple[str, Any]]:
    node = db.reference(f"{source_root}/tickets").get() or {}
    if isinstance(node, list):
        return [(str(i), row) for i, row in enumerate(node) if row]
    if isinstance(node, dict):
        return list(node.items())
    return []


def load_approved_tickets(source_root: str, date_from: str, date_to: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for fallback_key, raw in firebase_ticket_entries(source_root):
        ticket, _roles, tid_raw = normalize_row(raw, fallback_key)
        ticket_id = normalize_ticket_id(tid_raw)
        if not ticket_id:
            continue
        if approval_decision(ticket) != "approved":
            continue

        approved_date = approval_decision_date(ticket)
        if date_from and (not approved_date or approved_date < date_from):
            continue
        if date_to and (not approved_date or approved_date > date_to):
            continue

        rows.append(
            {
                "Ticket Number": ticket_id,
                "TicketID": ticket_id,
                "Approved Date": approved_date,
                "Ticket Status": clean(ticket.get("TicketStatus")),
                "Ticket Status Text": clean(ticket.get("TicketStatusText")),
                "Dealer ID": clean(ticket.get("DealerID") or ticket.get("WarrantyHandlingDealerID")),
                "Dealer Name": clean(ticket.get("DealerName")),
                "C4C AmountIncludingTax": parse_amount(ticket.get("AmountIncludingTax")),
                "ERP Purchase Order": clean(ticket.get("ERPPurchaseOrder")),
                "ERP Free Order": clean(ticket.get("ERPFreeOrder")),
                "Sales Order": clean(ticket.get("Sales Order") or ticket.get("SalesOrder") or ticket.get("LookupSalesOrder")),
                "Created On": clean(ticket.get("CreatedOn") or ticket.get("createdOn")),
                "Chassis Number": clean(ticket.get("ChassisNumber") or ticket.get("chassis") or ticket.get("Ticket Chassis Number")),
                "Claim Approved On": clean(
                    ticket.get("ClaimApprovedOnDateTime")
                    or ticket.get("ClaimApprovedOnDate")
                    or ticket.get("ClaimApprovedOn")
                    or ticket.get("Claim Approved On")
                ),
            }
        )

    df = pd.DataFrame(rows, columns=APPROVED_COLUMNS)
    if not df.empty:
        df = df.drop_duplicates(subset=["Ticket Number"], keep="first").sort_values(["Approved Date", "Ticket Number"])
    logger.info("Approved C4C tickets loaded: %s", len(df))
    return df


def load_ticket_status_lookup(source_root: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for fallback_key, raw in firebase_ticket_entries(source_root):
        ticket, _roles, tid_raw = normalize_row(raw, fallback_key)
        ticket_id = normalize_ticket_id(tid_raw)
        if not ticket_id:
            continue
        status_code = clean(ticket.get("TicketStatus") or ticket.get("statusCode") or ticket.get("Status")).upper()
        status_text = clean(ticket.get("TicketStatusText") or ticket.get("statusText") or ticket.get("StatusText"))
        if status_code and status_text:
            current_status = f"{status_code} | {status_text}"
        else:
            current_status = status_text or status_code
        rows.append(
            {
                "Ticket Number": ticket_id,
                "C4C Current Status": current_status,
                "C4C Status Code": status_code,
                "C4C Status Text": status_text,
            }
        )

    df = pd.DataFrame(rows, columns=["Ticket Number", "C4C Current Status", "C4C Status Code", "C4C Status Text"])
    if not df.empty:
        df["Ticket Number"] = df["Ticket Number"].fillna("").astype(str).str.strip()
        df = df[df["Ticket Number"] != ""].drop_duplicates(subset=["Ticket Number"], keep="last")
    logger.info("Ticket status lookup loaded: %s", len(df))
    return df


def aggregate_po_by_ticket(regular_po: pd.DataFrame, amount_column: str) -> pd.DataFrame:
    if regular_po.empty:
        return pd.DataFrame(columns=PO_AGG_COLUMNS)
    if amount_column not in regular_po.columns:
        raise RuntimeError(f"SAP PO result does not have amount column: {amount_column}")

    work = regular_po.copy()
    work[amount_column] = pd.to_numeric(work[amount_column], errors="coerce").fillna(0.0)
    if "Net Price" in work.columns:
        work["Net Price"] = pd.to_numeric(work["Net Price"], errors="coerce").fillna(0.0)
    if "Net Order Value" in work.columns:
        work["Net Order Value"] = pd.to_numeric(work["Net Order Value"], errors="coerce").fillna(0.0)

    def unique_join(values: pd.Series) -> str:
        out: list[str] = []
        seen: set[str] = set()
        for value in values:
            text = clean(value)
            if text and text not in seen:
                out.append(text)
                seen.add(text)
        return ", ".join(out[:20])

    grouped = (
        work.groupby("Ticket Number", dropna=False)
        .agg(
            SAP_PO_Item_Count=("Ticket Number", "size"),
            SAP_PO_Amount=(amount_column, "sum"),
            SAP_Net_Order_Value=("Net Order Value", "sum") if "Net Order Value" in work.columns else (amount_column, "sum"),
            SAP_Net_Price_Sum=("Net Price", "sum") if "Net Price" in work.columns else (amount_column, "sum"),
            SAP_PO_Documents=("Purchasing Document", unique_join) if "Purchasing Document" in work.columns else ("Ticket Number", unique_join),
            SAP_Currencies=("Currency", unique_join) if "Currency" in work.columns else ("Ticket Number", unique_join),
            SAP_Short_Text_Sample=("Short Text", unique_join) if "Short Text" in work.columns else ("Ticket Number", unique_join),
        )
        .reset_index()
    )
    grouped["SAP_PO_Amount"] = grouped["SAP_PO_Amount"].round(2)
    grouped["SAP_Net_Order_Value"] = grouped["SAP_Net_Order_Value"].round(2)
    grouped["SAP_Net_Price_Sum"] = grouped["SAP_Net_Price_Sum"].round(2)
    return grouped


def add_c4c_status_to_po_rows(po_rows: pd.DataFrame, ticket_status_lookup_df: pd.DataFrame) -> pd.DataFrame:
    if po_rows.empty:
        return po_rows.copy()
    if ticket_status_lookup_df.empty or "Ticket Number" not in po_rows.columns or "Ticket Number" not in ticket_status_lookup_df.columns:
        out = po_rows.copy()
        if "C4C Current Status" not in out.columns:
            out["C4C Current Status"] = ""
        return out

    lookup = ticket_status_lookup_df[
        [col for col in ["Ticket Number", "C4C Current Status"] if col in ticket_status_lookup_df.columns]
    ].copy()
    if lookup.empty or "Ticket Number" not in lookup.columns:
        out = po_rows.copy()
        if "C4C Current Status" not in out.columns:
            out["C4C Current Status"] = ""
        return out

    lookup["Ticket Number"] = lookup["Ticket Number"].fillna("").astype(str).str.strip()
    lookup = lookup[lookup["Ticket Number"] != ""].drop_duplicates(subset=["Ticket Number"], keep="last")

    out = po_rows.copy()
    out["Ticket Number"] = out["Ticket Number"].fillna("").astype(str).str.strip()
    out = out.merge(lookup, how="left", on="Ticket Number")
    out["C4C Current Status"] = out["C4C Current Status"].fillna("")

    cols = list(out.columns)
    if "C4C Current Status" in cols:
        cols.remove("C4C Current Status")
        insert_at = cols.index("Ticket Number") + 1 if "Ticket Number" in cols else len(cols)
        cols.insert(insert_at, "C4C Current Status")
        out = out[cols]
    return out


def build_outputs(
    approved_df: pd.DataFrame,
    approved_membership_df: pd.DataFrame,
    ticket_status_lookup_df: pd.DataFrame,
    regular_po: pd.DataFrame,
    exceptions: pd.DataFrame,
    amount_column: str,
    date_from: str = "",
    date_to: str = "",
) -> dict[str, pd.DataFrame]:
    po_by_ticket = aggregate_po_by_ticket(regular_po, amount_column)
    summary = approved_df.merge(po_by_ticket, how="left", on="Ticket Number")
    if not summary.empty:
        for col in PO_AGG_COLUMNS:
            if col not in summary.columns:
                summary[col] = 0 if col in {"SAP_PO_Item_Count", "SAP_PO_Amount", "SAP_Net_Order_Value", "SAP_Net_Price_Sum"} else ""
        summary["SAP Match Status"] = summary["SAP_PO_Amount"].notna().map({True: "Matched", False: "No PO Match"})
        summary["SAP_PO_Amount"] = summary["SAP_PO_Amount"].fillna(0.0).round(2)
        summary["SAP_Net_Order_Value"] = summary["SAP_Net_Order_Value"].fillna(0.0).round(2)
        summary["SAP_Net_Price_Sum"] = summary["SAP_Net_Price_Sum"].fillna(0.0).round(2)
        summary["SAP_PO_Item_Count"] = summary["SAP_PO_Item_Count"].fillna(0).astype(int)
        summary["C4C AmountIncludingTax"] = summary["C4C AmountIncludingTax"].fillna(0.0).round(2)
        summary["Delta SAP Amount - C4C Amount"] = (
            summary["SAP_PO_Amount"] - summary["C4C AmountIncludingTax"]
        ).round(2)
    summary = add_c4c_status_to_po_rows(summary, ticket_status_lookup_df)

    approved_ids = set(approved_df["Ticket Number"].astype(str).tolist()) if not approved_df.empty else set()
    approved_membership_ids = (
        set(approved_membership_df["Ticket Number"].astype(str).tolist())
        if not approved_membership_df.empty
        else set()
    )
    if regular_po.empty or "Ticket Number" not in regular_po.columns:
        matched_items = pd.DataFrame(columns=list(regular_po.columns))
        approved_outside_window = pd.DataFrame(columns=list(regular_po.columns))
        po_not_approved = pd.DataFrame(columns=list(regular_po.columns))
    else:
        matched_items = regular_po[regular_po["Ticket Number"].astype(str).isin(approved_ids)].copy()
        approved_outside_window = regular_po[
            regular_po["Ticket Number"].astype(str).isin(approved_membership_ids - approved_ids)
        ].copy()
        po_not_approved = regular_po[
            ~regular_po["Ticket Number"].astype(str).isin(approved_membership_ids)
        ].copy()
    matched_items = add_c4c_status_to_po_rows(matched_items, ticket_status_lookup_df)
    approved_outside_window = add_c4c_status_to_po_rows(approved_outside_window, ticket_status_lookup_df)
    po_not_approved = add_c4c_status_to_po_rows(po_not_approved, ticket_status_lookup_df)
    exceptions = add_c4c_status_to_po_rows(exceptions, ticket_status_lookup_df)
    no_match = summary[summary["SAP Match Status"] == "No PO Match"].copy() if not summary.empty else pd.DataFrame()

    outside_window_metric = len(approved_outside_window)
    outside_window_label = "SAP PO rows approved outside selected date range"
    if not (date_from or date_to):
        outside_window_metric = 0
        outside_window_label = "SAP PO rows approved outside selected date range (date filter not used)"

    totals = pd.DataFrame(
        [
            {"Metric": "Approved tickets", "Value": len(approved_df)},
            {"Metric": "Approved tickets matched to SAP PO Short Text", "Value": int((summary["SAP Match Status"] == "Matched").sum()) if not summary.empty else 0},
            {"Metric": "Approved tickets without SAP PO match", "Value": len(no_match)},
            {"Metric": f"Approved cost from SAP PO {amount_column}", "Value": round(float(summary["SAP_PO_Amount"].sum()) if not summary.empty else 0.0, 2)},
            {"Metric": "Approved cost from C4C AmountIncludingTax", "Value": round(float(summary["C4C AmountIncludingTax"].sum()) if not summary.empty else 0.0, 2)},
            {"Metric": "SAP PO rows with ticket number", "Value": len(regular_po)},
            {"Metric": "SAP PO short text exception rows", "Value": len(exceptions)},
            {"Metric": outside_window_label, "Value": outside_window_metric},
            {"Metric": "SAP PO rows with ticket number but not approved in C4C", "Value": len(po_not_approved)},
        ]
    )

    return {
        "Run_Summary": totals,
        "Approved_PO_Cost_Summary": summary,
        "Matched_PO_Items": matched_items,
        "Approved_No_PO_Match": no_match,
        "PO_Approved_Outside_Date_Range": approved_outside_window,
        "PO_Ticket_Not_Approved": po_not_approved,
        "Short_Text_Exceptions": exceptions,
    }


def write_excel(path: str, sheets: dict[str, pd.DataFrame]) -> None:
    logger.info("Writing Excel: %s", path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.sheets[safe_name]
            ws.freeze_panes = "A2"
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            for col_cells in ws.columns:
                header = col_cells[0].value
                col_letter = col_cells[0].column_letter
                max_len = 10
                for cell in col_cells[:250]:
                    if cell.value is not None:
                        max_len = max(max_len, min(len(str(cell.value)), 55))
                width = max_len + 2
                if header in {"Short Text", "SAP_Short_Text_Sample"}:
                    width = min(max(width, 35), 60)
                else:
                    width = min(width, 34)
                ws.column_dimensions[col_letter].width = width

            for col_idx, header in enumerate([cell.value for cell in ws[1]], start=1):
                if header and any(token in str(header) for token in ["Amount", "Value", "Price", "Quantity", "Delta"]):
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        thin = Side(style="thin", color="DDDDDD")
        for ws in writer.sheets.values():
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin)
    logger.info("Excel saved: %s", path)


def valid_date(value: str) -> str:
    value = clean(value)
    if not value:
        return ""
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"Date must be YYYY-MM-DD: {value}") from exc
    return value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Calculate approved ticket cost from SAP PO Short Text ticket numbers.")
    parser.add_argument("--output", default=os.getenv("OUTPUT_FILE", DEFAULT_OUTPUT))
    parser.add_argument("--log-file", default=os.getenv("LOG_FILE", "approved_cost_from_sap_po_short_text.log"))
    parser.add_argument("--firebase-db-url", default=os.getenv("FIREBASE_DB_URL", DEFAULT_DB_URL))
    parser.add_argument("--firebase-sa-path", default=os.getenv("FIREBASE_SA_PATH", str(Path.cwd() / "firebase-service-account.json")))
    parser.add_argument("--source-root", default=os.getenv("SOURCE_ROOT", DEFAULT_SOURCE_ROOT))
    parser.add_argument("--sap-hana-dsn", default=os.getenv("SAP_HANA_DSN", ""))
    parser.add_argument("--sap-schema", default=os.getenv("SAP_SCHEMA", DEFAULT_SCHEMA))
    parser.add_argument("--sap-client", default=os.getenv("SAP_CLIENT", DEFAULT_CLIENT))
    parser.add_argument("--purchasing-org", default=os.getenv("PURCHASING_ORG", DEFAULT_PURCHASING_ORG))
    parser.add_argument("--purchasing-group", default=os.getenv("PURCHASING_GROUP", DEFAULT_PURCHASING_GROUP))
    parser.add_argument("--po-plant-filter", default=os.getenv("PO_PLANT_FILTER", ""))
    parser.add_argument("--exclude-deleted-po", action="store_true", default=os.getenv("EXCLUDE_DELETED_PO", "").lower() in {"1", "true", "yes", "y"})
    parser.add_argument("--date-from", type=valid_date, default=os.getenv("APPROVED_DATE_FROM", ""))
    parser.add_argument("--date-to", type=valid_date, default=os.getenv("APPROVED_DATE_TO", ""))
    parser.add_argument(
        "--amount-column",
        default=os.getenv("SAP_PO_AMOUNT_COLUMN", DEFAULT_AMOUNT_COLUMN),
        help="SAP PO column used as approved cost. Default: Net Order Value.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    setup_logging(args.log_file)

    logger.info("Started approved cost from SAP PO Short Text proof-of-concept.")
    logger.info("Output file: %s", args.output)
    logger.info("Approved date filter: %s to %s", args.date_from or "(beginning)", args.date_to or "(latest)")
    logger.info("SAP amount column for approved cost: %s", args.amount_column)

    init_firebase(args.firebase_db_url, args.firebase_sa_path)
    ticket_status_lookup_df = load_ticket_status_lookup(args.source_root)
    approved_membership_df = load_approved_tickets(args.source_root, "", "")
    approved_df = approved_membership_df
    if args.date_from or args.date_to:
        approved_df = load_approved_tickets(args.source_root, args.date_from, args.date_to)

    hana_dsn = resolve_hana_dsn(args.sap_hana_dsn)
    with connect_hana(hana_dsn) as conn:
        po_df = fetch_sap_po_items(
            conn,
            schema=args.sap_schema,
            client=args.sap_client,
            purchasing_org=args.purchasing_org,
            purchasing_group=args.purchasing_group,
            plant_filter=clean(args.po_plant_filter),
            exclude_deleted=bool(args.exclude_deleted_po),
        )

    regular_po, exceptions = add_ticket_number_to_po(po_df)
    sheets = build_outputs(
        approved_df,
        approved_membership_df,
        ticket_status_lookup_df,
        regular_po,
        exceptions,
        args.amount_column,
        date_from=args.date_from,
        date_to=args.date_to,
    )
    write_excel(args.output, sheets)

    summary = sheets["Run_Summary"]
    for _, row in summary.iterrows():
        logger.info("%s: %s", row["Metric"], row["Value"])
    logger.info("Done.")


if __name__ == "__main__":
    main()
