#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Enrich the delivery-flow "Open Parts" export with the latest per-material
prices from three SAP sources:

    3090 SO      -- latest VBAP line under VKORG='3090'   (CN sales price)
    3110 PO      -- latest EKPO line under EKORG='3111'   (AU purchase price)
    3091 PO      -- latest EKPO line under EKORG='3091'   (CN factory purchase price)

For each source we add four columns right after "Amount Including Tax":
    <label> Latest Net Price      -- unit price (NETPR / PEINH or KPEIN)
    <label> Latest Currency       -- WAERS (PO) or WAERK (SO)
    <label> Latest Date           -- AEDAT (PO) or ERDAT (SO), rendered YYYY-MM-DD
    <label> Latest Doc Number     -- EBELN (PO) or VBELN (SO), for click-through

Design notes
------------
Match is BY MATERIAL only, not by PO/SO number. Investigation on this HANA
instance showed there is no 1:1 link between AU warranty POs (EKORG=3111,
doctype ZCRM) and CN-side documents; the intercompany relationship is
by-material, so the useful comparison is "what's the latest price this
material was transacted at, on each org".

Latest is chosen via ROW_NUMBER() OVER (PARTITION BY MATNR ORDER BY date DESC,
doc DESC, item DESC). Deleted / rejected lines are excluded (LOEKZ blank
for POs, ABGRU blank for SOs).

Env matches fetch_all_tickets_fast_*.py:
    SAP_HANA_DSN, SAP_CLIENT (default "800")
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import xml.etree.ElementTree as ET
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


SAP_HANA_DSN = os.getenv(
    "SAP_HANA_DSN",
    "DRIVER={HDBODBC};SERVERNODE=10.11.2.25:30241;UID=BAOJIANFENG;PWD=Xja@2025ABC;",
)
SAP_CLIENT = os.getenv("SAP_CLIENT", "800")

SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
ITEM_SHEET = "Open Parts Item Detail"
INSERT_AFTER = "Amount Including Tax"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("enrich_open_parts")


# ---------------------------------------------------------------------------
# Source configuration
# ---------------------------------------------------------------------------

class Source:
    __slots__ = ("label", "kind", "org_value", "org_field")

    def __init__(self, label: str, kind: str, org_value: str, org_field: str):
        self.label = label        # column-name prefix, e.g. "3090 SO"
        self.kind = kind          # "SO" or "PO"
        self.org_value = org_value
        self.org_field = org_field


SOURCES: List[Source] = [
    Source(label="3090 SO", kind="SO", org_value="3090", org_field="VKORG"),
    Source(label="3110 PO", kind="PO", org_value="3111", org_field="EKORG"),
    Source(label="3091 PO", kind="PO", org_value="3091", org_field="EKORG"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"", "#", "nan", "none"} else s


def sql_quote(v: str) -> str:
    return str(v).replace("'", "''")


def chunks(lst: list, n: int) -> List[list]:
    return [lst[i:i + n] for i in range(0, len(lst), n)]


def fmt_sap_date(s: str) -> str:
    s = clean_str(s)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


# ---------------------------------------------------------------------------
# Read SpreadsheetML 2003 (.xls XML)
# ---------------------------------------------------------------------------

def read_spreadsheetml(path: str) -> List[Dict[str, object]]:
    tree = ET.parse(path)
    root = tree.getroot()
    sheets: List[Dict[str, object]] = []
    for ws in root.findall(f"{{{SS_NS}}}Worksheet"):
        name = ws.get(f"{{{SS_NS}}}Name", "")
        rows: List[List[str]] = []
        table = ws.find(f"{{{SS_NS}}}Table")
        if table is None:
            sheets.append({"name": name, "rows": rows})
            continue
        for row in table.findall(f"{{{SS_NS}}}Row"):
            cells: List[str] = []
            expected_index = 1
            for cell in row.findall(f"{{{SS_NS}}}Cell"):
                idx_attr = cell.get(f"{{{SS_NS}}}Index")
                if idx_attr:
                    idx = int(idx_attr)
                    while expected_index < idx:
                        cells.append("")
                        expected_index += 1
                data = cell.find(f"{{{SS_NS}}}Data")
                cells.append("" if data is None or data.text is None else data.text)
                expected_index += 1
            rows.append(cells)
        sheets.append({"name": name, "rows": rows})
    return sheets


# ---------------------------------------------------------------------------
# SAP queries
# ---------------------------------------------------------------------------

def _sql_so(materials_in_list: str) -> str:
    """Latest VBAP row per material under the given VKORG. KPEIN is the
    'price unit' scaling factor."""
    return f"""
WITH ranked AS (
    SELECT
        vbap."MATNR",
        vbap."VBELN"                              AS "DOCNUM",
        vbap."POSNR"                              AS "DOCITEM",
        vbak."WAERK"                              AS "WAERS",
        vbap."NETPR"                              AS "NETPR",
        COALESCE(NULLIF(vbap."KPEIN", 0), 1)      AS "PEINH",
        vbap."ERDAT"                              AS "DOCDATE",
        ROW_NUMBER() OVER (
            PARTITION BY vbap."MATNR"
            ORDER BY vbap."ERDAT" DESC, vbap."VBELN" DESC, vbap."POSNR" DESC
        ) AS rn
    FROM "SAPHANADB"."VBAP" vbap
    INNER JOIN "SAPHANADB"."VBAK" vbak
        ON vbak."MANDT" = vbap."MANDT"
       AND vbak."VBELN" = vbap."VBELN"
    WHERE vbap."MANDT" = '{sql_quote(SAP_CLIENT)}'
      AND vbak."{{ORG_FIELD}}" = '{{ORG_VALUE}}'
      AND vbap."MATNR" IN ('{materials_in_list}')
      AND (vbap."ABGRU" IS NULL OR vbap."ABGRU" = '')
)
SELECT "MATNR","DOCNUM","DOCITEM","WAERS","NETPR","PEINH","DOCDATE"
FROM ranked WHERE rn = 1
"""


def _sql_po(materials_in_list: str) -> str:
    """Latest EKPO row per material under the given EKORG. Skip deleted
    header or item lines. EKKO carries EKORG, WAERS and AEDAT; EKPO
    carries NETPR and PEINH."""
    return f"""
WITH ranked AS (
    SELECT
        ekpo."MATNR",
        ekpo."EBELN"                                  AS "DOCNUM",
        ekpo."EBELP"                                  AS "DOCITEM",
        ekko."WAERS"                                  AS "WAERS",
        ekpo."NETPR"                                  AS "NETPR",
        COALESCE(NULLIF(ekpo."PEINH", 0), 1)          AS "PEINH",
        ekko."AEDAT"                                  AS "DOCDATE",
        ROW_NUMBER() OVER (
            PARTITION BY ekpo."MATNR"
            ORDER BY ekko."AEDAT" DESC, ekpo."EBELN" DESC, ekpo."EBELP" DESC
        ) AS rn
    FROM "SAPHANADB"."EKPO" ekpo
    INNER JOIN "SAPHANADB"."EKKO" ekko
        ON ekko."MANDT" = ekpo."MANDT"
       AND ekko."EBELN" = ekpo."EBELN"
    WHERE ekpo."MANDT" = '{sql_quote(SAP_CLIENT)}'
      AND ekko."{{ORG_FIELD}}" = '{{ORG_VALUE}}'
      AND ekpo."MATNR" IN ('{materials_in_list}')
      AND (ekpo."LOEKZ" IS NULL OR ekpo."LOEKZ" = '')
      AND (ekko."LOEKZ" IS NULL OR ekko."LOEKZ" = '')
)
SELECT "MATNR","DOCNUM","DOCITEM","WAERS","NETPR","PEINH","DOCDATE"
FROM ranked WHERE rn = 1
"""


def fetch_latest_by_material(
    source: Source,
    materials: List[str],
    debug: bool = False,
) -> Dict[str, Dict[str, object]]:
    """Return {MATNR: {NETPR, WAERS, DOCDATE, DOCNUM}} for the latest doc
    of the given source per material."""
    unique_mats = sorted({clean_str(m) for m in materials if clean_str(m)})
    if not unique_mats:
        return {}

    out: Dict[str, Dict[str, object]] = {}
    batches = chunks(unique_mats, 500)
    template = _sql_so if source.kind == "SO" else _sql_po

    import pyodbc  # deferred: only required when actually querying
    with pyodbc.connect(SAP_HANA_DSN, autocommit=True) as conn:
        for i, batch in enumerate(batches, start=1):
            in_list = "','".join(sql_quote(m) for m in batch)
            sql = template(in_list).format(
                ORG_FIELD=source.org_field,
                ORG_VALUE=sql_quote(source.org_value),
            )
            logger.info("[%s] batch %s/%s (%s materials)", source.label, i, len(batches), len(batch))
            df = pd.read_sql(sql, conn)
            if debug and i == 1:
                logger.info("[%s] DEBUG batch 1 returned %s rows", source.label, len(df))
                for _, dr in df.head(5).iterrows():
                    logger.info("  %r", dr.to_dict())
            for _, row in df.iterrows():
                mat = clean_str(row["MATNR"])
                if not mat:
                    continue
                netpr = float(row.get("NETPR") or 0)
                peinh = float(row.get("PEINH") or 1) or 1
                out[mat] = {
                    "NETPR": netpr / peinh,
                    "WAERS": clean_str(row.get("WAERS")),
                    "DOCDATE": clean_str(row.get("DOCDATE")),
                    "DOCNUM": clean_str(row.get("DOCNUM")),
                }
    return out


# ---------------------------------------------------------------------------
# Enrich Item Detail
# ---------------------------------------------------------------------------

def enrich_item_detail(sheet: Dict[str, object], debug: bool = False) -> Tuple[List[str], List[List[object]], Dict[str, Dict[str, int]]]:
    rows: List[List[str]] = sheet["rows"]  # type: ignore[assignment]
    if not rows:
        return [], [], {}

    header = rows[0]
    header_idx = {name: i for i, name in enumerate(header)}
    for req in ("Material", INSERT_AFTER):
        if req not in header_idx:
            raise SystemExit(f"Column not found in {ITEM_SHEET!r}: {req}")

    materials: List[str] = []
    for r in rows[1:]:
        mat = clean_str(r[header_idx["Material"]] if header_idx["Material"] < len(r) else "")
        if mat:
            materials.append(mat)
    unique_mats = set(materials)
    logger.info("Item Detail rows: %s, unique materials: %s", len(rows) - 1, len(unique_mats))

    price_maps: Dict[str, Dict[str, Dict[str, object]]] = {}
    stats: Dict[str, Dict[str, int]] = {}
    for src in SOURCES:
        pm = fetch_latest_by_material(src, sorted(unique_mats), debug=debug) if unique_mats else {}
        price_maps[src.label] = pm
        stats[src.label] = {"materials_with_price": len(pm), "unique_materials": len(unique_mats)}
        logger.info("[%s] found for %s / %s materials", src.label, len(pm), len(unique_mats))

    new_cols: List[str] = []
    for src in SOURCES:
        new_cols += [
            f"{src.label} Latest Net Price",
            f"{src.label} Latest Currency",
            f"{src.label} Latest Date",
            f"{src.label} Latest Doc Number",
        ]
    insert_at = header_idx[INSERT_AFTER] + 1
    new_header = header[:insert_at] + new_cols + header[insert_at:]

    new_rows: List[List[object]] = [new_header]
    matched_row_any = 0
    for r in rows[1:]:
        r = list(r) + [""] * (len(header) - len(r))
        mat = clean_str(r[header_idx["Material"]])
        extras: List[object] = []
        row_any = False
        for src in SOURCES:
            slot = price_maps[src.label].get(mat) if mat else None
            if slot:
                row_any = True
                extras += [
                    round(float(slot["NETPR"]), 4),
                    slot["WAERS"],
                    fmt_sap_date(str(slot["DOCDATE"])),
                    slot["DOCNUM"],
                ]
            else:
                extras += ["", "", "", ""]
        if row_any:
            matched_row_any += 1
        new_rows.append(r[:insert_at] + extras + r[insert_at:])

    stats["_rows"] = {"input_rows": len(rows) - 1, "matched_any": matched_row_any}
    return new_header, new_rows, stats


# ---------------------------------------------------------------------------
# Write .xlsx (preserve leading zeros on ID/date columns)
# ---------------------------------------------------------------------------

def looks_numeric(v: object) -> bool:
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if not s:
        return False
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", s.replace(",", "")))


def to_number(v: object) -> object:
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return v


PRESERVE_TEXT_COLS = {
    "Ticket Key", "Ticket ID", "Sales Order", "SO Created Date", "Dealer ID",
    "Sales Order Item", "Material", "ERP Purchase Order", "First Issue Date",
}
for _src in SOURCES:
    PRESERVE_TEXT_COLS.add(f"{_src.label} Latest Date")
    PRESERVE_TEXT_COLS.add(f"{_src.label} Latest Doc Number")


def write_xlsx(sheets: List[Dict[str, object]], out_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="EEF6FF", end_color="EEF6FF")
    for sheet in sheets:
        name = str(sheet["name"])[:31] or "Sheet"
        ws = wb.create_sheet(title=name)
        rows: List[List[object]] = sheet["rows"]  # type: ignore[assignment]
        header_names: List[str] = [str(v) for v in rows[0]] if rows else []
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if r_idx == 1:
                    cell.value = val
                    cell.font = header_font
                    cell.fill = header_fill
                    continue
                col_name = header_names[c_idx - 1] if c_idx - 1 < len(header_names) else ""
                if col_name in PRESERVE_TEXT_COLS:
                    cell.value = val
                else:
                    cell.value = to_number(val) if looks_numeric(val) else val
    wb.save(out_path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    logger.info("Reading %s", args.input)
    sheets = read_spreadsheetml(args.input)
    logger.info("Sheets: %s", [s["name"] for s in sheets])

    item_sheet = next((s for s in sheets if s["name"] == ITEM_SHEET), None)
    if item_sheet is None:
        raise SystemExit(f"Sheet not found: {ITEM_SHEET!r}")

    if args.dry_run:
        rows = item_sheet["rows"]
        header = rows[0] if rows else []
        idx = {n: i for i, n in enumerate(header)}
        mats = {clean_str(r[idx["Material"]]) for r in rows[1:] if idx["Material"] < len(r) and clean_str(r[idx["Material"]])}
        logger.info("Dry-run: %s rows, %s unique materials", len(rows) - 1, len(mats))
        return

    new_header, new_rows, stats = enrich_item_detail(item_sheet, debug=args.debug)
    item_sheet["rows"] = new_rows
    logger.info("Summary: %s", stats)

    write_xlsx(sheets, args.output)
    logger.info("Wrote %s", args.output)


if __name__ == "__main__":
    main()
