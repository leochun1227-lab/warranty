import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    return "" if value is None else str(value).strip()


def column_has_data_below_header(sheet, column, header_row):
    for row in range(header_row + 1, sheet.max_row + 1):
        if clean(sheet.cell(row, column).value):
            return True
    return False


def trim_sheet(sheet, header_row=1):
    deleted = []
    for column in range(sheet.max_column, 0, -1):
        header = clean(sheet.cell(header_row, column).value)
        if not header:
            continue
        if column_has_data_below_header(sheet, column, header_row):
            continue
        deleted.append({"column": column, "header": header})
        sheet.delete_cols(column, 1)
    deleted.reverse()
    return deleted


def main():
    parser = argparse.ArgumentParser(description="Remove spreadsheet columns where the header exists but every data cell is blank.")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook)
    report = {}
    for sheet in workbook.worksheets:
        deleted = trim_sheet(sheet)
        if deleted:
            report[sheet.title] = deleted
    workbook.save(args.workbook)
    print(json.dumps({"workbook": str(args.workbook), "deleted": report}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
