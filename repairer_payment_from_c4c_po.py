# -*- coding: utf-8 -*-
"""
Compare repairer payment by C4C ticket -> C4C PO -> SAP PO amount.

This script is intentionally standalone and does not change any dashboard logic.
It can start from a Ticket Detail workbook export or directly from current C4C
Firebase tickets, then:

1. Reads C4C ticket numbers.
2. Looks up each ticket's C4C ERP Purchase Order and AmountIncludingTax.
3. Fetches SAP PO rows by those C4C PO numbers.
4. Sums SAP PO amounts per ticket.
5. Shows Short Text plus extracted ticket number beside the original ticket.
6. Exports an Excel workbook for manual comparison.
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import warnings
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook

from approved_cost_from_sap_po_short_text import (
    DEFAULT_AMOUNT_COLUMN,
    DEFAULT_CLIENT,
    DEFAULT_DB_URL,
    DEFAULT_SCHEMA,
    DEFAULT_SOURCE_ROOT,
    add_select,
    clean,
    connect_hana,
    extract_ticket_number,
    full_table,
    get_table_columns,
    init_firebase,
    normalize_ticket_id,
    parse_amount,
    read_sql,
    real_col,
    resolve_hana_dsn,
    setup_logging,
    sql_quote,
    write_excel,
)
from ctm_v44_history_safe_mandt800_rejection_filter import get_field, normalize_po_number, normalize_row
from firebase_admin import db


warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable.*",
    category=UserWarning,
)


DEFAULT_OUTPUT = "repairer_payment_from_c4c_po.xlsx"
DEFAULT_TICKET_DETAIL_FILE = Path("outputs") / "claim_ytd_comparison" / "claim_ytd_comparison_tickets_detail_latest.xlsx"

logger = logging.getLogger("repairer_payment_from_c4c_po")


def text_or_blank(value: Any) -> str:
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = clean(value)
    return "" if text.lower() in {"nan", "none", "null"} else text


def normalize_ticket_detail_sheet_name(value: str) -> str:
    return text_or_blank(value)


def detect_header_row(ws) -> int:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = {clean(cell).lower().replace(" ", "") for cell in row if clean(cell)}
        if "ticketid" in headers or "ticketnumber" in headers:
            return row_idx
    raise RuntimeError(f"Could not find a TicketID/Ticket Number header row in sheet {ws.title!r}.")


def sanitize_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for idx, value in enumerate(values, start=1):
        base = clean(value) or f"Column{idx}"
        count = used.get(base, 0)
        used[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def load_ticket_detail_rows(path: str, sheet_name: str = "") -> pd.DataFrame:
    ticket_path = Path(path)
    if not ticket_path.exists():
        logger.warning("Ticket detail input file not found: %s", ticket_path)
        return pd.DataFrame()

    wb = load_workbook(ticket_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row_idx = detect_header_row(ws)
        header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
        headers = sanitize_headers(list(header_cells))

        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            ticket_id = normalize_ticket_id(get_field(raw, ["Ticket Number", "TicketID", "Ticket Id"]))
            if not ticket_id:
                continue
            rows.append(
                {
                    "Ticket Number": ticket_id,
                    "Ticket Detail Claim Scope": get_field(raw, ["ClaimScope", "Claim Scope"]),
                    "Ticket Detail Status Group": get_field(raw, ["StatusGroup", "Status Group"]),
                    "Ticket Detail Status Text": get_field(raw, ["StatusText", "Status Text"]),
                    "Ticket Detail Status Code": get_field(raw, ["StatusCode", "Status Code"]),
                    "Ticket Detail Created On": get_field(raw, ["CreatedOn", "Created On"]),
                    "Ticket Detail Claim Approved On": get_field(raw, ["ClaimApprovedOn", "Claim Approved On"]),
                    "Ticket Detail Resolved On": get_field(raw, ["ResolvedOn", "Resolved On"]),
                    "Ticket Detail ERP Purchase Order": normalize_po_number(
                        get_field(raw, ["ERPPurchaseOrder", "ERP Purchase Order", "ERP Purchase Order ID", "Purchasing Document"])
                    ),
                    "Ticket Detail PO Net Value": parse_amount(
                        get_field(raw, ["PoNetValue", "PO Net Value", "NetValue", "Net Value"])
                    ),
                    "Ticket Detail Claim Amount Value": parse_amount(
                        get_field(raw, ["ClaimAmountValue", "Claim Amount Value", "AmountIncludingTax", "Amount Including Tax"])
                    ),
                    "Ticket Detail Dealer Name": get_field(raw, ["DealerName", "Dealer Name"]),
                    "Ticket Detail Ticket Type": get_field(raw, ["TicketType", "Ticket Type"]),
                    "Ticket Detail Ticket Type Text": get_field(raw, ["TicketTypeText", "Ticket Type Text"]),
                }
            )
    finally:
        wb.close()

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.drop_duplicates(subset=["Ticket Number"], keep="first").reset_index(drop=True)
    logger.info("Ticket detail rows loaded: %s", len(df))
    return df


def firebase_ticket_entries(source_root: str) -> Iterable[tuple[str, Any]]:
    node = db.reference(f"{source_root}/tickets").get() or {}
    if isinstance(node, list):
        return [(str(i), row) for i, row in enumerate(node) if row]
    if isinstance(node, dict):
        return list(node.items())
    return []


def load_c4c_ticket_lookup(source_root: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for fallback_key, raw in firebase_ticket_entries(source_root):
        ticket, _roles, tid_raw = normalize_row(raw, fallback_key)
        ticket_id = normalize_ticket_id(tid_raw)
        if not ticket_id:
            continue
        rows.append(
            {
                "Ticket Number": ticket_id,
                "C4C ERP Purchase Order": normalize_po_number(
                    get_field(ticket, ["ERPPurchaseOrder", "ERP Purchase Order", "ERP Purchase Order ID", "Purchasing Document"])
                ),
                "C4C Sales Order": get_field(ticket, ["Sales Order", "SalesOrder", "LookupSalesOrder"]),
                "C4C AmountIncludingTax": parse_amount(
                    ticket.get("AmountIncludingTax")
                    or ticket.get("Amount Including Tax")
                    or ticket.get("ClaimAmountValue")
                    or ticket.get("ClaimTotalAmount")
                ),
                "C4C Chassis Number": get_field(ticket, ["ChassisNumber", "Chassis Number", "Ticket Chassis Number"]),
                "C4C Dealer Name": get_field(ticket, ["DealerName", "Dealer Name"]),
                "C4C Dealer ID": get_field(ticket, ["DealerID", "Dealer ID", "WarrantyHandlingDealerID"]),
                "C4C Ticket Status": get_field(ticket, ["TicketStatus", "StatusCode", "Status"]),
                "C4C Ticket Status Text": get_field(ticket, ["TicketStatusText", "StatusText", "Status"]),
                "C4C Created On": get_field(ticket, ["CreatedOn", "Created On", "createdOn"]),
                "C4C Claim Approved On": get_field(
                    ticket,
                    ["ClaimApprovedOnDateTime", "ClaimApprovedOnDate", "ClaimApprovedOn", "Claim Approved On"],
                ),
                "C4C Resolved On": get_field(ticket, ["ResolvedOnDateTime", "ResolvedOnDate", "ResolvedOn", "Resolved On"]),
                "C4C Ticket Type": get_field(ticket, ["TicketType", "Ticket Type"]),
                "C4C Ticket Type Text": get_field(ticket, ["TicketTypeText", "Ticket Type Text"]),
            }
        )

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.drop_duplicates(subset=["Ticket Number"], keep="last").reset_index(drop=True)
    logger.info("C4C ticket lookup loaded: %s", len(df))
    return df


def first_nonempty_text(*values: Any) -> str:
    for value in values:
        text = text_or_blank(value)
        if text:
            return text
    return ""


def first_positive_amount(*values: Any) -> float:
    for value in values:
        amount = parse_amount(value)
        if amount > 0:
            return round(amount, 2)
    return 0.0


def split_po_numbers(value: Any) -> list[str]:
    raw = text_or_blank(value)
    if not raw:
        return []

    number_matches = re.findall(r"\b\d{6,}\b", raw)
    out: list[str] = []
    seen: set[str] = set()
    for token in number_matches:
        po = normalize_po_number(token)
        if po and po not in seen:
            out.append(po)
            seen.add(po)
    if out:
        return out
    po = normalize_po_number(raw)
    if not re.fullmatch(r"\d{6,}", po):
        return []
    return [po] if po else []


def build_base_ticket_rows(ticket_detail_df: pd.DataFrame, c4c_lookup_df: pd.DataFrame) -> pd.DataFrame:
    if ticket_detail_df.empty and c4c_lookup_df.empty:
        return pd.DataFrame()

    if ticket_detail_df.empty:
        base = c4c_lookup_df.copy()
        base["Ticket Detail Claim Scope"] = ""
        base["Ticket Detail Status Group"] = ""
        base["Ticket Detail Status Text"] = ""
        base["Ticket Detail Status Code"] = ""
        base["Ticket Detail Created On"] = ""
        base["Ticket Detail Claim Approved On"] = ""
        base["Ticket Detail Resolved On"] = ""
        base["Ticket Detail ERP Purchase Order"] = ""
        base["Ticket Detail PO Net Value"] = 0.0
        base["Ticket Detail Claim Amount Value"] = 0.0
        base["Ticket Detail Dealer Name"] = ""
        base["Ticket Detail Ticket Type"] = ""
        base["Ticket Detail Ticket Type Text"] = ""
    else:
        base = ticket_detail_df.merge(c4c_lookup_df, how="left", on="Ticket Number")

    base["C4C ERP Purchase Order"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C ERP Purchase Order"), row.get("Ticket Detail ERP Purchase Order")),
        axis=1,
    )
    base["C4C AmountIncludingTax"] = base.apply(
        lambda row: first_positive_amount(row.get("C4C AmountIncludingTax"), row.get("Ticket Detail Claim Amount Value")),
        axis=1,
    )
    base["Dealer Name"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Dealer Name"), row.get("Ticket Detail Dealer Name")),
        axis=1,
    )
    base["Ticket Type"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Ticket Type"), row.get("Ticket Detail Ticket Type")),
        axis=1,
    )
    base["Ticket Type Text"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Ticket Type Text"), row.get("Ticket Detail Ticket Type Text")),
        axis=1,
    )
    base["Ticket Status Group"] = base.apply(
        lambda row: first_nonempty_text(row.get("Ticket Detail Status Group")),
        axis=1,
    )
    base["Ticket Status Text"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Ticket Status Text"), row.get("Ticket Detail Status Text")),
        axis=1,
    )
    base["Ticket Status Code"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Ticket Status"), row.get("Ticket Detail Status Code")),
        axis=1,
    )
    base["Created On"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Created On"), row.get("Ticket Detail Created On")),
        axis=1,
    )
    base["Claim Approved On"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Claim Approved On"), row.get("Ticket Detail Claim Approved On")),
        axis=1,
    )
    base["Resolved On"] = base.apply(
        lambda row: first_nonempty_text(row.get("C4C Resolved On"), row.get("Ticket Detail Resolved On")),
        axis=1,
    )
    base["Claim Scope"] = base.apply(
        lambda row: first_nonempty_text(row.get("Ticket Detail Claim Scope")),
        axis=1,
    )
    base["C4C Chassis Number"] = base["C4C Chassis Number"].fillna("").astype(str).str.strip()
    base["C4C PO Number Count"] = base["C4C ERP Purchase Order"].apply(lambda value: len(split_po_numbers(value)))

    preferred_cols = [
        "Ticket Number",
        "Claim Scope",
        "Ticket Status Group",
        "Ticket Status Text",
        "Ticket Status Code",
        "Dealer Name",
        "Ticket Type",
        "Ticket Type Text",
        "Created On",
        "Claim Approved On",
        "Resolved On",
        "C4C ERP Purchase Order",
        "C4C PO Number Count",
        "C4C AmountIncludingTax",
        "C4C Sales Order",
        "C4C Chassis Number",
        "Ticket Detail PO Net Value",
    ]
    for col in preferred_cols:
        if col not in base.columns:
            base[col] = ""
    base = base.drop_duplicates(subset=["Ticket Number"], keep="first").reset_index(drop=True)
    return base


def fetch_sap_po_items_by_documents(
    conn,
    schema: str,
    client: str,
    po_numbers: list[str],
    exclude_deleted: bool,
) -> pd.DataFrame:
    if not po_numbers:
        return pd.DataFrame()

    ekko_cols = set(get_table_columns(conn, schema, "EKKO"))
    ekpo_cols = set(get_table_columns(conn, schema, "EKPO"))

    select_parts: list[str] = []
    add_select(select_parts, ekpo_cols, "p", "EBELN", "Purchasing Document")
    add_select(select_parts, ekpo_cols, "p", "EBELP", "Item")
    add_select(select_parts, ekko_cols, "h", "EKORG", "Purchasing Organization")
    add_select(select_parts, ekko_cols, "h", "EKGRP", "Purchasing Group")
    add_select(select_parts, ekpo_cols, "p", "WERKS", "Plant")
    add_select(select_parts, ekko_cols, "h", "BSART", "Document Type")
    add_select(select_parts, ekko_cols, "h", "BEDAT", "Document Date")
    add_select(select_parts, ekko_cols, "h", "AEDAT", "Changed On")
    add_select(select_parts, ekko_cols, "h", "ERNAM", "Created By")
    add_select(select_parts, ekko_cols, "h", "LIFNR", "Supplier")
    add_select(select_parts, ekpo_cols, "p", "MATNR", "Material")
    add_select(select_parts, ekpo_cols, "p", "TXZ01", "Short Text")
    add_select(select_parts, ekpo_cols, "p", "MATKL", "Material Group")
    add_select(select_parts, ekpo_cols, "p", "KNTTP", "Account Assignment Category")
    add_select(select_parts, ekpo_cols, "p", "PSTYP", "Item Category")
    add_select(select_parts, ekpo_cols, "p", "MENGE", "Order Quantity")
    add_select(select_parts, ekpo_cols, "p", "MEINS", "Order Unit")
    add_select(select_parts, ekpo_cols, "p", "NETPR", "Net Price")
    add_select(select_parts, ekpo_cols, "p", "PEINH", "Price Unit")
    add_select(select_parts, ekpo_cols, "p", "NETWR", "Net Order Value")
    add_select(select_parts, ekpo_cols, "p", "BRTWR", "Gross Order Value")
    add_select(select_parts, ekko_cols, "h", "WAERS", "Currency")
    add_select(select_parts, ekpo_cols, "p", "MWSKZ", "Tax Code")
    add_select(select_parts, ekko_cols, "h", "LOEKZ", "Header Deletion Indicator")
    add_select(select_parts, ekpo_cols, "p", "LOEKZ", "Item Deletion Indicator")
    add_select(select_parts, ekpo_cols, "p", "ELIKZ", "Delivery Completed")
    add_select(select_parts, ekpo_cols, "p", "EREKZ", "Final Invoice")
    add_select(select_parts, ekpo_cols, "p", "LGORT", "Storage Location")

    if "Short Text" not in [part.split(" AS ")[-1].replace('"', "") for part in select_parts]:
        raise RuntimeError("EKPO.TXZ01 / Short Text is not available in the SAP query result.")

    chunks = [po_numbers[i:i + 500] for i in range(0, len(po_numbers), 500)]
    frames: list[pd.DataFrame] = []
    logger.info("Fetching SAP PO rows by C4C PO numbers: uniquePOs=%s chunks=%s", len(po_numbers), len(chunks))

    for idx, chunk in enumerate(chunks, start=1):
        po_sql = ", ".join(f"'{sql_quote(po)}'" for po in chunk)
        where_parts = [
            f'h."MANDT" = \'{sql_quote(client)}\'',
            f'p."EBELN" IN ({po_sql})',
        ]
        if exclude_deleted:
            if real_col(ekko_cols, "LOEKZ"):
                where_parts.append("COALESCE(h.\"LOEKZ\", '') = ''")
            if real_col(ekpo_cols, "LOEKZ"):
                where_parts.append("COALESCE(p.\"LOEKZ\", '') = ''")

        sql = f"""
        SELECT
            {", ".join(select_parts)}
        FROM {full_table(schema, "EKPO")} p
        INNER JOIN {full_table(schema, "EKKO")} h
            ON h."MANDT" = p."MANDT"
           AND h."EBELN" = p."EBELN"
        WHERE {" AND ".join(where_parts)}
        ORDER BY p."EBELN", p."EBELP"
        """
        frame = read_sql(conn, sql)
        logger.info("SAP PO chunk %s/%s fetched rows=%s", idx, len(chunks), len(frame))
        frames.append(frame)

    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    logger.info("SAP PO rows fetched by documents total=%s", len(df))
    return df


def add_short_text_columns(po_df: pd.DataFrame) -> pd.DataFrame:
    if po_df.empty:
        return po_df.copy()
    work = po_df.copy()
    ticket_extracted = work["Short Text"].apply(extract_ticket_number)

    work["SAP Short Text Ticket Number"] = ticket_extracted.apply(lambda item: item[0])
    work["SAP Short Text Ticket Note"] = ticket_extracted.apply(lambda item: item[1])

    for col in ["Net Price", "Net Order Value", "Gross Order Value", "Order Quantity"]:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)
    return work


def unique_join(values: pd.Series, limit: int = 20) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean(value)
        if text and text not in seen:
            out.append(text)
            seen.add(text)
        if len(out) >= limit:
            break
    return ", ".join(out)


def build_ticket_po_link_rows(base_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for record in base_df.to_dict("records"):
        po_numbers = split_po_numbers(record.get("C4C ERP Purchase Order"))
        if not po_numbers:
            rows.append({**record, "C4C ERP Purchase Order Single": ""})
            continue
        for po_number in po_numbers:
            rows.append({**record, "C4C ERP Purchase Order Single": po_number})
    return pd.DataFrame(rows)


def enrich_line_items(link_df: pd.DataFrame, po_df: pd.DataFrame, amount_column: str) -> pd.DataFrame:
    if link_df.empty:
        return pd.DataFrame()

    if po_df.empty:
        out = link_df.copy()
        out["SAP Short Text Resolved Ticket Number"] = ""
        out["SAP Short Text Match Source"] = ""
        out["SAP Short Text Match Status"] = ""
        return out

    work = po_df.copy()
    work["Purchasing Document"] = work["Purchasing Document"].fillna("").astype(str).str.strip()
    work["SAP_PO_Line_Amount"] = pd.to_numeric(work[amount_column], errors="coerce").fillna(0.0)

    link_work = link_df.copy()
    link_work["C4C ERP Purchase Order Single"] = link_work["C4C ERP Purchase Order Single"].fillna("").astype(str).str.strip()

    merged = link_work.merge(
        work,
        how="left",
        left_on="C4C ERP Purchase Order Single",
        right_on="Purchasing Document",
    )

    def resolve_row(row: pd.Series) -> tuple[str, str, str]:
        if not text_or_blank(row.get("Purchasing Document")):
            return "", "", "no_sap_po_line"

        c4c_ticket = text_or_blank(row.get("Ticket Number"))
        short_text_ticket = text_or_blank(row.get("SAP Short Text Ticket Number"))

        if short_text_ticket:
            resolved_ticket = short_text_ticket
            source = "short_text_ticket_number"
            status = "ticket_matches_c4c" if short_text_ticket == c4c_ticket else "ticket_differs_from_c4c"
            return resolved_ticket, source, status

        return "", "", "no_short_text_ticket_number"

    resolved = merged.apply(resolve_row, axis=1)
    merged["SAP Short Text Resolved Ticket Number"] = resolved.apply(lambda item: item[0])
    merged["SAP Short Text Match Source"] = resolved.apply(lambda item: item[1])
    merged["SAP Short Text Match Status"] = resolved.apply(lambda item: item[2])
    return merged


def aggregate_ticket_summary(base_df: pd.DataFrame, line_items_df: pd.DataFrame, amount_column: str) -> pd.DataFrame:
    summary = base_df.copy()
    if line_items_df.empty or "Purchasing Document" not in line_items_df.columns:
        summary["SAP Purchasing Documents"] = ""
        summary["SAP_PO_Item_Count"] = 0
        summary["SAP PO Amount"] = 0.0
        summary["SAP Net Order Value"] = 0.0
        summary["SAP Net Price Sum"] = 0.0
        summary["SAP Currencies"] = ""
        summary["SAP Short Text"] = ""
        summary["SAP Short Text Resolved Ticket Number"] = ""
        summary["SAP Short Text Extracted Ticket Number"] = ""
        summary["SAP Short Text Match Source"] = ""
        summary["SAP Short Text Match Status"] = ""
    else:
        matched_lines = line_items_df[line_items_df["Purchasing Document"].fillna("").astype(str).str.strip() != ""].copy()
        grouped = (
            matched_lines.groupby("Ticket Number", dropna=False)
            .agg(
                **{
                    "SAP Purchasing Documents": ("Purchasing Document", unique_join),
                    "SAP_PO_Item_Count": ("Purchasing Document", "size"),
                    "SAP PO Amount": ("SAP_PO_Line_Amount", "sum"),
                    "SAP Net Order Value": ("Net Order Value", "sum") if "Net Order Value" in matched_lines.columns else ("SAP_PO_Line_Amount", "sum"),
                    "SAP Net Price Sum": ("Net Price", "sum") if "Net Price" in matched_lines.columns else ("SAP_PO_Line_Amount", "sum"),
                    "SAP Currencies": ("Currency", unique_join) if "Currency" in matched_lines.columns else ("Purchasing Document", unique_join),
                    "SAP Short Text": ("Short Text", unique_join) if "Short Text" in matched_lines.columns else ("Purchasing Document", unique_join),
                    "SAP Short Text Resolved Ticket Number": ("SAP Short Text Resolved Ticket Number", unique_join),
                    "SAP Short Text Extracted Ticket Number": ("SAP Short Text Ticket Number", unique_join),
                    "SAP Short Text Match Source": ("SAP Short Text Match Source", unique_join),
                    "SAP Short Text Match Status": ("SAP Short Text Match Status", unique_join),
                }
            )
            .reset_index()
        )
        for col in ["SAP PO Amount", "SAP Net Order Value", "SAP Net Price Sum"]:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce").fillna(0.0).round(2)
        summary = summary.merge(grouped, how="left", on="Ticket Number")

    for col in [
        "SAP_PO_Item_Count",
        "SAP PO Amount",
        "SAP Net Order Value",
        "SAP Net Price Sum",
    ]:
        summary[col] = pd.to_numeric(summary.get(col), errors="coerce").fillna(0)
    summary["SAP_PO_Item_Count"] = summary["SAP_PO_Item_Count"].astype(int)
    for col in [
        "SAP Purchasing Documents",
        "SAP Currencies",
        "SAP Short Text",
        "SAP Short Text Resolved Ticket Number",
        "SAP Short Text Extracted Ticket Number",
        "SAP Short Text Match Source",
        "SAP Short Text Match Status",
    ]:
        summary[col] = summary.get(col, "").fillna("").astype(str)

    summary["Delta SAP Amount - C4C Amount"] = (
        pd.to_numeric(summary["SAP PO Amount"], errors="coerce").fillna(0.0)
        - pd.to_numeric(summary["C4C AmountIncludingTax"], errors="coerce").fillna(0.0)
    ).round(2)

    def match_status(row: pd.Series) -> str:
        if int(row.get("C4C PO Number Count") or 0) <= 0:
            return "No C4C PO"
        if int(row.get("SAP_PO_Item_Count") or 0) > 0:
            return "Matched by C4C PO"
        return "C4C PO not found in SAP"

    summary["SAP Match Status"] = summary.apply(match_status, axis=1)

    ordered_columns = [
        "Ticket Number",
        "SAP Short Text Resolved Ticket Number",
        "SAP Short Text Extracted Ticket Number",
        "SAP Short Text",
        "C4C ERP Purchase Order",
        "SAP Purchasing Documents",
        "C4C AmountIncludingTax",
        "SAP PO Amount",
        "Delta SAP Amount - C4C Amount",
        "Ticket Detail PO Net Value",
        "SAP Net Order Value",
        "SAP Net Price Sum",
        "SAP_PO_Item_Count",
        "SAP Match Status",
        "SAP Short Text Match Source",
        "SAP Short Text Match Status",
        "C4C Chassis Number",
        "Claim Scope",
        "Ticket Status Group",
        "Ticket Status Text",
        "Ticket Status Code",
        "Dealer Name",
        "Ticket Type",
        "Ticket Type Text",
        "Created On",
        "Claim Approved On",
        "Resolved On",
        "C4C Sales Order",
        "C4C PO Number Count",
    ]
    for col in ordered_columns:
        if col not in summary.columns:
            summary[col] = ""
    summary = summary[ordered_columns]
    return summary.sort_values(["SAP Match Status", "Ticket Number"]).reset_index(drop=True)


def build_output_sheets(summary_df: pd.DataFrame, line_items_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    no_c4c_po = summary_df[summary_df["SAP Match Status"] == "No C4C PO"].copy() if not summary_df.empty else pd.DataFrame()
    c4c_po_not_in_sap = summary_df[summary_df["SAP Match Status"] == "C4C PO not found in SAP"].copy() if not summary_df.empty else pd.DataFrame()
    matched = summary_df[summary_df["SAP Match Status"] == "Matched by C4C PO"].copy() if not summary_df.empty else pd.DataFrame()

    totals = pd.DataFrame(
        [
            {"Metric": "Tickets in summary", "Value": len(summary_df)},
            {"Metric": "Tickets matched by C4C PO", "Value": len(matched)},
            {"Metric": "Tickets with C4C PO not found in SAP", "Value": len(c4c_po_not_in_sap)},
            {"Metric": "Tickets with no C4C PO", "Value": len(no_c4c_po)},
            {"Metric": "SAP PO line rows", "Value": len(line_items_df)},
            {
                "Metric": "Total C4C AmountIncludingTax",
                "Value": round(float(pd.to_numeric(summary_df.get("C4C AmountIncludingTax"), errors="coerce").fillna(0.0).sum()), 2),
            },
            {
                "Metric": "Total SAP PO Amount",
                "Value": round(float(pd.to_numeric(summary_df.get("SAP PO Amount"), errors="coerce").fillna(0.0).sum()), 2),
            },
        ]
    )

    line_cols = [
        "Ticket Number",
        "SAP Short Text Resolved Ticket Number",
        "SAP Short Text Ticket Number",
        "Short Text",
        "C4C ERP Purchase Order",
        "C4C ERP Purchase Order Single",
        "C4C AmountIncludingTax",
        "Purchasing Document",
        "Item",
        "SAP_PO_Line_Amount",
        "Net Order Value",
        "Net Price",
        "Gross Order Value",
        "Currency",
        "SAP Short Text Match Source",
        "SAP Short Text Match Status",
        "C4C Chassis Number",
        "Document Date",
        "Changed On",
        "Created On",
        "Dealer Name",
        "Ticket Status Text",
    ]
    if not line_items_df.empty:
        for col in line_cols:
            if col not in line_items_df.columns:
                line_items_df[col] = ""
        line_items_export = line_items_df[line_cols].copy()
    else:
        line_items_export = pd.DataFrame(columns=line_cols)

    return {
        "Run_Summary": totals,
        "Ticket_PO_Summary": summary_df,
        "SAP_PO_Line_Items": line_items_export,
        "Matched_By_C4C_PO": matched,
        "C4C_PO_Not_In_SAP": c4c_po_not_in_sap,
        "No_C4C_PO": no_c4c_po,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare repairer payment by C4C ticket -> C4C PO -> SAP PO amount.")
    parser.add_argument("--output", default=os.getenv("OUTPUT_FILE", DEFAULT_OUTPUT))
    parser.add_argument("--log-file", default=os.getenv("LOG_FILE", "repairer_payment_from_c4c_po.log"))
    parser.add_argument("--ticket-detail-file", default=os.getenv("TICKET_DETAIL_FILE", str(DEFAULT_TICKET_DETAIL_FILE)))
    parser.add_argument("--ticket-sheet", default=os.getenv("TICKET_DETAIL_SHEET", ""))
    parser.add_argument("--firebase-db-url", default=os.getenv("FIREBASE_DB_URL", DEFAULT_DB_URL))
    parser.add_argument("--firebase-sa-path", default=os.getenv("FIREBASE_SA_PATH", str(Path.cwd() / "firebase-service-account.json")))
    parser.add_argument("--source-root", default=os.getenv("SOURCE_ROOT", DEFAULT_SOURCE_ROOT))
    parser.add_argument("--sap-hana-dsn", default=os.getenv("SAP_HANA_DSN", ""))
    parser.add_argument("--sap-schema", default=os.getenv("SAP_SCHEMA", DEFAULT_SCHEMA))
    parser.add_argument("--sap-client", default=os.getenv("SAP_CLIENT", DEFAULT_CLIENT))
    parser.add_argument(
        "--amount-column",
        default=os.getenv("SAP_PO_AMOUNT_COLUMN", DEFAULT_AMOUNT_COLUMN),
        help="SAP PO column used as payment amount. Default: Net Order Value.",
    )
    parser.add_argument("--exclude-deleted-po", action="store_true", default=os.getenv("EXCLUDE_DELETED_PO", "").lower() in {"1", "true", "yes", "y"})
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    setup_logging(args.log_file)

    logger.info("Started repairer payment compare by C4C PO.")
    logger.info("Output file: %s", args.output)
    logger.info("Ticket detail file: %s", args.ticket_detail_file or "(not provided)")
    logger.info("SAP amount column: %s", args.amount_column)

    init_firebase(args.firebase_db_url, args.firebase_sa_path)
    ticket_detail_df = load_ticket_detail_rows(args.ticket_detail_file, normalize_ticket_detail_sheet_name(args.ticket_sheet))
    c4c_lookup_df = load_c4c_ticket_lookup(args.source_root)
    base_df = build_base_ticket_rows(ticket_detail_df, c4c_lookup_df)
    if base_df.empty:
        raise SystemExit("No tickets found from Ticket Detail input or Firebase lookup.")

    po_numbers = sorted(
        {
            po
            for value in base_df["C4C ERP Purchase Order"].tolist()
            for po in split_po_numbers(value)
        }
    )
    logger.info("Unique C4C PO numbers to query: %s", len(po_numbers))

    hana_dsn = resolve_hana_dsn(args.sap_hana_dsn)
    with connect_hana(hana_dsn) as conn:
        po_df = fetch_sap_po_items_by_documents(
            conn,
            schema=args.sap_schema,
            client=args.sap_client,
            po_numbers=po_numbers,
            exclude_deleted=bool(args.exclude_deleted_po),
        )

    po_df = add_short_text_columns(po_df)
    link_df = build_ticket_po_link_rows(base_df)
    line_items_df = enrich_line_items(link_df, po_df, args.amount_column)
    summary_df = aggregate_ticket_summary(base_df, line_items_df, args.amount_column)
    sheets = build_output_sheets(summary_df, line_items_df)
    write_excel(args.output, sheets)

    for _, row in sheets["Run_Summary"].iterrows():
        logger.info("%s: %s", row["Metric"], row["Value"])
    logger.info("Done.")


if __name__ == "__main__":
    main()
